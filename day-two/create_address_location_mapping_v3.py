#!/usr/bin/env python3
"""
Generate ES Address__c to BBF Location__c field mapping Excel workbook
with field matching and picklist value mapping (Version 3 - Fixed Picklist Matching)
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from difflib import SequenceMatcher
import re

# Color definitions (EXACT hex codes per spec)
HEADER_FILL = PatternFill(start_color="FF366092", end_color="FF366092", fill_type="solid")
HIGH_CONFIDENCE_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
MEDIUM_CONFIDENCE_FILL = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")
LOW_CONFIDENCE_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
NONE_CONFIDENCE_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

HEADER_FONT = Font(bold=True, color="FFFFFF")

# System fields to exclude
SYSTEM_FIELDS = {
    'Id', 'IsDeleted', 'MasterRecordId', 'CreatedDate', 'CreatedById',
    'LastModifiedDate', 'LastModifiedById', 'SystemModstamp',
    'LastActivityDate', 'LastViewedDate', 'LastReferencedDate',
    'JigsawContactId', 'Jigsaw', 'PhotoUrl', 'CleanStatus'
}

# Day 1 migration fields to exclude for Location__c
DAY_1_FIELDS_LOCATION = {
    'Name', 'OwnerId', 'ES_Legacy_ID__c', 'BBF_New_Id__c',
    'Address_Line_1__c', 'Address_Line_2__c', 'City__c',
    'State_Province__c', 'Postal_Code__c', 'Country__c'
}

# Manual picklist field mapping (BBF -> ES) for fields that don't match by name
PICKLIST_FIELD_OVERRIDES = {
    'Address_Validated_By__c': 'Verification_Used__c',  # BBF field -> ES field
}

def normalize_api_name(name):
    """Normalize API name for comparison"""
    if not name:
        return ""
    # Remove __c suffix and convert to lowercase
    normalized = name.replace('__c', '').lower()
    # Remove common prefixes
    normalized = re.sub(r'^(es_|bbf_|legacy_)', '', normalized)
    return normalized

def similarity_score(str1, str2):
    """Calculate similarity score between two strings"""
    if not str1 or not str2:
        return 0.0
    return SequenceMatcher(None, str1.lower(), str2.lower()).ratio()

def match_fields(bbf_field, es_fields_df):
    """
    Match a BBF field to ES fields with confidence level
    Returns: (es_field_api_name, es_field_label, es_data_type, confidence, needs_transformer, notes)
    """
    bbf_api = bbf_field['Field API Name']
    bbf_label = bbf_field['Field Label']
    bbf_type = bbf_field['Field Type']

    # Try exact API name match first
    exact_match = es_fields_df[es_fields_df['Field API Name'] == bbf_api]
    if not exact_match.empty:
        es_field = exact_match.iloc[0]
        needs_xform = 'Y' if es_field['Field Type'] != bbf_type else 'N'
        return (es_field['Field API Name'], es_field['Field Label'], es_field['Field Type'],
                'High', needs_xform, 'Exact API name match')

    # Try normalized API name similarity
    bbf_normalized = normalize_api_name(bbf_api)
    best_api_match = None
    best_api_score = 0.0

    for _, es_field in es_fields_df.iterrows():
        es_normalized = normalize_api_name(es_field['Field API Name'])
        score = similarity_score(bbf_normalized, es_normalized)
        if score > best_api_score and score >= 0.8:
            best_api_score = score
            best_api_match = es_field

    if best_api_match is not None:
        needs_xform = 'Y' if best_api_match['Field Type'] != bbf_type else 'N'
        return (best_api_match['Field API Name'], best_api_match['Field Label'],
                best_api_match['Field Type'], 'High', needs_xform,
                f'API name similarity: {best_api_score:.2f}')

    # Try label match
    label_match = es_fields_df[es_fields_df['Field Label'].str.lower() == bbf_label.lower()]
    if not label_match.empty:
        es_field = label_match.iloc[0]
        needs_xform = 'Y' if es_field['Field Type'] != bbf_type else 'N'
        return (es_field['Field API Name'], es_field['Field Label'], es_field['Field Type'],
                'Medium', needs_xform, 'Exact label match')

    # Try semantic/fuzzy label match
    best_label_match = None
    best_label_score = 0.0

    for _, es_field in es_fields_df.iterrows():
        score = similarity_score(bbf_label, es_field['Field Label'])
        if score > best_label_score and score >= 0.6:
            best_label_score = score
            best_label_match = es_field

    if best_label_match is not None:
        needs_xform = 'Y' if best_label_match['Field Type'] != bbf_type else 'N'
        return (best_label_match['Field API Name'], best_label_match['Field Label'],
                best_label_match['Field Type'], 'Medium', needs_xform,
                f'Label similarity: {best_label_score:.2f}')

    # No match found
    return ('', '', '', 'None', 'N', 'No obvious ES source field identified')

def load_metadata():
    """Load ES and BBF metadata from CSV files"""
    es_fields = pd.read_csv('day-two/exports/es_Address__c_fields_with_picklists.csv')
    bbf_fields = pd.read_csv('day-two/exports/bbf_Location__c_fields_with_picklists.csv')
    es_picklists = pd.read_csv('day-two/exports/es_Address__c_picklist_values.csv')
    bbf_picklists = pd.read_csv('day-two/exports/bbf_Location__c_picklist_values.csv')

    return es_fields, bbf_fields, es_picklists, bbf_picklists

def filter_fields(bbf_fields):
    """Filter out system and Day 1 fields"""
    excluded = SYSTEM_FIELDS | DAY_1_FIELDS_LOCATION
    return bbf_fields[~bbf_fields['Field API Name'].isin(excluded)]

def create_field_mapping(es_fields, bbf_fields):
    """Create field mapping sheet data"""
    bbf_filtered = filter_fields(bbf_fields)

    mapping_data = []

    for _, bbf_field in bbf_filtered.iterrows():
        es_api, es_label, es_type, confidence, transformer, notes = match_fields(bbf_field, es_fields)

        # Check if field lengths differ and ES field is longer
        if es_api and es_type == 'string' and bbf_field['Field Type'] == 'string':
            es_length = es_fields[es_fields['Field API Name'] == es_api]['Length'].iloc[0]
            bbf_length = bbf_field['Length']
            if pd.notna(es_length) and pd.notna(bbf_length) and es_length > bbf_length:
                transformer = 'Y'
                notes += f' (ES length {int(es_length)} > BBF length {int(bbf_length)})'

        # Check for picklist type mismatches
        if es_type != bbf_field['Field Type'] and es_api:
            transformer = 'Y'
            notes += f' (Type mismatch: ES {es_type} -> BBF {bbf_field["Field Type"]})'

        mapping_data.append({
            'BBF_Field_API_Name': bbf_field['Field API Name'],
            'BBF_Field_Label': bbf_field['Field Label'],
            'BBF_Data_Type': bbf_field['Field Type'],
            'BBF_Is_Required': 'No' if bbf_field['Is Nillable'] else 'Yes',
            'ES_Field_API_Name': es_api,
            'ES_Field_Label': es_label,
            'ES_Data_Type': es_type,
            'Match_Confidence': confidence,
            'Transformer_Needed': transformer,
            'Notes': notes
        })

    return pd.DataFrame(mapping_data)

def create_picklist_mapping(es_picklists, bbf_picklists, field_mapping_df, es_fields):
    """Create picklist mapping sheet data - with manual overrides for semantic matches"""
    picklist_data = []

    # Get all BBF picklist fields
    bbf_picklist_fields = bbf_picklists['Field API Name'].unique()

    print(f"\nBBF picklist fields: {bbf_picklist_fields}")
    print(f"ES picklist fields: {es_picklists['Field API Name'].unique()}")

    for bbf_field in bbf_picklist_fields:
        print(f"\n=== Processing BBF field: {bbf_field} ===")

        # Check for manual override first
        if bbf_field in PICKLIST_FIELD_OVERRIDES:
            es_field = PICKLIST_FIELD_OVERRIDES[bbf_field]
            print(f"  Using manual override: {bbf_field} -> {es_field}")
        else:
            # Find matching ES field from field mapping
            field_match = field_mapping_df[field_mapping_df['BBF_Field_API_Name'] == bbf_field]

            if not field_match.empty:
                es_field = field_match.iloc[0]['ES_Field_API_Name']
                es_type = field_match.iloc[0]['ES_Data_Type']
                print(f"  Found ES match from field mapping: {es_field} (type: {es_type})")

                # Only use if it's actually a picklist in ES
                if es_type != 'picklist':
                    print(f"  WARNING: ES field is {es_type}, not picklist - skipping")
                    es_field = None
            else:
                print(f"  No field mapping found")
                es_field = None

        if not es_field or es_field not in es_picklists['Field API Name'].values:
            print(f"  No valid ES picklist field - showing BBF values only")
            # Show BBF field alone with all valid values
            bbf_values = bbf_picklists[bbf_picklists['Field API Name'] == bbf_field]['Picklist Value'].tolist()
            picklist_data.append({
                'ES_Field': '',
                'ES_Picklist_Value': 'N/A - No ES source field',
                'BBF_Field': bbf_field,
                'Suggested_Mapping': ' | '.join(bbf_values),
                'Notes': 'No Match - Select from list',
                'BBF_Final_Value': ''
            })
            continue

        # Get all ES values for this field
        es_values = es_picklists[es_picklists['Field API Name'] == es_field]['Picklist Value'].tolist()
        # Get all BBF values for this field
        bbf_values = bbf_picklists[bbf_picklists['Field API Name'] == bbf_field]['Picklist Value'].tolist()

        print(f"  Mapping {len(es_values)} ES values to {len(bbf_values)} BBF values")

        # Map each ES value to BBF
        for es_value in es_values:
            # Try exact match
            if es_value in bbf_values:
                picklist_data.append({
                    'ES_Field': es_field,
                    'ES_Picklist_Value': es_value,
                    'BBF_Field': bbf_field,
                    'Suggested_Mapping': es_value,
                    'Notes': 'Exact Match',
                    'BBF_Final_Value': ''
                })
            else:
                # Try fuzzy match
                best_match = None
                best_score = 0.0
                for bbf_value in bbf_values:
                    score = similarity_score(es_value, bbf_value)
                    if score > best_score:
                        best_score = score
                        best_match = bbf_value

                if best_score >= 0.7:
                    picklist_data.append({
                        'ES_Field': es_field,
                        'ES_Picklist_Value': es_value,
                        'BBF_Field': bbf_field,
                        'Suggested_Mapping': best_match,
                        'Notes': 'Close Match',
                        'BBF_Final_Value': ''
                    })
                else:
                    # No match - show all BBF values
                    all_bbf_values = ' | '.join(bbf_values)
                    picklist_data.append({
                        'ES_Field': es_field,
                        'ES_Picklist_Value': es_value,
                        'BBF_Field': bbf_field,
                        'Suggested_Mapping': all_bbf_values,
                        'Notes': 'No Match - Select from list',
                        'BBF_Final_Value': ''
                    })

    return pd.DataFrame(picklist_data)

def apply_formatting(ws, df, sheet_type='field'):
    """Apply formatting to worksheet"""
    # Header row formatting
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Set column widths
    if sheet_type == 'field':
        ws.column_dimensions['A'].width = 25  # BBF_Field_API_Name
        ws.column_dimensions['B'].width = 25  # BBF_Field_Label
        ws.column_dimensions['C'].width = 15  # BBF_Data_Type
        ws.column_dimensions['D'].width = 15  # BBF_Is_Required
        ws.column_dimensions['E'].width = 25  # ES_Field_API_Name
        ws.column_dimensions['F'].width = 25  # ES_Field_Label
        ws.column_dimensions['G'].width = 15  # ES_Data_Type
        ws.column_dimensions['H'].width = 15  # Match_Confidence
        ws.column_dimensions['I'].width = 15  # Transformer_Needed
        ws.column_dimensions['J'].width = 50  # Notes

        # Apply confidence-based row colors
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(df)+1), start=2):
            confidence = df.iloc[idx-2]['Match_Confidence']

            if confidence == 'High':
                fill = HIGH_CONFIDENCE_FILL
            elif confidence == 'Medium':
                fill = MEDIUM_CONFIDENCE_FILL
            elif confidence == 'Low':
                fill = LOW_CONFIDENCE_FILL
            else:  # None
                fill = NONE_CONFIDENCE_FILL

            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    else:  # picklist sheet
        ws.column_dimensions['A'].width = 25  # ES_Field
        ws.column_dimensions['B'].width = 30  # ES_Picklist_Value
        ws.column_dimensions['C'].width = 25  # BBF_Field
        ws.column_dimensions['D'].width = 50  # Suggested_Mapping
        ws.column_dimensions['E'].width = 30  # Notes
        ws.column_dimensions['F'].width = 30  # BBF_Final_Value

        # Apply match-based row colors
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(df)+1), start=2):
            notes = df.iloc[idx-2]['Notes']

            if notes == 'Exact Match':
                fill = HIGH_CONFIDENCE_FILL
            elif notes == 'Close Match':
                fill = MEDIUM_CONFIDENCE_FILL
            else:  # No Match
                fill = LOW_CONFIDENCE_FILL

            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(vertical='top', wrap_text=True)

def main():
    print("Loading metadata from CSV exports...")
    es_fields, bbf_fields, es_picklists, bbf_picklists = load_metadata()

    print(f"ES Address__c fields: {len(es_fields)}")
    print(f"BBF Location__c fields: {len(bbf_fields)}")

    print("\nCreating field mapping...")
    field_mapping_df = create_field_mapping(es_fields, bbf_fields)

    print(f"\nBBF fields after filtering: {len(field_mapping_df)}")
    print(f"High confidence matches: {len(field_mapping_df[field_mapping_df['Match_Confidence'] == 'High'])}")
    print(f"Medium confidence matches: {len(field_mapping_df[field_mapping_df['Match_Confidence'] == 'Medium'])}")
    print(f"Low confidence matches: {len(field_mapping_df[field_mapping_df['Match_Confidence'] == 'Low'])}")
    print(f"No match: {len(field_mapping_df[field_mapping_df['Match_Confidence'] == 'None'])}")
    print(f"Transformers needed: {len(field_mapping_df[field_mapping_df['Transformer_Needed'] == 'Y'])}")

    print("\n" + "="*60)
    print("Creating picklist mapping...")
    print("="*60)
    picklist_mapping_df = create_picklist_mapping(es_picklists, bbf_picklists, field_mapping_df, es_fields)

    if not picklist_mapping_df.empty:
        print(f"\n=== PICKLIST MAPPING SUMMARY ===")
        print(f"Total picklist rows: {len(picklist_mapping_df)}")
        print(f"Exact matches: {len(picklist_mapping_df[picklist_mapping_df['Notes'] == 'Exact Match'])}")
        print(f"Close matches: {len(picklist_mapping_df[picklist_mapping_df['Notes'] == 'Close Match'])}")
        print(f"No match: {len(picklist_mapping_df[picklist_mapping_df['Notes'] == 'No Match - Select from list'])}")
    else:
        print("\nNo picklist mappings found")

    print("\nCreating Excel workbook...")
    wb = Workbook()

    # Create Field_Mapping sheet
    ws1 = wb.active
    ws1.title = "Field_Mapping"

    for r in dataframe_to_rows(field_mapping_df, index=False, header=True):
        ws1.append(r)

    apply_formatting(ws1, field_mapping_df, sheet_type='field')

    # Create Picklist_Mapping sheet
    ws2 = wb.create_sheet("Picklist_Mapping")

    if not picklist_mapping_df.empty:
        for r in dataframe_to_rows(picklist_mapping_df, index=False, header=True):
            ws2.append(r)
        apply_formatting(ws2, picklist_mapping_df, sheet_type='picklist')
    else:
        # Add headers even if no data
        ws2.append(['ES_Field', 'ES_Picklist_Value', 'BBF_Field', 'Suggested_Mapping', 'Notes', 'BBF_Final_Value'])
        for cell in ws2[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

    # Save workbook
    output_path = 'day-two/mappings/ES_Address__c_to_BBF_Location__c_mapping.xlsx'
    wb.save(output_path)
    print(f"\nMapping saved to: {output_path}")

    return field_mapping_df, picklist_mapping_df

if __name__ == '__main__':
    field_df, picklist_df = main()
