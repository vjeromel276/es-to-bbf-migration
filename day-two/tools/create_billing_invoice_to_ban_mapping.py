#!/usr/bin/env python3
"""
Semantic Field Mapping Generator for ES Billing_Invoice__c to BBF BAN__c

This script performs AI-powered semantic matching to map fields between
EverStream Billing_Invoice__c and BlueBird Fiber BAN__c objects.

Uses domain knowledge of telecom billing accounts to intelligently match fields.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
import re

# Configuration
ES_FIELDS_FILE = "day-two/exports/es_Billing_Invoice__c_fields_with_picklists.csv"
BBF_FIELDS_FILE = "day-two/exports/bbf_BAN__c_fields_with_picklists.csv"
ES_PICKLIST_FILE = "day-two/exports/es_Billing_Invoice__c_picklist_values.csv"
BBF_PICKLIST_FILE = "day-two/exports/bbf_BAN__c_picklist_values.csv"
OUTPUT_FILE = "day-two/mappings/ES_Billing_Invoice__c_to_BBF_BAN__c_mapping.xlsx"

# System fields to exclude
SYSTEM_FIELDS = {
    'Id', 'IsDeleted', 'MasterRecordId', 'CreatedDate', 'CreatedById',
    'LastModifiedDate', 'LastModifiedById', 'SystemModstamp',
    'LastActivityDate', 'LastViewedDate', 'LastReferencedDate',
    'JigsawContactId', 'Jigsaw', 'PhotoUrl', 'CleanStatus'
}

# Day 1 migration fields for BAN__c (already populated)
DAY1_FIELDS = {
    'Name', 'Account__c', 'OwnerId', 'ES_Legacy_ID__c', 'BBF_New_Id__c'
}

# Color scheme (EXACT hex codes)
COLORS = {
    'header': 'FF366092',  # Dark blue
    'high': 'FFC6EFCE',    # Light green
    'medium': 'FFFFEB9C',  # Light yellow
    'low': 'FFFFC7CE',     # Light red/pink
    'none': 'FFFFC7CE'     # Light red/pink
}

def normalize_text(text):
    """Normalize text for comparison."""
    if pd.isna(text):
        return ""
    text = str(text).lower()
    text = re.sub(r'[_\s]+', '', text)
    return text

def find_best_es_match(bbf_field, es_df):
    """
    Find the best matching ES field for a given BBF field.

    Returns: (confidence, es_field_dict, transformer_needed, notes)
    """
    bbf_api = bbf_field['Field API Name']
    bbf_label = bbf_field['Field Label']
    bbf_type = bbf_field['Field Type']
    bbf_norm_api = normalize_text(bbf_api)
    bbf_norm_label = normalize_text(bbf_label)

    # Check if it's a calculated/formula field
    if bbf_field['Is Calculated'] == 'True':
        return ('None', None, 'N', f'BBF calculated field - auto-populated by system')

    # Semantic matching rules
    for _, es_row in es_df.iterrows():
        es_api = es_row['Field API Name']
        es_label = es_row['Field Label']
        es_type = es_row['Field Type']
        es_norm_api = normalize_text(es_api)
        es_norm_label = normalize_text(es_label)

        # Exact API name match
        if es_api == bbf_api:
            transformer = 'N' if es_type == bbf_type else 'Y'
            notes = f"Exact API name match. {'Type mismatch requires transformer.' if transformer == 'Y' else ''}"
            return ('High', es_row, transformer, notes)

        # Billing City
        if es_api == 'Billing_City__c' and bbf_api == 'Billing_City__c':
            return ('High', es_row, 'N', 'Billing city - exact semantic match')

        # Billing State
        if es_api == 'Billing_State__c' and bbf_api == 'Billing_State__c':
            transformer = 'Y' if es_type == 'picklist' and bbf_type == 'string' else 'N'
            notes = 'Billing state - semantic match. ES picklist to BBF string conversion needed.' if transformer == 'Y' else 'Billing state - exact match'
            return ('High', es_row, transformer, notes)

        # Billing Address 1 -> Billing Street
        if es_api == 'Billing_Address_1__c' and bbf_api == 'Billing_Street__c':
            return ('High', es_row, 'N', 'Billing street address line 1 - semantic match')

        # Billing ZIP -> Billing Postal Code
        if es_api == 'Billing_ZIP__c' and bbf_api == 'Billing_PostalCode__c':
            return ('High', es_row, 'N', 'Billing postal code - semantic match')

        # Payment Terms
        if es_api == 'Payment_Terms__c' and bbf_api == 'Payment_Terms__c':
            return ('High', es_row, 'Y', 'Payment terms - requires picklist value mapping (NET30,NET45,NET60 -> NET 30)')

        # Invoice Delivery Preference -> Invoice Type
        if es_api == 'Invoice_Delivery_Preference__c' and bbf_api == 'invType__c':
            return ('Medium', es_row, 'Y', 'Invoice delivery method - requires value transformation (Paper/Email -> Print/Email/Both/Manual)')

        # Billing Cycle -> Billing Schedule Group
        if es_api == 'Invoice_cycle_cd__c' and bbf_api == 'Billing_Schedule_Group__c':
            return ('Medium', es_row, 'Y', 'Billing schedule - requires value transformation (Manual/Monthly -> Monthly/Annual/Semiannual)')

        # Account Name -> Billing Company Name
        if es_api == 'Account_Name__c' and bbf_api == 'Billing_Company_Name__c':
            return ('High', es_row, 'N', 'Company/Account name for billing - semantic match')

        # Billing Notes -> General Description
        if es_api == 'Billing_Notes__c' and bbf_api == 'General_Description__c':
            return ('Medium', es_row, 'N', 'Billing notes to general description - semantic match for billing-specific notes')

        # Description -> BAN Description
        if es_api == 'Description__c' and bbf_api == 'BAN_Description__c':
            return ('Medium', es_row, 'N', 'Description fields - semantic match for sales description')

    # BBF-specific fields with no ES equivalent
    if bbf_api in ['busUnit__c', 'Federal_Tax_ID__c', 'State_Tax_ID__c',
                    'Customer_Posting_Group__c', 'Contract_Type__c', 'BAN_Type__c',
                    'Industry__c', 'Billing_Location__c', 'MatchKey__c',
                    'Smart_Hands_Rate__c', 'Smart_hands_ad_hoc_after_hours_rate__c',
                    'BAN_Team_Manager__c', 'BAN_Team_Leader__c', 'BAN_Team_Person_1__c',
                    'BAN_Team_Person_2__c', 'BAN_Team_Person_3__c', 'BAN_Team_Person_4__c',
                    'BAN_Team_Person_5__c', 'BAN_Team_Person_6__c', 'BAN_Team_Person_7__c',
                    'BAN_Team_Person_8__c', 'BAN_Team_Manager_Status__c', 'BAN_Team_Manager_Name__c',
                    'BAN_Team_Leader_Name__c', 'BAN_Team_Person_1_Name__c', 'BAN_Team_Person_2_Name__c',
                    'BAN_Team_Person_3_Name__c', 'BAN_Team_Person_4_Name__c', 'BAN_Team_Person_5_Name__c',
                    'BAN_Team_Person_6_Name__c', 'BAN_Team_Person_7_Name__c', 'BAN_Team_Person_8_Name__c',
                    'BAN_Team_Name_String__c', 'Account_Owner_Manager__c', 'BAN_Team_ID_String__c',
                    'BAN_Team_Leader_Username__c', 'BAN_Team_String_Trigger__c', 'BAN_Type_COPS__c',
                    'Datacenter_Notes__c', 'IT_Team__c', 'Customer_Segment__c',
                    'Customer_Notice_Requirement__c', 'Sort_Billing_Document_By_Location__c',
                    'Outstanding_Balance__c', 'Bill_To__c', 'Active_Charges__c',
                    'Company_Class__c', 'Vertical__c', 'Full_Address__c',
                    'Num_Active_Services_With_Charges__c', 'Account_Owner__c',
                    'Account_Owner_Matches_BTM__c', 'Num_Inactive_Services__c',
                    'Num_Active_Services_Without_Charges__c', 'Num_Billing_Contacts__c',
                    'Num_Won_Opportunities__c', 'Blankspot__c', 'Blankspot2__c',
                    'Billing_invoices_for_this_BAN__c', 'SOLs_for_this_BAN__c',
                    'Qualified_SOLs_for_this_BAN__c', 'Not_Yet_Qualified_SOLs_for_this_BAN__c',
                    'Services_for_this_BAN__c', 'Active_Services_for_this_BAN__c',
                    'Inactive_Services_for_this_BAN__c', 'Canceled_SOLs_for_this_BAN__c',
                    'Text_Billings_For_This_BAN__c', 'Text_SOLs_For_This_BAN__c',
                    'Text_Services_For_This_BAN__c', 'All_Files_For_BAN__c']:

        # Provide specific guidance
        if 'Team' in bbf_api:
            return ('None', None, 'N', 'BBF BAN Team management field - no ES equivalent. Requires business rule for team assignment.')
        elif 'Tax' in bbf_label:
            return ('None', None, 'N', 'Tax ID field - no ES equivalent. May need to extract from Account or manual entry.')
        elif bbf_api == 'busUnit__c':
            return ('None', None, 'N', 'Business Unit - no ES equivalent. Requires business rule/default (e.g., "EVS" for EverStream accounts).')
        elif bbf_api == 'BAN_Type__c':
            return ('None', None, 'N', 'BAN Type - no ES equivalent. Likely "Customer" for all migrated BAN records.')
        elif bbf_api == 'Customer_Posting_Group__c':
            return ('None', None, 'N', 'Customer Posting Group - financial/ERP field. Requires business rule or Account classification.')
        elif bbf_api == 'Contract_Type__c':
            return ('None', None, 'N', 'Contract Type - no ES equivalent. May derive from Account or Order data.')
        elif bbf_api == 'Industry__c':
            return ('None', None, 'N', 'Industry classification - deprecated field. May copy from Account.Industry if needed.')
        elif bbf_api == 'Outstanding_Balance__c':
            return ('None', None, 'N', 'Outstanding Balance - financial field. Will be updated by billing system integration.')
        elif bbf_api == 'Bill_To__c':
            return ('None', None, 'N', 'Bill To attention line - no direct ES equivalent. May derive from AP_Contact__c name.')
        elif 'Num_' in bbf_api or 'Active_Charges' in bbf_api:
            return ('None', None, 'N', 'Calculated/rollup field - auto-populated from related Service records.')
        elif 'for_this_BAN' in bbf_api or 'Blankspot' in bbf_api or 'Text_' in bbf_api:
            return ('None', None, 'N', 'Formula/display field for UI - auto-calculated, no migration needed.')
        else:
            return ('None', None, 'N', f'BBF-specific field - no ES equivalent. Requires business rule or default value.')

    # No match found
    return ('None', None, 'N', f'No semantic match found for BBF field: {bbf_label}')

def create_field_mapping_sheet(wb, es_df, bbf_df):
    """Create the Field_Mapping sheet."""
    ws = wb.create_sheet("Field_Mapping", 0)

    # Headers
    headers = [
        'BBF_Field_API_Name', 'BBF_Field_Label', 'BBF_Data_Type', 'BBF_Is_Required',
        'ES_Field_API_Name', 'ES_Field_Label', 'ES_Data_Type',
        'Match_Confidence', 'Transformer_Needed', 'Notes',
        'ES_Final_Field', 'Include_in_Migration', 'Business_Notes'
    ]

    # Write headers with formatting
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Filter BBF fields (exclude system fields and Day 1 fields)
    bbf_filtered = bbf_df[
        (~bbf_df['Field API Name'].isin(SYSTEM_FIELDS)) &
        (~bbf_df['Field API Name'].isin(DAY1_FIELDS))
    ].copy()

    # Filter ES fields (exclude system fields)
    es_filtered = es_df[~es_df['Field API Name'].isin(SYSTEM_FIELDS)].copy()

    # Process each BBF field
    row_num = 2
    stats = {'High': 0, 'Medium': 0, 'Low': 0, 'None': 0}
    transformer_count = 0

    for _, bbf_row in bbf_filtered.iterrows():
        bbf_api = bbf_row['Field API Name']
        bbf_label = bbf_row['Field Label']
        bbf_type = bbf_row['Field Type']
        bbf_required = 'No' if bbf_row['Is Nillable'] == 'True' else 'Yes'

        # Find matching ES field
        confidence, es_match, transformer, notes = find_best_es_match(bbf_row, es_filtered)

        if es_match is not None:
            es_api = es_match['Field API Name']
            es_label = es_match['Field Label']
            es_type = es_match['Field Type']
        else:
            es_api = ''
            es_label = ''
            es_type = ''

        stats[confidence] += 1
        if transformer == 'Y':
            transformer_count += 1

        # Write row data
        row_data = [
            bbf_api, bbf_label, bbf_type, bbf_required,
            es_api, es_label, es_type,
            confidence, transformer, notes,
            '',  # ES_Final_Field - Business decision: override AI suggestion if needed
            'Yes' if confidence == 'High' else 'TBD',  # Include_in_Migration - Business decision
            ''  # Business_Notes - Business decision: reasoning for overrides
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            # Apply color based on confidence
            if confidence in COLORS:
                cell.fill = PatternFill(start_color=COLORS[confidence.lower()],
                                       end_color=COLORS[confidence.lower()],
                                       fill_type="solid")

        row_num += 1

    # Adjust column widths (includes new business decision columns)
    column_widths = [30, 30, 15, 15, 30, 30, 15, 15, 15, 60, 35, 20, 40]
    for col_num, width in enumerate(column_widths, 1):
        col_letter = chr(64 + col_num) if col_num <= 26 else 'A' + chr(64 + col_num - 26)
        ws.column_dimensions[col_letter].width = width

    return stats, transformer_count

def create_picklist_mapping_sheet(wb, es_picklist_df, bbf_picklist_df):
    """Create the Picklist_Mapping sheet."""
    ws = wb.create_sheet("Picklist_Mapping", 1)

    # Headers
    headers = [
        'ES_Field', 'ES_Picklist_Value', 'BBF_Field',
        'Suggested_Mapping', 'Notes', 'BBF_Final_Value'
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    row_num = 2
    picklist_field_count = 0

    # Payment Terms mapping
    payment_terms_map = {
        'NET30,NET45,NET60': 'NET 30',
        'NET90': 'NET 60'
    }

    if len(es_picklist_df[es_picklist_df['Field API Name'] == 'Payment_Terms__c']) > 0:
        picklist_field_count += 1
        bbf_payment_values = bbf_picklist_df[bbf_picklist_df['Field API Name'] == 'Payment_Terms__c']['Picklist Value'].tolist()
        bbf_payment_str = ' | '.join(bbf_payment_values)

        for es_val in ['NET30,NET45,NET60', 'NET90']:
            suggestion = payment_terms_map.get(es_val, bbf_payment_str)
            match_type = 'Close Match' if es_val in payment_terms_map else 'No Match - Select from list'
            color_key = 'medium' if match_type == 'Close Match' else 'low'

            row_data = [
                'Payment_Terms__c', es_val, 'Payment_Terms__c',
                suggestion, match_type, ''
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.fill = PatternFill(start_color=COLORS[color_key],
                                       end_color=COLORS[color_key],
                                       fill_type="solid")
            row_num += 1

    # Billing State (ES picklist) -> Billing State (BBF string)
    es_states = es_picklist_df[es_picklist_df['Field API Name'] == 'Billing_State__c']['Picklist Value'].tolist()
    if es_states:
        picklist_field_count += 1
        cell = ws.cell(row=row_num, column=1, value='Billing_State__c')
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.fill = PatternFill(start_color=COLORS['high'], end_color=COLORS['high'], fill_type="solid")

        cell = ws.cell(row=row_num, column=2, value='[All 53 state values]')
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.fill = PatternFill(start_color=COLORS['high'], end_color=COLORS['high'], fill_type="solid")

        cell = ws.cell(row=row_num, column=3, value='Billing_State__c')
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.fill = PatternFill(start_color=COLORS['high'], end_color=COLORS['high'], fill_type="solid")

        cell = ws.cell(row=row_num, column=4, value='Direct copy - BBF field is string type')
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.fill = PatternFill(start_color=COLORS['high'], end_color=COLORS['high'], fill_type="solid")

        cell = ws.cell(row=row_num, column=5, value='Exact Match - ES picklist to BBF string')
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.fill = PatternFill(start_color=COLORS['high'], end_color=COLORS['high'], fill_type="solid")

        cell = ws.cell(row=row_num, column=6, value='')
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.fill = PatternFill(start_color=COLORS['high'], end_color=COLORS['high'], fill_type="solid")

        row_num += 1

    # Invoice Delivery Preference -> Invoice Type
    invoice_delivery_map = {
        'Paper': 'Print',
        'E-mail': 'Email'
    }

    bbf_inv_values = bbf_picklist_df[bbf_picklist_df['Field API Name'] == 'invType__c']['Picklist Value'].tolist()
    bbf_inv_str = ' | '.join(bbf_inv_values)

    if len(es_picklist_df[es_picklist_df['Field API Name'] == 'Invoice_Delivery_Preference__c']) > 0:
        picklist_field_count += 1
        for es_val in ['Paper', 'E-mail']:
            suggestion = invoice_delivery_map.get(es_val, bbf_inv_str)
            match_type = 'Close Match'
            color_key = 'medium'

            row_data = [
                'Invoice_Delivery_Preference__c', es_val, 'invType__c',
                suggestion, match_type, ''
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.fill = PatternFill(start_color=COLORS[color_key],
                                       end_color=COLORS[color_key],
                                       fill_type="solid")
            row_num += 1

    # Billing Cycle -> Billing Schedule Group
    bbf_schedule_values = bbf_picklist_df[bbf_picklist_df['Field API Name'] == 'Billing_Schedule_Group__c']['Picklist Value'].tolist()
    bbf_schedule_str = ' | '.join(bbf_schedule_values)

    billing_cycle_map = {
        'Monthly': 'Monthly',
        'Manual': bbf_schedule_str
    }

    if len(es_picklist_df[es_picklist_df['Field API Name'] == 'Invoice_cycle_cd__c']) > 0:
        picklist_field_count += 1
        for es_val in ['Monthly', 'Manual']:
            suggestion = billing_cycle_map.get(es_val, bbf_schedule_str)
            match_type = 'Exact Match' if es_val == 'Monthly' else 'No Match - Select from list'
            color_key = 'high' if match_type == 'Exact Match' else 'low'

            row_data = [
                'Invoice_cycle_cd__c', es_val, 'Billing_Schedule_Group__c',
                suggestion, match_type, ''
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.fill = PatternFill(start_color=COLORS[color_key],
                                       end_color=COLORS[color_key],
                                       fill_type="solid")
            row_num += 1

    # Adjust column widths
    column_widths = [35, 30, 35, 50, 40, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width

    return picklist_field_count

def main():
    """Main execution function."""
    print("Loading field metadata...")
    es_df = pd.read_csv(ES_FIELDS_FILE)
    bbf_df = pd.read_csv(BBF_FIELDS_FILE)

    print("Loading picklist metadata...")
    es_picklist_df = pd.read_csv(ES_PICKLIST_FILE)
    bbf_picklist_df = pd.read_csv(BBF_PICKLIST_FILE)

    print(f"Loaded {len(es_df)} ES fields and {len(bbf_df)} BBF fields")

    print("Creating Excel workbook...")
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    print("Creating Field_Mapping sheet with semantic matching...")
    stats, transformer_count = create_field_mapping_sheet(wb, es_df, bbf_df)

    print("Creating Picklist_Mapping sheet...")
    picklist_field_count = create_picklist_mapping_sheet(wb, es_picklist_df, bbf_picklist_df)

    # Save workbook
    output_path = Path(OUTPUT_FILE)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"Saving workbook to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)

    print("\n" + "="*70)
    print("MAPPING COMPLETE")
    print("="*70)
    print(f"\nOutput: {OUTPUT_FILE}")
    print(f"\nField Mapping Statistics:")
    print(f"  High Confidence Matches: {stats['High']}")
    print(f"  Medium Confidence Matches: {stats['Medium']}")
    print(f"  Low Confidence Matches: {stats['Low']}")
    print(f"  No Match (Need Review): {stats['None']}")
    print(f"\nTotal BBF Fields Mapped: {sum(stats.values())}")
    print(f"Transformers Needed: {transformer_count}")
    print(f"Picklist Fields Mapped: {picklist_field_count}")
    print("="*70)

if __name__ == "__main__":
    main()
