#!/usr/bin/env python3
"""
Create Contact to Contact field mapping Excel file
ES Contact → BBF Contact
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from difflib import SequenceMatcher
import re

# Define system fields to exclude
SYSTEM_FIELDS = {
    'Id', 'IsDeleted', 'MasterRecordId', 'CreatedDate', 'CreatedById',
    'LastModifiedDate', 'LastModifiedById', 'SystemModstamp', 'LastActivityDate',
    'LastViewedDate', 'LastReferencedDate', 'JigsawContactId', 'Jigsaw',
    'PhotoUrl', 'CleanStatus'
}

# Define Day 1 migration fields to exclude (already migrated)
DAY_1_FIELDS = {
    'FirstName', 'LastName', 'Name', 'AccountId', 'OwnerId',
    'ES_Legacy_ID__c', 'BBF_New_Id__c', 'Email', 'Phone'
}

# Color codes (EXACT per spec)
COLOR_HEADER = 'FF366092'
COLOR_HIGH = 'FFC6EFCE'
COLOR_MEDIUM = 'FFFFEB9C'
COLOR_LOW = 'FFFFC7CE'
COLOR_NONE = 'FFFFC7CE'

def similarity_score(s1, s2):
    """Calculate similarity between two strings"""
    if not s1 or not s2:
        return 0.0
    return SequenceMatcher(None, s1.lower(), s2.lower()).ratio()

def normalize_field_name(name):
    """Normalize field name for matching"""
    # Remove common prefixes/suffixes
    name = re.sub(r'__c$', '', name)
    name = re.sub(r'^(ES_|BBF_)', '', name)
    return name.lower()

def match_fields(bbf_field, es_fields_df):
    """
    Find best matching ES field for a BBF field
    Returns: (es_field_api_name, es_field_label, es_data_type, confidence, notes)
    """
    bbf_api = bbf_field['Field API Name']
    bbf_label = bbf_field['Field Label']
    bbf_type = bbf_field['Field Type']

    best_match = None
    best_score = 0
    confidence = 'None'
    notes = ''

    # Priority 1: Exact API name match
    exact_match = es_fields_df[es_fields_df['Field API Name'] == bbf_api]
    if not exact_match.empty:
        match = exact_match.iloc[0]
        if match['Field Type'] == bbf_type:
            return (match['Field API Name'], match['Field Label'], match['Field Type'], 'High', 'Exact API name match, same type')
        else:
            return (match['Field API Name'], match['Field Label'], match['Field Type'], 'Medium', f'Exact API name match, type differs (ES: {match["Field Type"]}, BBF: {bbf_type})')

    # Priority 2: Normalized API name similarity
    bbf_norm = normalize_field_name(bbf_api)
    for _, es_field in es_fields_df.iterrows():
        es_norm = normalize_field_name(es_field['Field API Name'])
        score = similarity_score(bbf_norm, es_norm)

        if score > 0.9:  # Very high similarity
            if es_field['Field Type'] == bbf_type:
                return (es_field['Field API Name'], es_field['Field Label'], es_field['Field Type'], 'High', f'API name very similar ({score:.0%} match), same type')
            else:
                return (es_field['Field API Name'], es_field['Field Label'], es_field['Field Type'], 'Medium', f'API name very similar ({score:.0%} match), type differs')

        if score > best_score:
            best_score = score
            best_match = es_field

    # Priority 3: Label match
    for _, es_field in es_fields_df.iterrows():
        label_score = similarity_score(bbf_label, es_field['Field Label'])

        if label_score > 0.95:  # Very high label similarity
            if es_field['Field Type'] == bbf_type:
                return (es_field['Field API Name'], es_field['Field Label'], es_field['Field Type'], 'High', f'Label match ({label_score:.0%} similar), same type')
            else:
                return (es_field['Field API Name'], es_field['Field Label'], es_field['Field Type'], 'Medium', f'Label match ({label_score:.0%} similar), type differs')

    # Priority 4: Good API name similarity (0.7-0.9)
    if best_score > 0.7 and best_match is not None:
        if best_match['Field Type'] == bbf_type:
            return (best_match['Field API Name'], best_match['Field Label'], best_match['Field Type'], 'Medium', f'API name similar ({best_score:.0%} match), same type')
        else:
            return (best_match['Field API Name'], best_match['Field Label'], best_match['Field Type'], 'Low', f'API name similar ({best_score:.0%} match), type differs')

    # No good match found
    return ('', '', '', 'None', 'No obvious ES source field found - may need business rules or default value')

def needs_transformer(bbf_type, es_type, bbf_api, es_api, confidence):
    """Determine if a transformer is needed"""
    if confidence == 'None':
        return 'N'

    # Type mismatch
    if bbf_type != es_type:
        return 'Y'

    # Picklist fields - may need value mapping
    if bbf_type in ['picklist', 'multipicklist']:
        return 'Y'

    return 'N'

def create_field_mapping(bbf_df, es_df):
    """Create field mapping dataframe"""
    mapping_data = []

    for _, bbf_field in bbf_df.iterrows():
        es_api, es_label, es_type, confidence, notes = match_fields(bbf_field, es_df)

        transformer = needs_transformer(
            bbf_field['Field Type'],
            es_type,
            bbf_field['Field API Name'],
            es_api,
            confidence
        )

        mapping_data.append({
            'BBF_Field_API_Name': bbf_field['Field API Name'],
            'BBF_Field_Label': bbf_field['Field Label'],
            'BBF_Data_Type': bbf_field['Field Type'],
            'BBF_Is_Required': 'No' if bbf_field['Is Nillable'] == 'True' else 'Yes',
            'ES_Field_API_Name': es_api,
            'ES_Field_Label': es_label,
            'ES_Data_Type': es_type,
            'Match_Confidence': confidence,
            'Transformer_Needed': transformer,
            'Notes': notes,
            'ES_Final_Field': '',  # Business decision: override AI suggestion if needed
            'Include_in_Migration': 'Yes' if confidence == 'High' else 'TBD',  # Business decision
            'Business_Notes': ''  # Business decision: reasoning for overrides
        })

    return pd.DataFrame(mapping_data)

def parse_picklist_values(value_str):
    """Parse picklist values from CSV format"""
    if pd.isna(value_str) or value_str == '':
        return []
    # Values are separated by "; "
    return [v.strip() for v in str(value_str).split(';') if v.strip()]

def create_picklist_mapping(bbf_fields_df, es_fields_df, bbf_picklist_df, es_picklist_df):
    """Create picklist value mapping"""
    mapping_data = []

    # Get BBF picklist fields
    bbf_picklists = bbf_fields_df[bbf_fields_df['Field Type'].isin(['picklist', 'multipicklist'])]

    for _, bbf_field in bbf_picklists.iterrows():
        bbf_api = bbf_field['Field API Name']

        # Skip if no picklist values
        if pd.isna(bbf_field['Picklist Values']) or bbf_field['Picklist Values'] == '':
            continue

        # Get valid BBF values
        bbf_values_list = bbf_picklist_df[bbf_picklist_df['Field API Name'] == bbf_api]['Picklist Value'].tolist()

        if not bbf_values_list:
            continue

        # Find matching ES field
        es_match = es_fields_df[es_fields_df['Field API Name'] == bbf_api]

        if es_match.empty:
            # No ES field found - BBF-only field
            for bbf_val in bbf_values_list:
                mapping_data.append({
                    'ES_Field': '',
                    'ES_Picklist_Value': '',
                    'BBF_Field': bbf_api,
                    'Suggested_Mapping': bbf_val,
                    'Notes': 'BBF-only field - no ES source',
                    'BBF_Final_Value': ''
                })
            continue

        es_api = es_match.iloc[0]['Field API Name']

        # Get ES values for this field
        es_values_list = es_picklist_df[es_picklist_df['Field API Name'] == es_api]['Picklist Value'].tolist()

        if not es_values_list:
            # ES field exists but has no picklist values
            for bbf_val in bbf_values_list:
                mapping_data.append({
                    'ES_Field': es_api,
                    'ES_Picklist_Value': 'N/A - ES field not a picklist',
                    'BBF_Field': bbf_api,
                    'Suggested_Mapping': bbf_val,
                    'Notes': 'ES field exists but is not a picklist',
                    'BBF_Final_Value': ''
                })
            continue

        # Map each ES value to BBF
        for es_val in es_values_list:
            # Look for exact match
            exact_match = [bv for bv in bbf_values_list if bv.lower() == es_val.lower()]
            if exact_match:
                mapping_data.append({
                    'ES_Field': es_api,
                    'ES_Picklist_Value': es_val,
                    'BBF_Field': bbf_api,
                    'Suggested_Mapping': exact_match[0],
                    'Notes': 'Exact Match',
                    'BBF_Final_Value': ''
                })
                continue

            # Look for close match
            best_match = None
            best_score = 0
            for bv in bbf_values_list:
                score = similarity_score(es_val, bv)
                if score > best_score:
                    best_score = score
                    best_match = bv

            if best_score > 0.8:
                mapping_data.append({
                    'ES_Field': es_api,
                    'ES_Picklist_Value': es_val,
                    'BBF_Field': bbf_api,
                    'Suggested_Mapping': best_match,
                    'Notes': 'Close Match',
                    'BBF_Final_Value': ''
                })
            else:
                # No match - show all valid BBF values
                all_values = ' | '.join(bbf_values_list)
                mapping_data.append({
                    'ES_Field': es_api,
                    'ES_Picklist_Value': es_val,
                    'BBF_Field': bbf_api,
                    'Suggested_Mapping': all_values,
                    'Notes': 'No Match - Select from list',
                    'BBF_Final_Value': ''
                })

    return pd.DataFrame(mapping_data)

def apply_formatting(ws, df, is_picklist=False):
    """Apply formatting to worksheet"""
    # Header formatting
    header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Apply row colors based on confidence/match
    if is_picklist:
        # Picklist mapping - column 5 is Notes
        notes_col = 5
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(df)+1), start=2):
            notes_val = ws.cell(row_idx, notes_col).value

            if notes_val == 'Exact Match':
                fill = PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type='solid')
            elif notes_val == 'Close Match':
                fill = PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type='solid')
            elif notes_val and 'No Match' in notes_val:
                fill = PatternFill(start_color=COLOR_NONE, end_color=COLOR_NONE, fill_type='solid')
            else:
                continue

            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    else:
        # Field mapping - column 8 is Match_Confidence
        conf_col = 8
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(df)+1), start=2):
            conf_val = ws.cell(row_idx, conf_col).value

            if conf_val == 'High':
                fill = PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type='solid')
            elif conf_val == 'Medium':
                fill = PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type='solid')
            elif conf_val in ['Low', 'None']:
                fill = PatternFill(start_color=COLOR_LOW, end_color=COLOR_LOW, fill_type='solid')
            else:
                continue

            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Set column widths
    column_widths = {
        'A': 35, 'B': 35, 'C': 20, 'D': 15, 'E': 35,
        'F': 35, 'G': 20, 'H': 18, 'I': 18, 'J': 60
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

def main():
    print("Loading CSV exports...")

    # Load ES data
    es_fields = pd.read_csv('/mnt/c/Users/vjero/Documents/repos/python-repo/python_scripts/ES-to-BBF-migration/day-two/exports/es_Contact_fields_with_picklists.csv')
    es_picklist = pd.read_csv('/mnt/c/Users/vjero/Documents/repos/python-repo/python_scripts/ES-to-BBF-migration/day-two/exports/es_Contact_picklist_values.csv')

    # Load BBF data
    bbf_fields = pd.read_csv('/mnt/c/Users/vjero/Documents/repos/python-repo/python_scripts/ES-to-BBF-migration/day-two/exports/bbf_Contact_fields_with_picklists.csv')
    bbf_picklist = pd.read_csv('/mnt/c/Users/vjero/Documents/repos/python-repo/python_scripts/ES-to-BBF-migration/day-two/exports/bbf_Contact_picklist_values.csv')

    print(f"ES Fields: {len(es_fields)}")
    print(f"BBF Fields: {len(bbf_fields)}")

    # Filter out excluded fields from BBF
    all_excluded = SYSTEM_FIELDS | DAY_1_FIELDS
    bbf_fields_filtered = bbf_fields[~bbf_fields['Field API Name'].isin(all_excluded)].copy()

    print(f"BBF Fields after exclusions: {len(bbf_fields_filtered)}")
    print(f"Excluded: {len(bbf_fields) - len(bbf_fields_filtered)} fields")

    # Create field mapping
    print("\nCreating field mappings...")
    field_mapping_df = create_field_mapping(bbf_fields_filtered, es_fields)

    # Create picklist mapping
    print("Creating picklist mappings...")
    picklist_mapping_df = create_picklist_mapping(bbf_fields_filtered, es_fields, bbf_picklist, es_picklist)

    # Calculate statistics
    high_conf = len(field_mapping_df[field_mapping_df['Match_Confidence'] == 'High'])
    medium_conf = len(field_mapping_df[field_mapping_df['Match_Confidence'] == 'Medium'])
    low_conf = len(field_mapping_df[field_mapping_df['Match_Confidence'] == 'Low'])
    none_conf = len(field_mapping_df[field_mapping_df['Match_Confidence'] == 'None'])
    transformers = len(field_mapping_df[field_mapping_df['Transformer_Needed'] == 'Y'])

    print(f"\nField Mapping Statistics:")
    print(f"  High Confidence: {high_conf}")
    print(f"  Medium Confidence: {medium_conf}")
    print(f"  Low Confidence: {low_conf}")
    print(f"  No Match: {none_conf}")
    print(f"  Transformers Needed: {transformers}")

    if len(picklist_mapping_df) > 0:
        exact_match = len(picklist_mapping_df[picklist_mapping_df['Notes'] == 'Exact Match'])
        close_match = len(picklist_mapping_df[picklist_mapping_df['Notes'] == 'Close Match'])
        no_match = len(picklist_mapping_df[picklist_mapping_df['Notes'].str.contains('No Match', na=False)])

        print(f"\nPicklist Mapping Statistics:")
        print(f"  Total Picklist Values: {len(picklist_mapping_df)}")
        print(f"  Exact Match: {exact_match}")
        print(f"  Close Match: {close_match}")
        print(f"  No Match: {no_match}")

    # Create Excel workbook
    print("\nCreating Excel workbook...")
    wb = Workbook()

    # Sheet 1: Field Mapping
    ws1 = wb.active
    ws1.title = "Field_Mapping"

    for r_idx, row in enumerate(dataframe_to_rows(field_mapping_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws1.cell(row=r_idx, column=c_idx, value=value)

    apply_formatting(ws1, field_mapping_df, is_picklist=False)

    # Sheet 2: Picklist Mapping
    ws2 = wb.create_sheet("Picklist_Mapping")

    if len(picklist_mapping_df) > 0:
        for r_idx, row in enumerate(dataframe_to_rows(picklist_mapping_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws2.cell(row=r_idx, column=c_idx, value=value)

        apply_formatting(ws2, picklist_mapping_df, is_picklist=True)
    else:
        # Empty picklist sheet
        ws2['A1'] = 'ES_Field'
        ws2['B1'] = 'ES_Picklist_Value'
        ws2['C1'] = 'BBF_Field'
        ws2['D1'] = 'Suggested_Mapping'
        ws2['E1'] = 'Notes'
        ws2['F1'] = 'BBF_Final_Value'

    # Save workbook
    output_path = '/mnt/c/Users/vjero/Documents/repos/python-repo/python_scripts/ES-to-BBF-migration/day-two/mappings/ES_Contact_to_BBF_Contact_mapping.xlsx'
    wb.save(output_path)

    print(f"\nMapping file created successfully:")
    print(f"  {output_path}")

    return {
        'bbf_total': len(bbf_fields),
        'bbf_after_exclusions': len(bbf_fields_filtered),
        'es_total': len(es_fields),
        'high_conf': high_conf,
        'medium_conf': medium_conf,
        'low_conf': low_conf,
        'none_conf': none_conf,
        'transformers': transformers,
        'picklist_total': len(picklist_mapping_df) if len(picklist_mapping_df) > 0 else 0,
        'picklist_exact': exact_match if len(picklist_mapping_df) > 0 else 0,
        'picklist_close': close_match if len(picklist_mapping_df) > 0 else 0,
        'picklist_no_match': no_match if len(picklist_mapping_df) > 0 else 0
    }

if __name__ == '__main__':
    stats = main()
