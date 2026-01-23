#!/usr/bin/env python3
"""
Create field mapping Excel file from JSON data.
Used by the AI-powered day-two-field-mapping agent.

Usage:
    python create_mapping_excel.py --mapping-json mapping_data.json --output output.xlsx
"""

import argparse
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Color codes (EXACT per spec)
COLOR_HEADER = 'FF366092'
COLOR_HIGH = 'FFC6EFCE'
COLOR_MEDIUM = 'FFFFEB9C'
COLOR_LOW = 'FFFFC7CE'
COLOR_NONE = 'FFFFC7CE'


def create_field_mapping_sheet(ws, field_mappings):
    """Create the Field_Mapping sheet"""
    # Headers - includes business decision columns
    headers = [
        'BBF_Field_API_Name', 'BBF_Field_Label', 'BBF_Data_Type', 'BBF_Is_Required',
        'ES_Field_API_Name', 'ES_Field_Label', 'ES_Data_Type',
        'Match_Confidence', 'Transformer_Needed', 'Notes',
        'ES_Final_Field', 'Include_in_Migration', 'Business_Notes'
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Write data rows
    for row_idx, mapping in enumerate(field_mappings, 2):
        ws.cell(row=row_idx, column=1, value=mapping.get('bbf_field_api', ''))
        ws.cell(row=row_idx, column=2, value=mapping.get('bbf_label', ''))
        ws.cell(row=row_idx, column=3, value=mapping.get('bbf_type', ''))
        ws.cell(row=row_idx, column=4, value=mapping.get('bbf_required', 'No'))
        ws.cell(row=row_idx, column=5, value=mapping.get('es_field_api', ''))
        ws.cell(row=row_idx, column=6, value=mapping.get('es_label', ''))
        ws.cell(row=row_idx, column=7, value=mapping.get('es_type', ''))
        ws.cell(row=row_idx, column=8, value=mapping.get('confidence', 'None'))
        ws.cell(row=row_idx, column=9, value=mapping.get('transformer_needed', 'N'))
        ws.cell(row=row_idx, column=10, value=mapping.get('notes', ''))

        # Business decision columns - read from mapping if present, otherwise default
        confidence = mapping.get('confidence', 'None')
        es_final = mapping.get('es_final_field', '')
        include_mig = mapping.get('include_in_migration', 'Yes' if confidence == 'High' else 'TBD')
        biz_notes = mapping.get('business_notes', '')
        ws.cell(row=row_idx, column=11, value=es_final)  # ES_Final_Field - empty means accept AI suggestion
        ws.cell(row=row_idx, column=12, value=include_mig)  # Include_in_Migration
        ws.cell(row=row_idx, column=13, value=biz_notes)  # Business_Notes

        # Apply color based on confidence
        if confidence == 'High':
            fill_color = COLOR_HIGH
        elif confidence == 'Medium':
            fill_color = COLOR_MEDIUM
        else:
            fill_color = COLOR_LOW

        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        for col in range(1, 14):  # Updated to include new columns
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = fill
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Set column widths - includes new columns
    column_widths = [35, 35, 20, 15, 35, 35, 20, 18, 18, 60, 35, 20, 40]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def create_picklist_mapping_sheet(ws, picklist_mappings):
    """Create the Picklist_Mapping sheet"""
    # Headers
    headers = [
        'ES_Field', 'ES_Picklist_Value', 'BBF_Field',
        'Suggested_Mapping', 'Notes', 'BBF_Final_Value'
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Write data rows
    for row_idx, mapping in enumerate(picklist_mappings, 2):
        ws.cell(row=row_idx, column=1, value=mapping.get('es_field', ''))
        ws.cell(row=row_idx, column=2, value=mapping.get('es_value', ''))
        ws.cell(row=row_idx, column=3, value=mapping.get('bbf_field', ''))
        ws.cell(row=row_idx, column=4, value=mapping.get('suggested_mapping', ''))
        ws.cell(row=row_idx, column=5, value=mapping.get('notes', ''))
        ws.cell(row=row_idx, column=6, value=mapping.get('bbf_final_value', ''))

        # Apply color based on match type
        notes = mapping.get('notes', '')
        if notes == 'Exact Match':
            fill_color = COLOR_HIGH
        elif notes == 'Close Match':
            fill_color = COLOR_MEDIUM
        else:
            fill_color = COLOR_LOW

        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        for col in range(1, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = fill
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Set column widths
    column_widths = [35, 40, 35, 60, 30, 30]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def main():
    parser = argparse.ArgumentParser(description='Create field mapping Excel from JSON')
    parser.add_argument('--mapping-json', required=True, help='Path to mapping JSON file')
    parser.add_argument('--output', required=True, help='Output Excel file path')
    args = parser.parse_args()

    # Load JSON data
    try:
        with open(args.mapping_json, 'r') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"Error: JSON file not found: {args.mapping_json}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON: {e}")
        sys.exit(1)

    field_mappings = data.get('field_mappings', [])
    picklist_mappings = data.get('picklist_mappings', [])

    print(f"Creating Excel with {len(field_mappings)} field mappings and {len(picklist_mappings)} picklist mappings...")

    # Create workbook
    wb = Workbook()

    # Sheet 1: Field Mapping
    ws1 = wb.active
    ws1.title = "Field_Mapping"
    create_field_mapping_sheet(ws1, field_mappings)

    # Sheet 2: Picklist Mapping
    ws2 = wb.create_sheet("Picklist_Mapping")
    create_picklist_mapping_sheet(ws2, picklist_mappings)

    # Save workbook
    wb.save(args.output)
    print(f"Excel file created successfully: {args.output}")

    # Print statistics
    high_conf = sum(1 for m in field_mappings if m.get('confidence') == 'High')
    medium_conf = sum(1 for m in field_mappings if m.get('confidence') == 'Medium')
    low_conf = sum(1 for m in field_mappings if m.get('confidence') == 'Low')
    none_conf = sum(1 for m in field_mappings if m.get('confidence') == 'None')
    transformers = sum(1 for m in field_mappings if m.get('transformer_needed') == 'Y')

    print(f"\nField Mapping Statistics:")
    print(f"  High Confidence: {high_conf}")
    print(f"  Medium Confidence: {medium_conf}")
    print(f"  Low Confidence: {low_conf}")
    print(f"  No Match: {none_conf}")
    print(f"  Transformers Needed: {transformers}")

    if picklist_mappings:
        exact = sum(1 for m in picklist_mappings if m.get('notes') == 'Exact Match')
        close = sum(1 for m in picklist_mappings if m.get('notes') == 'Close Match')
        no_match = sum(1 for m in picklist_mappings if 'No Match' in m.get('notes', ''))

        print(f"\nPicklist Mapping Statistics:")
        print(f"  Total Values: {len(picklist_mappings)}")
        print(f"  Exact Match: {exact}")
        print(f"  Close Match: {close}")
        print(f"  No Match: {no_match}")


if __name__ == '__main__':
    main()
