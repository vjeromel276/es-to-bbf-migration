#!/usr/bin/env python3
"""
Create BAN__c field mapping Excel workbook with proper formatting.
Maps ES Billing_Invoice__c to BBF BAN__c.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from difflib import SequenceMatcher

# System fields to exclude (from all objects)
SYSTEM_FIELDS = {
    'Id', 'IsDeleted', 'MasterRecordId', 'CreatedDate', 'CreatedById',
    'LastModifiedDate', 'LastModifiedById', 'SystemModstamp',
    'LastActivityDate', 'LastViewedDate', 'LastReferencedDate',
    'JigsawContactId', 'Jigsaw', 'PhotoUrl', 'CleanStatus'
}

# Day 1 migration fields for BAN__c (already populated)
DAY1_BAN_FIELDS = {
    'Name', 'OwnerId', 'ES_Legacy_ID__c', 'BBF_New_Id__c', 'Account__c'
}

# Manual field mappings (overrides for better matching)
MANUAL_FIELD_MAPPINGS = {
    'Billing_PostalCode__c': 'Billing_ZIP__c',
    'Billing_City__c': 'Billing_City__c',
    'Billing_State__c': 'Billing_State__c',
    'Billing_Street__c': 'Billing_Address_1__c',
    'BAN_Description__c': 'Description__c',
    'Payment_Terms__c': 'Payment_Terms__c',
    'Billing_Company_Name__c': 'Account_Name__c',
    'Federal_Tax_ID__c': None,  # No ES equivalent
    'invType__c': 'Invoice_Delivery_Preference__c',
}

# Color definitions (EXACT hex codes from spec)
HEADER_COLOR = 'FF366092'  # Dark blue
HIGH_CONFIDENCE_COLOR = 'FFC6EFCE'  # Light green
MEDIUM_CONFIDENCE_COLOR = 'FFFFEB9C'  # Light yellow
LOW_CONFIDENCE_COLOR = 'FFFFC7CE'  # Light red/pink
NONE_CONFIDENCE_COLOR = 'FFFFC7CE'  # Light red/pink

def similarity_ratio(str1, str2):
    """Calculate similarity ratio between two strings."""
    return SequenceMatcher(None, str1.lower(), str2.lower()).ratio()

def normalize_field_name(name):
    """Normalize field name for comparison."""
    # Remove common prefixes/suffixes
    normalized = name.lower()
    normalized = re.sub(r'__c$', '', normalized)
    normalized = re.sub(r'^billing_', '', normalized)
    normalized = re.sub(r'^invoice_', '', normalized)
    normalized = re.sub(r'^ban_', '', normalized)
    normalized = re.sub(r'^account_', '', normalized)
    return normalized

def determine_match_confidence(bbf_field, es_field, bbf_type, es_type, is_manual=False):
    """Determine match confidence level."""
    if not es_field:
        return 'None'

    # Manual mappings get high confidence if types match
    if is_manual:
        if bbf_type == es_type:
            return 'High'
        return 'Medium'

    bbf_api = bbf_field.lower()
    es_api = es_field.lower()

    # Exact API name match
    if bbf_api == es_api:
        if bbf_type == es_type:
            return 'High'
        return 'Medium'

    # Normalized API name match
    bbf_norm = normalize_field_name(bbf_field)
    es_norm = normalize_field_name(es_field)

    if bbf_norm == es_norm:
        if bbf_type == es_type:
            return 'High'
        return 'Medium'

    # High similarity in API names
    if similarity_ratio(bbf_api, es_api) > 0.85:
        if bbf_type == es_type:
            return 'High'
        return 'Medium'

    # Medium similarity
    if similarity_ratio(bbf_api, es_api) > 0.6:
        return 'Medium'

    return 'Low'

def needs_transformer(bbf_type, es_type, bbf_field, es_field):
    """Determine if a transformer is needed."""
    if not es_field:
        return 'N'

    # Different data types
    if bbf_type != es_type:
        return 'Y'

    # Picklist fields (always need validation/transformation)
    if bbf_type == 'picklist':
        return 'Y'

    return 'N'

def find_best_match(bbf_row, es_df):
    """Find best matching ES field for a BBF field."""
    bbf_api = bbf_row['Field API Name']
    bbf_label = bbf_row['Field Label']
    bbf_type = bbf_row['Field Type']

    # Check manual mappings first
    if bbf_api in MANUAL_FIELD_MAPPINGS:
        es_field_name = MANUAL_FIELD_MAPPINGS[bbf_api]
        if es_field_name is None:
            return None, 0, False

        es_match = es_df[es_df['Field API Name'] == es_field_name]
        if not es_match.empty:
            return es_match.iloc[0], 1.0, True

    best_match = None
    best_score = 0

    for _, es_row in es_df.iterrows():
        es_api = es_row['Field API Name']
        es_label = es_row['Field Label']
        es_type = es_row['Field Type']

        # Skip system fields
        if es_api in SYSTEM_FIELDS:
            continue

        # Exact API name match
        if bbf_api.lower() == es_api.lower():
            return es_row, 1.0, False

        # Normalized API name match
        if normalize_field_name(bbf_api) == normalize_field_name(es_api):
            score = 0.95 if bbf_type == es_type else 0.85
            if score > best_score:
                best_match = es_row
                best_score = score

        # Label match
        if bbf_label.lower() == es_label.lower():
            score = 0.80 if bbf_type == es_type else 0.70
            if score > best_score:
                best_match = es_row
                best_score = score

        # API name similarity
        api_sim = similarity_ratio(bbf_api, es_api)
        if api_sim > 0.7:
            score = api_sim * (0.9 if bbf_type == es_type else 0.7)
            if score > best_score:
                best_match = es_row
                best_score = score

        # Label similarity
        label_sim = similarity_ratio(bbf_label, es_label)
        if label_sim > 0.7:
            score = label_sim * (0.7 if bbf_type == es_type else 0.6)
            if score > best_score:
                best_match = es_row
                best_score = score

    return best_match, best_score, False

def create_field_mapping(bbf_df, es_df):
    """Create field mapping DataFrame."""
    mappings = []

    for _, bbf_row in bbf_df.iterrows():
        bbf_api = bbf_row['Field API Name']
        bbf_label = bbf_row['Field Label']
        bbf_type = bbf_row['Field Type']
        bbf_nillable = bbf_row['Is Nillable']

        # Skip excluded fields
        if bbf_api in SYSTEM_FIELDS or bbf_api in DAY1_BAN_FIELDS:
            continue

        # Find best ES match
        es_match, score, is_manual = find_best_match(bbf_row, es_df)

        if es_match is not None and score > 0.5:
            es_api = es_match['Field API Name']
            es_label = es_match['Field Label']
            es_type = es_match['Field Type']
        else:
            es_api = ''
            es_label = ''
            es_type = ''

        confidence = determine_match_confidence(bbf_api, es_api, bbf_type, es_type, is_manual)
        transformer = needs_transformer(bbf_type, es_type, bbf_api, es_api)

        # Create notes
        notes = []
        if not es_api:
            notes.append('No obvious ES source field identified')
        elif bbf_type != es_type and es_type:
            notes.append(f'Data type mismatch: ES {es_type} → BBF {bbf_type}')
        elif bbf_type == 'picklist':
            notes.append('Picklist field - requires value mapping')
        elif confidence == 'High':
            if is_manual:
                notes.append('Manual mapping - strong semantic match')
            else:
                notes.append('Strong field match - direct mapping likely')
        elif confidence == 'Medium':
            notes.append('Moderate match - verify field mapping')
        elif confidence == 'Low':
            notes.append('Weak match - manual review required')

        mappings.append({
            'BBF_Field_API_Name': bbf_api,
            'BBF_Field_Label': bbf_label,
            'BBF_Data_Type': bbf_type,
            'BBF_Is_Required': 'Yes' if not bbf_nillable else 'No',
            'ES_Field_API_Name': es_api,
            'ES_Field_Label': es_label,
            'ES_Data_Type': es_type,
            'Match_Confidence': confidence,
            'Transformer_Needed': transformer,
            'Notes': ' | '.join(notes)
        })

    return pd.DataFrame(mappings)

def create_picklist_mapping(bbf_picklist_df, es_picklist_df, field_mapping_df):
    """Create picklist value mapping DataFrame."""
    picklist_mappings = []

    # Get BBF picklist fields that have ES matches
    for _, row in field_mapping_df.iterrows():
        bbf_field = row['BBF_Field_API_Name']
        es_field = row['ES_Field_API_Name']
        bbf_type = row['BBF_Data_Type']

        if bbf_type != 'picklist' or not es_field:
            continue

        # Get ES picklist values for this field
        es_values = es_picklist_df[es_picklist_df['Field API Name'] == es_field]

        # Get BBF picklist values for this field
        bbf_values = bbf_picklist_df[bbf_picklist_df['Field API Name'] == bbf_field]
        bbf_value_list = bbf_values['Picklist Value'].tolist()

        # If no ES values, skip this field
        if es_values.empty:
            continue

        # Map each ES value to BBF
        for _, es_val_row in es_values.iterrows():
            es_value = es_val_row['Picklist Value']

            # Try to find exact match
            exact_match = None
            for bbf_val in bbf_value_list:
                if es_value.lower() == bbf_val.lower():
                    exact_match = bbf_val
                    break

            if exact_match:
                suggested = exact_match
                notes = 'Exact Match'
            else:
                # Try close match
                close_match = None
                best_sim = 0
                for bbf_val in bbf_value_list:
                    sim = similarity_ratio(es_value, bbf_val)
                    if sim > 0.7 and sim > best_sim:
                        close_match = bbf_val
                        best_sim = sim

                if close_match:
                    suggested = close_match
                    notes = 'Close Match'
                else:
                    # No match - show all BBF values
                    suggested = ' | '.join(bbf_value_list)
                    notes = 'No Match - Select from list'

            picklist_mappings.append({
                'ES_Field': es_field,
                'ES_Picklist_Value': es_value,
                'BBF_Field': bbf_field,
                'Suggested_Mapping': suggested,
                'Notes': notes,
                'BBF_Final_Value': ''
            })

    return pd.DataFrame(picklist_mappings)

def apply_formatting(wb, ws, sheet_name):
    """Apply formatting to worksheet."""
    # Header formatting
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Apply row colors based on confidence/match type
    if sheet_name == 'Field_Mapping':
        confidence_col = 8  # Match_Confidence column (H)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            confidence = row[confidence_col - 1].value

            if confidence == 'High':
                fill_color = HIGH_CONFIDENCE_COLOR
            elif confidence == 'Medium':
                fill_color = MEDIUM_CONFIDENCE_COLOR
            elif confidence == 'Low':
                fill_color = LOW_CONFIDENCE_COLOR
            else:  # None
                fill_color = NONE_CONFIDENCE_COLOR

            row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            for cell in row:
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    elif sheet_name == 'Picklist_Mapping':
        notes_col = 5  # Notes column (E)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            notes = row[notes_col - 1].value

            if notes and 'Exact Match' in notes:
                fill_color = HIGH_CONFIDENCE_COLOR
            elif notes and 'Close Match' in notes:
                fill_color = MEDIUM_CONFIDENCE_COLOR
            else:  # No Match
                fill_color = LOW_CONFIDENCE_COLOR

            row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            for cell in row:
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 50

def main():
    # Read exported metadata
    bbf_df = pd.read_csv('day-two/exports/bbf_BAN__c_fields_with_picklists.csv')
    es_df = pd.read_csv('day-two/exports/es_Billing_Invoice__c_fields_with_picklists.csv')
    bbf_picklist_df = pd.read_csv('day-two/exports/bbf_BAN__c_picklist_values.csv')
    es_picklist_df = pd.read_csv('day-two/exports/es_Billing_Invoice__c_picklist_values.csv')

    print(f"Loaded BBF BAN__c: {len(bbf_df)} fields")
    print(f"Loaded ES Billing_Invoice__c: {len(es_df)} fields")

    # Create field mapping
    field_mapping_df = create_field_mapping(bbf_df, es_df)
    print(f"Created field mapping: {len(field_mapping_df)} BBF fields (after exclusions)")

    # Create picklist mapping
    picklist_mapping_df = create_picklist_mapping(bbf_picklist_df, es_picklist_df, field_mapping_df)
    print(f"Created picklist mapping: {len(picklist_mapping_df)} picklist values")

    # Count statistics
    high_count = len(field_mapping_df[field_mapping_df['Match_Confidence'] == 'High'])
    medium_count = len(field_mapping_df[field_mapping_df['Match_Confidence'] == 'Medium'])
    low_count = len(field_mapping_df[field_mapping_df['Match_Confidence'] == 'Low'])
    none_count = len(field_mapping_df[field_mapping_df['Match_Confidence'] == 'None'])
    transformer_count = len(field_mapping_df[field_mapping_df['Transformer_Needed'] == 'Y'])

    exact_match_count = len(picklist_mapping_df[picklist_mapping_df['Notes'] == 'Exact Match'])
    close_match_count = len(picklist_mapping_df[picklist_mapping_df['Notes'] == 'Close Match'])
    no_match_count = len(picklist_mapping_df[picklist_mapping_df['Notes'] == 'No Match - Select from list'])

    print(f"\nField Mapping Stats:")
    print(f"  High Confidence: {high_count}")
    print(f"  Medium Confidence: {medium_count}")
    print(f"  Low Confidence: {low_count}")
    print(f"  No Match: {none_count}")
    print(f"  Transformers Needed: {transformer_count}")

    print(f"\nPicklist Mapping Stats:")
    print(f"  Exact Matches: {exact_match_count}")
    print(f"  Close Matches: {close_match_count}")
    print(f"  No Match: {no_match_count}")

    # Create Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: Field_Mapping
    ws1 = wb.create_sheet('Field_Mapping')
    for r_idx, row in enumerate(dataframe_to_rows(field_mapping_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws1.cell(row=r_idx, column=c_idx, value=value)
    apply_formatting(wb, ws1, 'Field_Mapping')

    # Sheet 2: Picklist_Mapping
    ws2 = wb.create_sheet('Picklist_Mapping')
    for r_idx, row in enumerate(dataframe_to_rows(picklist_mapping_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=value)
    apply_formatting(wb, ws2, 'Picklist_Mapping')

    # Save workbook
    output_file = 'day-two/mappings/ES_Billing_Invoice__c_to_BBF_BAN__c_mapping.xlsx'
    wb.save(output_file)
    print(f"\n✅ Mapping saved to: {output_file}")

    return {
        'bbf_fields': len(bbf_df),
        'es_fields': len(es_df),
        'mapped_fields': len(field_mapping_df),
        'high_confidence': high_count,
        'medium_confidence': medium_count,
        'low_confidence': low_count,
        'no_match': none_count,
        'transformers': transformer_count,
        'picklist_values': len(picklist_mapping_df),
        'exact_picklist': exact_match_count,
        'close_picklist': close_match_count,
        'no_match_picklist': no_match_count
    }

if __name__ == '__main__':
    stats = main()
