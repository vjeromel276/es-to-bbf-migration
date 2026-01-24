#!/usr/bin/env python3
"""
Generate Product Mapping Template Excel for Business Review

Creates an Excel file with:
1. ES Product_Family__c → BBF Service_Type_Charge__c mapping (with AI suggestions)
2. ES Product_Name__c → BBF Product_Simple__c mapping (with AI suggestions)
3. Bandwidth extraction patterns from Description field
"""

from collections import Counter
from simple_salesforce import Salesforce
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import re

# Color codes
COLOR_HEADER = 'FF366092'
COLOR_HIGH = 'FFC6EFCE'      # Green - High confidence match
COLOR_MEDIUM = 'FFFFEB9C'    # Yellow - Medium confidence
COLOR_LOW = 'FFFFC7CE'       # Red - Low confidence / needs review
COLOR_WHITE = 'FFFFFFFF'

# AI-suggested mappings: ES Product_Family__c → BBF Service_Type_Charge__c
FAMILY_TO_SERVICE_TYPE = {
    'Point-to-Point (PTPS)': ('EPL', 'High', 'Point-to-Point maps to Ethernet Private Line'),
    'Dark Fiber (DFBR)': ('DF', 'High', 'Dark Fiber direct match'),
    'Dark Fiber Pair IRU (DFBR)': ('DF', 'High', 'Dark Fiber IRU maps to DF'),
    'Dedicated Internet Access (DIAS)': ('DIA', 'High', 'DIA direct match'),
    'Internet Only (DIAS)': ('DIA', 'High', 'Internet maps to DIA'),
    'Virtual Dedicated Internet Access (VDIA)': ('DIA', 'Medium', 'Virtual DIA likely maps to DIA'),
    'Hosted Voice (VOIC)': ('VOICE', 'High', 'Voice direct match'),
    'PRI (SIP) (VOIC)': ('VOICE', 'High', 'PRI/SIP is Voice'),
    'SS7  (VOIC)': ('SS7', 'High', 'SS7 direct match'),
    'Collocation (COLO)': ('COLO', 'High', 'Collocation direct match'),
    'Rack (COLO)': ('COLO', 'High', 'Rack is Collocation'),
    'Power (COLO)': ('Power', 'High', 'Power direct match'),
    'Network-to-Network Interface (NNIS)': ('ENNI', 'High', 'NNI maps to ENNI'),
    'Point-to-MultiPoint (PMPS)': ('ELAN', 'Medium', 'Point-to-MultiPoint likely maps to ELAN'),
    'Ethernet (EPL/EVPL) (ETHS)': ('EPL', 'High', 'Ethernet EPL/EVPL maps to EPL'),
    'Ethernet Transport Only (ETHS)': ('EPL', 'Medium', 'Ethernet Transport maps to EPL'),
    'Dedicated DWDM (DWDM)': ('Wavelength', 'High', 'DWDM maps to Wavelength'),
    'Managed Wave (MWAV)': ('Wavelength', 'High', 'Managed Wave maps to Wavelength'),
    'Private Line (PVLS)': ('EPL', 'Medium', 'Private Line likely maps to EPL'),
    'Data Center Services (CLDS)': ('COLO', 'Medium', 'Data Center likely maps to COLO'),
    'IP': ('IP Subnet', 'Medium', 'IP likely maps to IP Subnet'),
    'Promotions': ('Credit', 'Medium', 'Promotions often result in Credits'),
    'NRC (Non-Recurring Charge)': ('Installation', 'Medium', 'NRC often is Installation'),
    'Work Order - Unknown (WOUK)': ('', 'Low', 'NEEDS BUSINESS DECISION - 28.7% of records'),
    'Work Order - Labor (WOLB)': ('Installation', 'Low', 'Labor likely maps to Installation'),
    'Managed Service (MSP)': ('', 'Low', 'NEEDS BUSINESS DECISION - what MSP type?'),
    'Tagged / Untagged': ('VLAN', 'Medium', 'VLAN tagging'),
    'Handoff Type': ('', 'Low', 'Handoff is attribute, not service type'),
    'Logical Attributes': ('', 'Low', 'Attribute, not service type'),
    'Diversity': ('', 'Low', 'Attribute, not service type'),
    'Routing': ('L3VPN', 'Medium', 'Routing likely maps to L3VPN'),
    'Additional Port': ('XCN', 'Medium', 'Additional port likely cross-connect'),
    'Tandem': ('SS7', 'Low', 'Tandem may map to SS7'),
    'TSP Fee': ('TSP', 'High', 'TSP Fee direct match'),
    'Equipment & Managed Equipment': ('Equipment', 'High', 'Equipment direct match'),
}

# Product_Name__c → Product_Simple__c direct mapping (same as transformer)
PRODUCT_NAME_TO_PRODUCT_SIMPLE = {
    # Internet/DIA products
    'Dedicated Internet Access': ('DIA', 'High', 'DIA direct match'),
    'Internet': ('DIA', 'High', 'Internet maps to DIA'),
    'Virtual Dedicated Internet Access': ('DIA', 'Medium', 'Virtual DIA'),

    # Dark Fiber products
    'Dark Fiber': ('Dark Fiber', 'High', 'Direct match'),
    'Dark Fiber IRU': ('Dark Fiber', 'High', 'IRU is Dark Fiber'),
    'Dark Fiber Lease': ('Dark Fiber', 'High', 'Lease is Dark Fiber'),
    '2 Strands': ('Dark Fiber', 'High', 'Strand count = Dark Fiber'),
    '4 Strands': ('Dark Fiber', 'High', 'Strand count = Dark Fiber'),
    '6 Strands': ('Dark Fiber', 'High', 'Strand count = Dark Fiber'),
    '8 Strands': ('Dark Fiber', 'High', 'Strand count = Dark Fiber'),
    '12 Strands': ('Dark Fiber', 'High', 'Strand count = Dark Fiber'),
    '24 Strands': ('Dark Fiber', 'High', 'Strand count = Dark Fiber'),
    '48 Strands': ('Dark Fiber', 'High', 'Strand count = Dark Fiber'),
    '60 Strands': ('Dark Fiber', 'High', 'Strand count = Dark Fiber'),
    '72 Strands': ('Dark Fiber', 'High', 'Strand count = Dark Fiber'),
    '96 Strands': ('Dark Fiber', 'High', 'Strand count = Dark Fiber'),
    '144 Strands': ('Dark Fiber', 'High', 'Strand count = Dark Fiber'),

    # Ethernet products
    'Ethernet': ('Ethernet Transport', 'High', 'Direct match'),
    'Ethernet (EPL/EVPL)': ('Ethernet Transport', 'High', 'EPL/EVPL = Ethernet'),
    'Point-to-Point': ('Ethernet Transport', 'High', 'P2P = Ethernet Transport'),
    'Point-to-MultiPoint': ('Ethernet Transport', 'Medium', 'P2MP = Ethernet'),
    'Private Line': ('Ethernet Transport', 'Medium', 'Private Line = Ethernet'),

    # IP products
    'IPv4 Blocks /30  (1 Gateway, 1 Usable)': ('/30 Static IP', 'High', 'Exact IP block match'),
    'IPv4 Blocks /29  (1 Gateway, 5 Usable)': ('/29 Static IP', 'High', 'Exact IP block match'),
    'IPv4 Blocks /28  (1 Gateway, 13 Usable)': ('/28 Static IP', 'High', 'Exact IP block match'),
    'IPv4 Blocks /27  (1 Gateway, 29 Usable)': ('/27 Static IP', 'High', 'Exact IP block match'),
    'IPv4 Blocks /26  (1 Gateway, 61 Usable)': ('/26 Static IP', 'High', 'Exact IP block match'),
    'IPv4 Blocks /25  (1 Gateway, 125 Usable)': ('/25 Static IP', 'High', 'Exact IP block match'),
    'IPv4 Blocks /24  (1 Gateway, 253 Usable)': ('/24 Static IP', 'High', 'Exact IP block match'),
    'eBGP': ('IP ADDR', 'Medium', 'BGP routing = IP'),

    # Voice products
    'Voice Recurring Charges': ('Voice Trunk', 'High', 'Voice = Voice Trunk'),
    'Hosted Voice': ('Voice Trunk', 'High', 'Hosted Voice = Voice Trunk'),
    'PRI': ('Voice Trunk', 'High', 'PRI = Voice Trunk'),
    'SIP': ('Voice Trunk', 'High', 'SIP = Voice Trunk'),
    'Toll Free': ('Toll Free', 'High', 'Direct match'),
    'DID': ('DID', 'High', 'Direct match'),

    # Colocation products
    'Collocation': ('COLO', 'High', 'Direct match'),
    'Colocation': ('COLO', 'High', 'Direct match (alt spelling)'),
    'Full Cabinet': ('Full Cabinet', 'High', 'Direct match'),
    'Half Cabinet': ('Half Cabinet', 'High', 'Direct match'),
    'Cabinet Relocate': ('COLO', 'Medium', 'Cabinet work = COLO'),
    'Rack Space': ('Rack Unit', 'High', 'Rack = Rack Unit'),
    'Floor Space': ('Floor Space', 'High', 'Direct match'),
    'Power': ('AC Power', 'High', 'Power = AC Power'),

    # Cross-connect products
    'Cross Connect': ('Cross Connect', 'High', 'Direct match'),
    'Handoff Type : Copper': ('Cross Connect', 'Medium', 'Handoff = Cross Connect'),
    'Handoff Type : Fiber': ('Cross Connect', 'Medium', 'Handoff = Cross Connect'),

    # DWDM/Wavelength products
    'DWDM': ('Wavelength', 'High', 'DWDM = Wavelength'),
    'Managed Wave': ('Wavelength', 'High', 'Managed Wave = Wavelength'),
    'Wavelength': ('Wavelength', 'High', 'Direct match'),

    # Network products
    'Network-to-Network Interface': ('ENNI', 'High', 'NNI = ENNI'),
    'NNI': ('ENNI', 'High', 'Direct match'),

    # Logical/VLAN products
    'Logical Attribute : Layer 2': ('VLAN', 'Medium', 'L2 = VLAN'),
    'Logical Attribute : Layer 3': ('L3VPN', 'Medium', 'L3 = L3VPN'),
    'Tagged': ('VLAN', 'Medium', 'Tagged = VLAN'),
    'Untagged': ('VLAN', 'Medium', 'Untagged = VLAN'),

    # Labor/Installation products
    'Labor': ('Install', 'High', 'Labor = Install'),
    'Installation': ('Install', 'High', 'Direct match'),
    'Expedite': ('Expedite', 'High', 'Direct match'),
    'Smart Hands': ('Smart Hands Hourly', 'High', 'Smart Hands match'),

    # Promotions/Credits/Discounts
    'Discount': ('DISCOUNT', 'High', 'Direct match'),
    'Credit': ('Credit', 'High', 'Direct match'),
    'New Service Promotion: First 3 Months Free': ('Credit', 'High', 'Promo = Credit'),
    'New Service Promotion: First 5 Months Free': ('Credit', 'High', 'Promo = Credit'),
    'New Service Promotion:  First 5 Months Free': ('Credit', 'High', 'Promo = Credit'),

    # Equipment
    'Rocket Fiber Service Delivery Device': ('EqpLease', 'Medium', 'Device = Equipment'),
    'Equipment': ('EqpLease', 'High', 'Equipment = EqpLease'),
    'Router': ('EqpLease', 'Medium', 'Router = Equipment'),

    # Diversity products
    'Diversity': ('Ethernet Transport', 'Low', 'Diversity attribute'),
    'Diverse Path': ('Ethernet Transport', 'Low', 'Diversity attribute'),

    # TSP products
    'TSP': ('TSP-MRC', 'High', 'TSP match'),
    'TSP Fee': ('TSP-MRC', 'High', 'TSP match'),

    # SS7 products
    'SS7': ('SS7', 'High', 'Direct match'),
    'ISUP': ('SS7', 'Medium', 'ISUP = SS7'),

    # Unknown/Undetermined - needs default
    'Unknown': ('ANNUAL', 'Low', 'NEEDS REVIEW - 28.7% of records'),
    '00000_UNDETERMINED': ('ANNUAL', 'Low', 'NEEDS REVIEW - undetermined product'),
}

# Bandwidth patterns to extract from Description
BANDWIDTH_PATTERNS = [
    (r'(\d+)\s*Gbps', 'Gbps'),
    (r'(\d+)\s*Mbps', 'Mbps'),
    (r'(\d+)Gbps', 'Gbps'),
    (r'(\d+)Mbps', 'Mbps'),
    (r'0*(\d+)Gbps', 'Gbps'),  # Handle leading zeros like "01Gbps"
    (r'0*(\d+)Mbps', 'Mbps'),  # Handle leading zeros like "0100Mbps"
]

def extract_bandwidth(description):
    """Extract bandwidth from description field."""
    if not description:
        return None, None

    for pattern, unit in BANDWIDTH_PATTERNS:
        match = re.search(pattern, description, re.IGNORECASE)
        if match:
            value = int(match.group(1))
            return value, unit
    return None, None

def bandwidth_to_product_simple(value, unit):
    """Convert bandwidth to BBF Product_Simple__c value."""
    if value is None:
        return None

    if unit == 'Gbps':
        return f"{value} Gbps"
    elif unit == 'Mbps':
        return f"{value} Mbps"
    return None

def main():
    print("=" * 80)
    print("GENERATING PRODUCT MAPPING TEMPLATE")
    print("=" * 80)

    # Connect to ES UAT
    print("\n📌 Connecting to ES UAT...")
    es_sf = Salesforce(
        username='sfdcapi@everstream.net.uat',
        password='ZXasqw1234!@#$',
        security_token='X0ation2CNmK5C0pV94M6vFYS',
        domain='test'
    )
    print(f"✅ Connected to ES: {es_sf.sf_instance}")

    # Connect to BBF
    print("\n📌 Connecting to BBF...")
    bbf_sf = Salesforce(
        username='vlettau@everstream.net',
        password='MNlkpo0987)(*&',
        security_token='I4xmQLmm03cXl1O9qI2Z3XAAX',
        domain='test'
    )
    print(f"✅ Connected to BBF: {bbf_sf.sf_instance}")

    # Query ES OrderItems
    print("\n📊 Querying ES OrderItem data...")
    query = '''
    SELECT Id, Product_Name__c, Product_Family__c, Description
    FROM OrderItem
    WHERE Order.Status IN ('Activated', 'Suspended (Late Payment)', 'Disconnect in Progress')
    AND (Order.Project_Group__c = null OR (NOT Order.Project_Group__c LIKE '%PA MARKET DECOM%'))
    LIMIT 10000
    '''
    result = es_sf.query_all(query)
    records = result['records']
    print(f"✅ Retrieved {len(records)} OrderItem records")

    # Aggregate data
    family_counter = Counter()
    product_counter = Counter()
    bandwidth_counter = Counter()
    family_bandwidth = {}  # Track bandwidths per family

    for r in records:
        family = r.get('Product_Family__c') or '(null)'
        product = r.get('Product_Name__c') or '(null)'
        desc = r.get('Description') or ''

        family_counter[family] += 1
        product_counter[product] += 1

        # Extract bandwidth
        bw_value, bw_unit = extract_bandwidth(desc)
        if bw_value:
            bw_str = bandwidth_to_product_simple(bw_value, bw_unit)
            bandwidth_counter[bw_str] += 1

            # Track bandwidths per family
            if family not in family_bandwidth:
                family_bandwidth[family] = Counter()
            family_bandwidth[family][bw_str] += 1

    # Get BBF picklist values
    print("\n📋 Getting BBF picklist values...")
    desc = bbf_sf.Service_Charge__c.describe()

    bbf_service_types = []
    bbf_products = []

    for field in desc['fields']:
        if field['name'] == 'Service_Type_Charge__c':
            bbf_service_types = [pv['value'] for pv in field['picklistValues'] if pv['active']]
        if field['name'] == 'Product_Simple__c':
            bbf_products = [pv['value'] for pv in field['picklistValues'] if pv['active']]

    print(f"   BBF Service_Type_Charge__c: {len(bbf_service_types)} values")
    print(f"   BBF Product_Simple__c: {len(bbf_products)} values")

    # Create Excel workbook
    print("\n📝 Creating Excel mapping template...")
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # =========================================================================
    # SHEET 1: Product_Family → Service_Type_Charge Mapping
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Family_to_ServiceType"

    headers1 = [
        'ES Product_Family__c', 'Count', '%',
        'AI Suggested BBF Service_Type_Charge__c', 'Confidence', 'AI Notes',
        'BUSINESS DECISION: Final BBF Value', 'Business Notes'
    ]

    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row = 2
    for family, count in family_counter.most_common():
        pct = count / len(records) * 100

        # Get AI suggestion
        suggestion = FAMILY_TO_SERVICE_TYPE.get(family, ('', 'Low', 'No mapping defined - needs business input'))
        suggested_value, confidence, notes = suggestion

        # Determine fill color
        if confidence == 'High':
            fill = PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type='solid')
        elif confidence == 'Medium':
            fill = PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type='solid')
        else:
            fill = PatternFill(start_color=COLOR_LOW, end_color=COLOR_LOW, fill_type='solid')

        ws1.cell(row=row, column=1, value=family).border = thin_border
        ws1.cell(row=row, column=2, value=count).border = thin_border
        ws1.cell(row=row, column=3, value=f"{pct:.1f}%").border = thin_border
        ws1.cell(row=row, column=4, value=suggested_value).border = thin_border
        ws1.cell(row=row, column=5, value=confidence).border = thin_border
        ws1.cell(row=row, column=6, value=notes).border = thin_border
        ws1.cell(row=row, column=7, value='').border = thin_border  # Business decision
        ws1.cell(row=row, column=8, value='').border = thin_border  # Business notes

        # Apply fill to row
        for col in range(1, 9):
            ws1.cell(row=row, column=col).fill = fill

        row += 1

    # Column widths
    ws1.column_dimensions['A'].width = 40
    ws1.column_dimensions['B'].width = 10
    ws1.column_dimensions['C'].width = 8
    ws1.column_dimensions['D'].width = 35
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 50
    ws1.column_dimensions['G'].width = 35
    ws1.column_dimensions['H'].width = 40
    ws1.freeze_panes = 'A2'

    # =========================================================================
    # SHEET 2: Bandwidth → Product_Simple Mapping
    # =========================================================================
    ws2 = wb.create_sheet("Bandwidth_to_Product")

    headers2 = [
        'Extracted Bandwidth', 'Count', '%',
        'BBF Product_Simple__c Match', 'Match Status',
        'BUSINESS DECISION: Final BBF Value', 'Notes'
    ]

    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row = 2
    for bw, count in bandwidth_counter.most_common():
        pct = count / len(records) * 100

        # Check if exact match exists in BBF
        if bw in bbf_products:
            match_status = 'Exact Match'
            fill = PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type='solid')
        else:
            match_status = 'No Exact Match - Review'
            fill = PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type='solid')

        ws2.cell(row=row, column=1, value=bw).border = thin_border
        ws2.cell(row=row, column=2, value=count).border = thin_border
        ws2.cell(row=row, column=3, value=f"{pct:.1f}%").border = thin_border
        ws2.cell(row=row, column=4, value=bw if bw in bbf_products else '').border = thin_border
        ws2.cell(row=row, column=5, value=match_status).border = thin_border
        ws2.cell(row=row, column=6, value='').border = thin_border
        ws2.cell(row=row, column=7, value='').border = thin_border

        for col in range(1, 8):
            ws2.cell(row=row, column=col).fill = fill

        row += 1

    # Add row for records without bandwidth
    no_bw_count = len(records) - sum(bandwidth_counter.values())
    ws2.cell(row=row, column=1, value='(No Bandwidth Detected)').border = thin_border
    ws2.cell(row=row, column=2, value=no_bw_count).border = thin_border
    ws2.cell(row=row, column=3, value=f"{no_bw_count/len(records)*100:.1f}%").border = thin_border
    ws2.cell(row=row, column=4, value='').border = thin_border
    ws2.cell(row=row, column=5, value='Needs Default Value').border = thin_border
    ws2.cell(row=row, column=6, value='').border = thin_border
    ws2.cell(row=row, column=7, value='Set default Product_Simple__c for records without bandwidth').border = thin_border

    fill = PatternFill(start_color=COLOR_LOW, end_color=COLOR_LOW, fill_type='solid')
    for col in range(1, 8):
        ws2.cell(row=row, column=col).fill = fill

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 8
    ws2.column_dimensions['D'].width = 30
    ws2.column_dimensions['E'].width = 20
    ws2.column_dimensions['F'].width = 30
    ws2.column_dimensions['G'].width = 50
    ws2.freeze_panes = 'A2'

    # =========================================================================
    # SHEET 3: Product_Name → Product_Simple Mapping (NEW - main mapping)
    # =========================================================================
    ws3 = wb.create_sheet("ProductName_to_ProductSimple")

    headers3 = [
        'ES Product_Name__c', 'Count', '%',
        'AI Suggested BBF Product_Simple__c', 'Confidence', 'AI Notes',
        'BUSINESS DECISION: Final BBF Value', 'Business Notes'
    ]

    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row = 2
    mapped_count = 0
    unmapped_count = 0

    for product_name, count in product_counter.most_common():
        pct = count / len(records) * 100

        # Get AI suggestion from mapping
        if product_name in PRODUCT_NAME_TO_PRODUCT_SIMPLE:
            suggested_value, confidence, notes = PRODUCT_NAME_TO_PRODUCT_SIMPLE[product_name]
            mapped_count += count
        else:
            suggested_value = ''
            confidence = 'None'
            notes = 'No mapping defined - needs business input'
            unmapped_count += count

        # Determine fill color
        if confidence == 'High':
            fill = PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type='solid')
        elif confidence == 'Medium':
            fill = PatternFill(start_color=COLOR_MEDIUM, end_color=COLOR_MEDIUM, fill_type='solid')
        elif confidence == 'Low':
            fill = PatternFill(start_color=COLOR_LOW, end_color=COLOR_LOW, fill_type='solid')
        else:
            fill = PatternFill(start_color=COLOR_LOW, end_color=COLOR_LOW, fill_type='solid')

        ws3.cell(row=row, column=1, value=product_name).border = thin_border
        ws3.cell(row=row, column=2, value=count).border = thin_border
        ws3.cell(row=row, column=3, value=f"{pct:.1f}%").border = thin_border
        ws3.cell(row=row, column=4, value=suggested_value).border = thin_border
        ws3.cell(row=row, column=5, value=confidence).border = thin_border
        ws3.cell(row=row, column=6, value=notes).border = thin_border
        ws3.cell(row=row, column=7, value='').border = thin_border  # Business decision
        ws3.cell(row=row, column=8, value='').border = thin_border  # Business notes

        # Apply fill to row
        for col in range(1, 9):
            ws3.cell(row=row, column=col).fill = fill

        row += 1

    # Column widths
    ws3.column_dimensions['A'].width = 50
    ws3.column_dimensions['B'].width = 10
    ws3.column_dimensions['C'].width = 8
    ws3.column_dimensions['D'].width = 35
    ws3.column_dimensions['E'].width = 12
    ws3.column_dimensions['F'].width = 40
    ws3.column_dimensions['G'].width = 35
    ws3.column_dimensions['H'].width = 40
    ws3.freeze_panes = 'A2'

    # =========================================================================
    # SHEET 4: BBF Service_Type_Charge__c Reference
    # =========================================================================
    ws4 = wb.create_sheet("BBF_ServiceType_Values")

    ws4.cell(row=1, column=1, value='BBF Service_Type_Charge__c Picklist Values').font = Font(bold=True, size=14)
    ws4.cell(row=2, column=1, value=f'Total: {len(bbf_service_types)} values')

    row = 4
    for val in sorted(bbf_service_types):
        ws4.cell(row=row, column=1, value=val)
        row += 1

    ws4.column_dimensions['A'].width = 30

    # =========================================================================
    # SHEET 5: BBF Product_Simple__c Reference
    # =========================================================================
    ws5 = wb.create_sheet("BBF_Product_Values")

    ws5.cell(row=1, column=1, value='BBF Product_Simple__c Picklist Values').font = Font(bold=True, size=14)
    ws5.cell(row=2, column=1, value=f'Total: {len(bbf_products)} values')

    row = 4
    for val in sorted(bbf_products):
        ws5.cell(row=row, column=1, value=val)
        row += 1

    ws5.column_dimensions['A'].width = 40

    # =========================================================================
    # SHEET 6: Summary & Instructions
    # =========================================================================
    ws6 = wb.create_sheet("Instructions")
    ws6.insert_rows(1)  # Move to first position
    wb.move_sheet(ws6, offset=-5)

    mapped_pct = mapped_count / len(records) * 100 if len(records) > 0 else 0
    unmapped_pct = unmapped_count / len(records) * 100 if len(records) > 0 else 0

    instructions = [
        ('ES → BBF Service_Charge__c Product Mapping Template', Font(bold=True, size=16)),
        ('', None),
        ('Generated: ' + datetime.now().strftime('%Y-%m-%d %H:%M:%S'), None),
        ('', None),
        ('PURPOSE:', Font(bold=True, size=12)),
        ('This template maps ES OrderItem Product data to BBF Service_Charge__c fields:', None),
        ('  • Product_Name__c → Product_Simple__c (PRIMARY - direct product mapping)', None),
        ('  • Product_Family__c → Service_Type_Charge__c (service category)', None),
        ('  • Description (bandwidth) → Product_Simple__c (SECONDARY - bandwidth extraction)', None),
        ('', None),
        ('MAPPING PRIORITY (for Product_Simple__c):', Font(bold=True, size=12)),
        ('  1. Direct Product_Name__c mapping (most accurate)', None),
        ('  2. Bandwidth extracted from Description field', None),
        ('  3. Bandwidth extracted from Product_Name__c field', None),
        ('  4. Service type default (based on Service_Type_Charge__c)', None),
        ('  5. Ultimate default: "ANNUAL"', None),
        ('', None),
        ('HOW TO USE:', Font(bold=True, size=12)),
        ('1. Review Sheet "ProductName_to_ProductSimple" - PRIMARY Product_Simple__c mapping', None),
        ('2. Review Sheet "Family_to_ServiceType" - AI suggestions for Service_Type_Charge__c', None),
        ('3. Review Sheet "Bandwidth_to_Product" - Bandwidth extraction for Product_Simple__c', None),
        ('4. Fill in "BUSINESS DECISION" columns where AI confidence is Low/None', None),
        ('5. Return completed file for transformer script generation', None),
        ('', None),
        ('COLOR LEGEND:', Font(bold=True, size=12)),
        ('  Green = High confidence match (likely correct)', None),
        ('  Yellow = Medium confidence (review recommended)', None),
        ('  Red = Low confidence or No mapping (business decision required)', None),
        ('', None),
        ('STATISTICS:', Font(bold=True, size=12)),
        (f'  Total ES OrderItems analyzed: {len(records):,}', None),
        ('', None),
        ('  Product_Name__c Mapping Coverage:', Font(bold=True)),
        (f'    Unique Product_Name__c values: {len(product_counter)}', None),
        (f'    Records with AI mapping: {mapped_count:,} ({mapped_pct:.1f}%)', None),
        (f'    Records needing business input: {unmapped_count:,} ({unmapped_pct:.1f}%)', None),
        ('', None),
        ('  Product_Family__c → Service_Type_Charge__c:', Font(bold=True)),
        (f'    Unique Product_Family__c values: {len(family_counter)}', None),
        ('', None),
        ('  Bandwidth Extraction (SECONDARY):', Font(bold=True)),
        (f'    Unique bandwidths extracted: {len(bandwidth_counter)}', None),
        (f'    Records with bandwidth: {sum(bandwidth_counter.values()):,} ({sum(bandwidth_counter.values())/len(records)*100:.1f}%)', None),
        (f'    Records without bandwidth: {no_bw_count:,} ({no_bw_count/len(records)*100:.1f}%)', None),
        ('', None),
        ('REFERENCE SHEETS:', Font(bold=True, size=12)),
        ('  • "BBF_ServiceType_Values" - All valid BBF Service_Type_Charge__c picklist values', None),
        ('  • "BBF_Product_Values" - All valid BBF Product_Simple__c picklist values', None),
    ]

    for row_idx, (text, font) in enumerate(instructions, 1):
        cell = ws6.cell(row=row_idx, column=1, value=text)
        if font:
            cell.font = font

    ws6.column_dimensions['A'].width = 80

    # Save workbook
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f'day-two/mappings/ES_Product_to_BBF_ServiceCharge_mapping_{timestamp}.xlsx'
    wb.save(output_file)

    print(f"\n✅ Mapping template created: {output_file}")
    print(f"\n📊 Summary:")
    print(f"   Total records analyzed: {len(records):,}")
    print(f"\n   Product_Name__c → Product_Simple__c:")
    print(f"      Unique Product_Name values: {len(product_counter)}")
    print(f"      Records with AI mapping: {mapped_count:,} ({mapped_pct:.1f}%)")
    print(f"      Records needing business input: {unmapped_count:,} ({unmapped_pct:.1f}%)")
    print(f"\n   Product_Family → Service_Type:")
    print(f"      Product_Family values: {len(family_counter)}")
    print(f"\n   Bandwidth extraction (secondary):")
    print(f"      Bandwidth patterns found: {len(bandwidth_counter)}")
    print(f"      Records with bandwidth: {sum(bandwidth_counter.values()):,} ({sum(bandwidth_counter.values())/len(records)*100:.1f}%)")

    return output_file


if __name__ == '__main__':
    main()
