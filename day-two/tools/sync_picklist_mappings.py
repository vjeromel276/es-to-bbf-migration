#!/usr/bin/env python3
"""
Sync Picklist Mappings Tool
===========================
Synchronizes the Picklist_Mapping sheet with ES_Final_Field values from Field_Mapping.

When business owners change ES_Final_Field to a different source field, this tool:
1. Reads Field_Mapping to get the actual ES source field (ES_Final_Field or ES_Field_API_Name)
2. Queries ES Salesforce for that field's picklist values
3. Updates Picklist_Mapping sheet with correct ES field name and values
4. Preserves existing BBF_Final_Value translations where possible

Usage:
    # Sync a specific mapping file
    python sync_picklist_mappings.py --mapping ES_Account_to_BBF_Account_mapping.xlsx

    # Sync all mapping files
    python sync_picklist_mappings.py --all

    # Dry run - show what would change without modifying
    python sync_picklist_mappings.py --mapping ES_Account_to_BBF_Account_mapping.xlsx --dry-run
"""

import argparse
import os
import sys
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime

import pandas as pd
from simple_salesforce import Salesforce
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ---------------------------------------------------------------------
# Salesforce credentials (ES org)
# ---------------------------------------------------------------------
ES_USERNAME = "sfdcapi@everstream.net"
ES_PASSWORD = "pV4CAxns8DQtJsBq!"
ES_TOKEN = "r1uoYiusK19RbrflARydi86TA"
ES_DOMAIN = "login"

# ---------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------
MAPPINGS_DIR = Path(__file__).parent.parent / "mappings"
EXPORTS_DIR = Path(__file__).parent.parent / "exports"


def load_bbf_picklist_values(object_name: str) -> Dict[str, List[str]]:
    """
    Load BBF picklist values from exports CSV.

    Returns dict of field_name -> list of active picklist values.
    """
    # Try to find the BBF picklist export file
    csv_path = EXPORTS_DIR / f"bbf_{object_name}_picklist_values.csv"
    xlsx_path = EXPORTS_DIR / f"bbf_{object_name}_picklist_values.xlsx"

    result = {}

    if csv_path.exists():
        df = pd.read_csv(csv_path)
    elif xlsx_path.exists():
        df = pd.read_excel(xlsx_path)
    else:
        print(f"  ⚠️ No BBF picklist export found for {object_name}")
        return result

    # Group by field and collect active values
    for field_name in df['Field API Name'].unique():
        field_rows = df[(df['Field API Name'] == field_name) & (df['Is Active'] == True)]
        values = field_rows['Picklist Value'].dropna().tolist()
        if values:
            result[field_name] = values

    return result


def get_bbf_object_from_filename(filename: str) -> str:
    """Extract BBF object name from mapping filename."""
    import re
    match = re.match(r'ES_.+?_to_BBF_(.+?)_mapping\.xlsx', filename)
    if match:
        return match.group(1)
    return None


def connect_to_es_salesforce() -> Salesforce:
    """Connect to ES Salesforce org."""
    print("🔌 Connecting to ES Salesforce...")
    sf = Salesforce(
        username=ES_USERNAME,
        password=ES_PASSWORD,
        security_token=ES_TOKEN,
        domain=ES_DOMAIN,
    )
    print(f"✅ Connected to: {sf.sf_instance}")
    return sf


def is_valid_api_name(field_name: str) -> bool:
    """Check if a string looks like a valid Salesforce API field name."""
    if not field_name or not isinstance(field_name, str):
        return False
    field_name = field_name.strip()
    if not field_name or ' ' in field_name:
        return False
    import re
    if field_name.endswith('__c') or re.match(r'^[A-Za-z][A-Za-z0-9_]*$', field_name):
        return True
    return False


def get_es_object_from_filename(filename: str) -> str:
    """Extract ES object name from mapping filename."""
    import re
    match = re.match(r'ES_(.+?)_to_BBF_', filename)
    if match:
        return match.group(1)
    return None


def get_picklist_values_for_field(sf: Salesforce, object_name: str, field_name: str) -> List[str]:
    """
    Get picklist values for a specific field from Salesforce.

    Returns list of active picklist values.
    """
    try:
        desc = sf.__getattr__(object_name).describe()
        for field in desc.get('fields', []):
            if field.get('name') == field_name:
                if field.get('type') not in ['picklist', 'multipicklist']:
                    return []
                values = []
                for pv in field.get('picklistValues', []):
                    if pv.get('active', False):
                        values.append(pv.get('value', ''))
                return values
        return []
    except Exception as e:
        print(f"  ⚠️ Error getting picklist values for {object_name}.{field_name}: {e}")
        return []


def get_all_picklist_fields(sf: Salesforce, object_name: str) -> Dict[str, List[str]]:
    """
    Get all picklist fields and their values for an object.

    Returns dict of field_name -> list of values.
    """
    result = {}
    try:
        desc = sf.__getattr__(object_name).describe()
        for field in desc.get('fields', []):
            if field.get('type') in ['picklist', 'multipicklist']:
                values = []
                for pv in field.get('picklistValues', []):
                    if pv.get('active', False):
                        values.append(pv.get('value', ''))
                result[field.get('name')] = values
    except Exception as e:
        print(f"⚠️ Error describing {object_name}: {e}")
    return result


def analyze_mapping_file(filepath: Path) -> Dict[str, Any]:
    """
    Analyze a mapping file to find picklist fields and their ES sources.

    Returns dict with:
    - picklist_fields: List of {bbf_field, es_field, bbf_type} for picklist fields
    - current_picklist_mapping: DataFrame of current Picklist_Mapping sheet
    """
    result = {
        'picklist_fields': [],
        'current_picklist_mapping': None,
        'field_mapping_df': None,
    }

    # Read Field_Mapping sheet
    try:
        df_fields = pd.read_excel(filepath, sheet_name='Field_Mapping')
        result['field_mapping_df'] = df_fields
    except Exception as e:
        print(f"  ❌ Error reading Field_Mapping: {e}")
        return result

    # Read current Picklist_Mapping sheet
    try:
        df_picklists = pd.read_excel(filepath, sheet_name='Picklist_Mapping')
        result['current_picklist_mapping'] = df_picklists
    except Exception:
        result['current_picklist_mapping'] = pd.DataFrame()

    # Find picklist fields
    for _, row in df_fields.iterrows():
        bbf_field = row.get('BBF_Field_API_Name')
        bbf_type = str(row.get('BBF_Data_Type', '')).lower()

        # Only process picklist fields
        if 'picklist' not in bbf_type:
            continue

        # Get the ES source field (prefer ES_Final_Field if valid)
        es_final = row.get('ES_Final_Field')
        es_api = row.get('ES_Field_API_Name')
        es_final_str = str(es_final).strip() if pd.notna(es_final) else ''
        es_field = es_final_str if is_valid_api_name(es_final_str) else es_api

        # Check Include_in_Migration
        include = str(row.get('Include_in_Migration', 'Yes')).strip().lower()
        if include == 'no':
            continue

        if pd.notna(bbf_field) and pd.notna(es_field) and str(es_field).strip():
            result['picklist_fields'].append({
                'bbf_field': bbf_field,
                'es_field': str(es_field).strip(),
                'bbf_type': bbf_type,
                'es_original': str(es_api) if pd.notna(es_api) else '',
            })

    return result


def build_new_picklist_mapping(
    sf: Salesforce,
    es_object: str,
    bbf_object: str,
    picklist_fields: List[Dict],
    current_mapping: pd.DataFrame
) -> pd.DataFrame:
    """
    Build new picklist mapping DataFrame.

    Preserves existing BBF_Final_Value translations where ES values match.
    """
    # Get all ES picklist values
    print(f"  📡 Fetching picklist values from ES {es_object}...")
    es_picklists = get_all_picklist_fields(sf, es_object)

    # Load BBF picklist values from exports
    print(f"  📂 Loading BBF picklist values for {bbf_object}...")
    bbf_picklists = load_bbf_picklist_values(bbf_object)

    # Build lookup of existing data: (bbf_field, es_value) -> {bbf_final_value, suggested_mapping}
    existing_data = {}
    if current_mapping is not None and not current_mapping.empty:
        for _, row in current_mapping.iterrows():
            bbf_field = row.get('BBF_Field')
            es_value = row.get('ES_Picklist_Value')
            bbf_final = row.get('BBF_Final_Value')
            suggested = row.get('Suggested_Mapping')
            if pd.notna(bbf_field) and pd.notna(es_value):
                existing_data[(str(bbf_field), str(es_value))] = {
                    'bbf_final': str(bbf_final) if pd.notna(bbf_final) else '',
                    'suggested': str(suggested) if pd.notna(suggested) else '',
                }

    # Build new rows
    new_rows = []
    for pf in picklist_fields:
        bbf_field = pf['bbf_field']
        es_field = pf['es_field']

        # Get ES picklist values for this field
        es_values = es_picklists.get(es_field, [])

        # Get BBF picklist values for suggestion (pipe-separated)
        bbf_values = bbf_picklists.get(bbf_field, [])
        bbf_suggestion = '|'.join(bbf_values) if bbf_values else ''

        if not es_values:
            print(f"    ⚠️ No picklist values found for ES field: {es_field}")
            # Still add a placeholder row
            new_rows.append({
                'BBF_Field': bbf_field,
                'ES_Field': es_field,
                'ES_Picklist_Value': '',
                'Suggested_Mapping': '',
                'BBF_Final_Value': '',
                'Notes': 'No ES picklist values found',
            })
            continue

        for es_value in es_values:
            # Check for existing data (translations and suggestions)
            existing = existing_data.get((bbf_field, es_value), {})
            existing_final = existing.get('bbf_final', '')
            existing_suggested = existing.get('suggested', '')

            # Use existing suggested if available, otherwise use BBF values from export
            suggested = existing_suggested if existing_suggested else bbf_suggestion

            # Determine notes (use 'No Match' for compatibility with recommend_picklist_values.py)
            if existing_final:
                notes = 'Preserved'
            elif existing_suggested:
                notes = 'Has Suggestion'
            else:
                notes = 'No Match'  # This triggers the AI recommendation tool

            new_rows.append({
                'BBF_Field': bbf_field,
                'ES_Field': es_field,
                'ES_Picklist_Value': es_value,
                'Suggested_Mapping': suggested,  # Preserve existing or populate with BBF options
                'BBF_Final_Value': existing_final,
                'Notes': notes,
            })

    return pd.DataFrame(new_rows)


def sync_mapping_file(
    filepath: Path,
    sf: Salesforce,
    dry_run: bool = False
) -> Dict[str, Any]:
    """
    Sync a single mapping file's Picklist_Mapping sheet.

    Returns summary of changes.
    """
    filename = filepath.name
    print(f"\n{'='*60}")
    print(f"Processing: {filename}")
    print(f"{'='*60}")

    # Get ES and BBF object names
    es_object = get_es_object_from_filename(filename)
    bbf_object = get_bbf_object_from_filename(filename)
    if not es_object or not bbf_object:
        print(f"  ❌ Could not determine object names from filename")
        return {'error': 'Invalid filename'}
    print(f"  ES Object: {es_object}")
    print(f"  BBF Object: {bbf_object}")

    # Analyze current mapping
    analysis = analyze_mapping_file(filepath)
    picklist_fields = analysis['picklist_fields']
    current_mapping = analysis['current_picklist_mapping']

    print(f"  Picklist fields found: {len(picklist_fields)}")

    if not picklist_fields:
        print(f"  ℹ️ No picklist fields to sync")
        return {'picklist_fields': 0, 'changes': 0}

    # Show what fields we're syncing
    print(f"\n  Picklist fields to sync:")
    for pf in picklist_fields:
        changed = pf['es_field'] != pf['es_original']
        indicator = " ⚠️ CHANGED" if changed else ""
        print(f"    {pf['bbf_field']:<35} <- {pf['es_field']}{indicator}")

    # Build new picklist mapping
    new_mapping = build_new_picklist_mapping(sf, es_object, bbf_object, picklist_fields, current_mapping)

    # Count changes
    current_count = len(current_mapping) if current_mapping is not None else 0
    new_count = len(new_mapping)
    preserved = len(new_mapping[new_mapping['Notes'] == 'Preserved']) if not new_mapping.empty else 0
    needs_translation = len(new_mapping[new_mapping['Notes'] == 'No Match']) if not new_mapping.empty else 0

    print(f"\n  Summary:")
    print(f"    Current picklist rows: {current_count}")
    print(f"    New picklist rows: {new_count}")
    print(f"    Preserved translations: {preserved}")
    print(f"    Needs translation: {needs_translation}")

    if dry_run:
        print(f"\n  [DRY RUN] Would update Picklist_Mapping sheet")
        if not new_mapping.empty:
            print(f"\n  Preview of new mapping (first 10 rows):")
            print(new_mapping.head(10).to_string())
    else:
        # Write updated Picklist_Mapping sheet
        print(f"\n  📝 Updating Picklist_Mapping sheet...")

        # Load workbook and replace sheet
        wb = load_workbook(filepath)

        # Remove existing Picklist_Mapping sheet if it exists
        if 'Picklist_Mapping' in wb.sheetnames:
            del wb['Picklist_Mapping']

        # Create new sheet
        ws = wb.create_sheet('Picklist_Mapping')

        # Write headers
        headers = ['BBF_Field', 'ES_Field', 'ES_Picklist_Value', 'Suggested_Mapping', 'BBF_Final_Value', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')

        # Write data
        for row_idx, row_data in new_mapping.iterrows():
            for col_idx, header in enumerate(headers):
                ws.cell(row=row_idx + 2, column=col_idx + 1, value=row_data.get(header, ''))

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        # Save
        wb.save(filepath)
        print(f"  ✅ Saved updated mapping file")

    return {
        'picklist_fields': len(picklist_fields),
        'current_rows': current_count,
        'new_rows': new_count,
        'preserved': preserved,
        'needs_translation': needs_translation,
    }


def main():
    parser = argparse.ArgumentParser(
        description='Sync Picklist_Mapping sheet with ES_Final_Field values'
    )
    parser.add_argument(
        '--mapping',
        type=str,
        help='Specific mapping file to sync (e.g., ES_Account_to_BBF_Account_mapping.xlsx)'
    )
    parser.add_argument(
        '--all',
        action='store_true',
        help='Sync all mapping files'
    )
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Show what would change without modifying files'
    )

    args = parser.parse_args()

    if not args.mapping and not args.all:
        parser.print_help()
        print("\nExamples:")
        print("  python sync_picklist_mappings.py --mapping ES_Account_to_BBF_Account_mapping.xlsx")
        print("  python sync_picklist_mappings.py --all")
        print("  python sync_picklist_mappings.py --all --dry-run")
        sys.exit(1)

    # Connect to ES Salesforce
    sf = connect_to_es_salesforce()

    # Get files to process
    if args.all:
        mapping_files = list(MAPPINGS_DIR.glob('ES_*_to_BBF_*_mapping.xlsx'))
        print(f"\nFound {len(mapping_files)} mapping files")
    else:
        filepath = MAPPINGS_DIR / args.mapping
        if not filepath.exists():
            print(f"❌ File not found: {filepath}")
            sys.exit(1)
        mapping_files = [filepath]

    # Process each file
    results = []
    for filepath in sorted(mapping_files):
        result = sync_mapping_file(filepath, sf, dry_run=args.dry_run)
        result['file'] = filepath.name
        results.append(result)

    # Summary
    print(f"\n{'='*60}")
    print("SYNC COMPLETE")
    print(f"{'='*60}")

    total_needs_translation = 0
    for r in results:
        if 'error' not in r:
            needs = r.get('needs_translation', 0)
            total_needs_translation += needs
            print(f"  {r['file']}: {r.get('new_rows', 0)} rows, {needs} need translation")

    if total_needs_translation > 0:
        print(f"\n⚠️  {total_needs_translation} picklist values need BBF translations")
        print("   Open the Excel files and fill in BBF_Final_Value column")


if __name__ == '__main__':
    main()
