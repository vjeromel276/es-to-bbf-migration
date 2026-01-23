#!/usr/bin/env python3
"""
AI-powered picklist value recommender for ES to BBF migration.

This script reads mapping Excel files and uses semantic understanding to
recommend the best BBF picklist values for ES values marked as "No Match".

Usage:
    # Process a specific mapping file
    python recommend_picklist_values.py --input path/to/mapping.xlsx

    # Process all mapping files
    python recommend_picklist_values.py --all

    # Preview recommendations without updating Excel
    python recommend_picklist_values.py --input path/to/mapping.xlsx --dry-run
"""

import argparse
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Error: Required packages not installed. Run: pip install pandas openpyxl")
    sys.exit(1)


# Mapping directory
MAPPINGS_DIR = Path(__file__).parent.parent / "mappings"


# ============================================================================
# SEMANTIC MATCHING KNOWLEDGE BASE
# ============================================================================

# Common semantic equivalences across domains
SEMANTIC_EQUIVALENCES = {
    # Payment Terms
    'net30': 'NET 30',
    'net 30': 'NET 30',
    'net15': 'NET 15',
    'net 15': 'NET 15',
    'net45': 'NET 45',
    'net 45': 'NET 45',
    'net60': 'NET 60',
    'net 60': 'NET 60',
    'due on receipt': 'Due Upon Receipt',
    'due upon receipt': 'Due Upon Receipt',
    'prepaid': 'Prepaid',
    'prepay': 'Prepaid',

    # Billing/Invoice Types
    'paper': 'Print',
    'print': 'Print',
    'electronic': 'Email',
    'email': 'Email',
    'e-mail': 'Email',
    'e-invoice': 'Email',
    'portal': 'Portal',
    'online': 'Portal',
    'edi': 'EDI',

    # Contact Types
    'decision maker': 'Executive',
    'decisionmaker': 'Executive',
    'executive': 'Executive',
    'c-level': 'Executive',
    'on-site': 'Local Contact',
    'onsite': 'Local Contact',
    'on site': 'Local Contact',
    'local': 'Local Contact',
    'hands & feet': 'Hands & Feet',
    'hands and feet': 'Hands & Feet',
    'portal user': 'Main',
    'primary': 'Main',
    'main': 'Main',
    'repair': 'Technician',
    'tech': 'Technician',
    'technician': 'Technician',
    'unknown': 'Prospect',
    'prospect': 'Prospect',

    # Status Values
    'active': 'Active',
    'in service': 'Active',
    'activated': 'Active',
    'pending': 'Pending',
    'pending activation': 'Pending',
    'in progress': 'In Progress',
    'cancelled': 'Disconnected',
    'canceled': 'Disconnected',
    'disconnected': 'Disconnected',
    'terminated': 'Disconnected',
    'suspended': 'Suspended',
    'on hold': 'Suspended',

    # Industry/Service Types
    'lit building': 'On-Net',
    'on-net': 'On-Net',
    'onnet': 'On-Net',
    'dark fiber': 'Off-Net',
    'off-net': 'Off-Net',
    'offnet': 'Off-Net',
    'metro ethernet': 'Ethernet',
    'ethernet': 'Ethernet',
    'fiber': 'Fiber',
    'dedicated': 'Dedicated',

    # Lead Sources
    'agent': 'Agent',
    'partner': 'Partner',
    'referral': 'Referral',
    'employee referral': 'Employee Referral',
    'customer referral': 'Customer Referral',
    'direct': 'Direct Contact',
    'direct contact': 'Direct Contact',
    'website': 'Website',
    'web': 'Website',
    'inbound': 'Inbound Web',
    'inbound web': 'Inbound Web',
    'trade show': 'Trade Show',
    'event': 'Event',
    'marketing': 'Marketing Campaign',
    'campaign': 'Marketing Campaign',

    # Yes/No variants
    'yes': 'Yes',
    'y': 'Yes',
    'true': 'Yes',
    'no': 'No',
    'n': 'No',
    'false': 'No',
}


# Category-specific mappings for more context-aware matching
CATEGORY_MAPPINGS = {
    'payment_terms': {
        'net30': 'NET 30',
        'net 30': 'NET 30',
        'net15': 'NET 15',
        'net 15': 'NET 15',
        'net45': 'NET 45',
        'net 45': 'NET 45',
        'net60': 'NET 60',
        'net 60': 'NET 60',
        'due on receipt': 'Due Upon Receipt',
        'prepaid': 'Prepaid',
        'credit card': 'Credit Card',
        'ach': 'ACH',
        'wire': 'Wire Transfer',
    },
    'billing_method': {
        'paper': 'Print',
        'electronic': 'Email',
        'email': 'Email',
        'portal': 'Portal',
        'edi': 'EDI',
    },
    'contact_type': {
        'billing': 'Billing',
        'technical': 'Technical',
        'sales': 'Sales',
        'support': 'Support',
        'executive': 'Executive',
        'decision maker': 'Executive',
        'local contact': 'Local Contact',
        'on-site': 'Local Contact',
        'technician': 'Technician',
        'repair': 'Technician',
        'maintenance': 'Maintenance',
    },
    'status': {
        'active': 'Active',
        'inactive': 'Inactive',
        'pending': 'Pending',
        'suspended': 'Suspended',
        'disconnected': 'Disconnected',
        'cancelled': 'Disconnected',
    },
}


def normalize_value(value: str) -> str:
    """Normalize a value for comparison."""
    if not value:
        return ''
    # Lowercase, remove extra whitespace, remove special chars
    normalized = value.lower().strip()
    normalized = re.sub(r'\s+', ' ', normalized)
    return normalized


def calculate_similarity(str1: str, str2: str) -> float:
    """Calculate simple similarity score between two strings."""
    if not str1 or not str2:
        return 0.0

    s1 = normalize_value(str1)
    s2 = normalize_value(str2)

    if s1 == s2:
        return 1.0

    # Check if one contains the other
    if s1 in s2 or s2 in s1:
        return 0.8

    # Check word overlap
    words1 = set(s1.split())
    words2 = set(s2.split())
    if words1 and words2:
        overlap = len(words1 & words2)
        total = len(words1 | words2)
        return overlap / total if total > 0 else 0.0

    return 0.0


def detect_field_category(field_name: str) -> Optional[str]:
    """Detect the category of a field based on its name."""
    field_lower = field_name.lower()

    if 'payment' in field_lower or 'terms' in field_lower:
        return 'payment_terms'
    elif 'billing' in field_lower and ('method' in field_lower or 'type' in field_lower or 'format' in field_lower):
        return 'billing_method'
    elif 'contact' in field_lower and 'type' in field_lower:
        return 'contact_type'
    elif 'status' in field_lower or 'state' in field_lower:
        return 'status'

    return None


def recommend_bbf_value(
    es_value: str,
    bbf_values: List[str],
    es_field: str = '',
    bbf_field: str = ''
) -> Tuple[str, str, float]:
    """
    Recommend the best BBF value for an ES value using semantic understanding.

    Args:
        es_value: The ES picklist value to match
        bbf_values: List of available BBF picklist values
        es_field: ES field name (for context)
        bbf_field: BBF field name (for context)

    Returns:
        Tuple of (recommended_value, reasoning, confidence_score)
    """
    if not es_value or not bbf_values:
        return ('', 'No value or options provided', 0.0)

    es_normalized = normalize_value(es_value)

    # 1. Check exact match (case-insensitive)
    for bbf_val in bbf_values:
        if normalize_value(bbf_val) == es_normalized:
            return (bbf_val, 'Exact match (case-insensitive)', 1.0)

    # 2. Check semantic equivalences
    if es_normalized in SEMANTIC_EQUIVALENCES:
        equiv = SEMANTIC_EQUIVALENCES[es_normalized]
        for bbf_val in bbf_values:
            if normalize_value(bbf_val) == normalize_value(equiv):
                return (bbf_val, f'Semantic equivalent: "{es_value}" means "{equiv}"', 0.95)

    # 3. Check category-specific mappings
    category = detect_field_category(bbf_field) or detect_field_category(es_field)
    if category and category in CATEGORY_MAPPINGS:
        cat_map = CATEGORY_MAPPINGS[category]
        if es_normalized in cat_map:
            target = cat_map[es_normalized]
            for bbf_val in bbf_values:
                if normalize_value(bbf_val) == normalize_value(target):
                    return (bbf_val, f'Category match ({category}): "{es_value}" -> "{target}"', 0.9)

    # 4. Check for partial semantic matches
    # Common patterns: spacing differences (NET30 vs NET 30)
    es_no_space = es_normalized.replace(' ', '')
    for bbf_val in bbf_values:
        bbf_no_space = normalize_value(bbf_val).replace(' ', '')
        if es_no_space == bbf_no_space:
            return (bbf_val, f'Spacing difference: "{es_value}" = "{bbf_val}"', 0.9)

    # 5. Word-based similarity matching
    best_match = None
    best_score = 0.0
    best_reason = ''

    for bbf_val in bbf_values:
        score = calculate_similarity(es_value, bbf_val)
        if score > best_score:
            best_score = score
            best_match = bbf_val
            best_reason = f'Word similarity ({score:.0%})'

    if best_match and best_score >= 0.5:
        return (best_match, best_reason, best_score)

    # 6. Substring matching
    for bbf_val in bbf_values:
        bbf_norm = normalize_value(bbf_val)
        if es_normalized in bbf_norm or bbf_norm in es_normalized:
            return (bbf_val, f'Substring match: "{es_value}" in "{bbf_val}"', 0.6)

    # 7. No good match found - return best guess with low confidence
    if best_match and best_score > 0:
        return (best_match, f'Best available match ({best_score:.0%} similarity) - NEEDS REVIEW', best_score)

    # 8. No match at all
    return ('', f'No semantic match found for "{es_value}" - manual mapping required', 0.0)


def process_picklist_sheet(picklist_df: pd.DataFrame) -> pd.DataFrame:
    """Process the Picklist_Mapping sheet and add recommendations."""
    if picklist_df is None or picklist_df.empty:
        return picklist_df

    # Add recommendation columns if not present
    if 'BBF_Final_Value' not in picklist_df.columns:
        picklist_df['BBF_Final_Value'] = ''
    if 'AI_Confidence' not in picklist_df.columns:
        picklist_df['AI_Confidence'] = ''
    if 'AI_Reasoning' not in picklist_df.columns:
        picklist_df['AI_Reasoning'] = ''

    recommendations_made = 0

    for idx, row in picklist_df.iterrows():
        notes = str(row.get('Notes', ''))

        # Only process rows marked as "No Match"
        if 'No Match' not in notes:
            continue

        # Skip if already has a final value
        current_final = row.get('BBF_Final_Value', '')
        if current_final and not pd.isna(current_final) and str(current_final).strip():
            continue

        es_value = str(row.get('ES_Picklist_Value', ''))
        suggested = str(row.get('Suggested_Mapping', ''))
        es_field = str(row.get('ES_Field', ''))
        bbf_field = str(row.get('BBF_Field', ''))

        # Parse BBF values from Suggested_Mapping (pipe-separated list)
        bbf_values = [v.strip() for v in suggested.split('|') if v.strip()]

        if not bbf_values:
            continue

        # Get recommendation
        recommended, reasoning, confidence = recommend_bbf_value(
            es_value=es_value,
            bbf_values=bbf_values,
            es_field=es_field,
            bbf_field=bbf_field
        )

        if recommended:
            picklist_df.at[idx, 'BBF_Final_Value'] = recommended
            picklist_df.at[idx, 'AI_Confidence'] = f'{confidence:.0%}'
            picklist_df.at[idx, 'AI_Reasoning'] = reasoning
            recommendations_made += 1

    print(f"  Recommendations made: {recommendations_made}")
    return picklist_df


def update_excel_file(excel_path: Path, picklist_df: pd.DataFrame, dry_run: bool = False) -> bool:
    """Update the Excel file with picklist recommendations."""
    if dry_run:
        print(f"\n  [DRY RUN] Would update: {excel_path}")
        print("\n  Recommendations preview:")
        no_match_rows = picklist_df[picklist_df['Notes'].str.contains('No Match', na=False)]
        for _, row in no_match_rows.head(10).iterrows():
            es_val = row.get('ES_Picklist_Value', '')
            bbf_final = row.get('BBF_Final_Value', '')
            reasoning = row.get('AI_Reasoning', '')
            confidence = row.get('AI_Confidence', '')
            if bbf_final and not pd.isna(bbf_final):
                print(f"    '{es_val}' -> '{bbf_final}' ({confidence}) - {reasoning}")
        if len(no_match_rows) > 10:
            print(f"    ... and {len(no_match_rows) - 10} more")
        return True

    try:
        # Load existing workbook to preserve formatting
        wb = load_workbook(excel_path)

        # Get or create Picklist_Mapping sheet
        if 'Picklist_Mapping' in wb.sheetnames:
            ws = wb['Picklist_Mapping']

            # Find column indices
            headers = [cell.value for cell in ws[1]]

            # Add new columns if needed
            bbf_final_col = None
            ai_conf_col = None
            ai_reason_col = None

            for i, h in enumerate(headers, 1):
                if h == 'BBF_Final_Value':
                    bbf_final_col = i
                elif h == 'AI_Confidence':
                    ai_conf_col = i
                elif h == 'AI_Reasoning':
                    ai_reason_col = i

            # Add columns if not present
            next_col = len(headers) + 1
            if bbf_final_col is None:
                bbf_final_col = next_col
                ws.cell(row=1, column=bbf_final_col, value='BBF_Final_Value')
                next_col += 1
            if ai_conf_col is None:
                ai_conf_col = next_col
                ws.cell(row=1, column=ai_conf_col, value='AI_Confidence')
                next_col += 1
            if ai_reason_col is None:
                ai_reason_col = next_col
                ws.cell(row=1, column=ai_reason_col, value='AI_Reasoning')

            # Update cells
            for row_idx, (_, row_data) in enumerate(picklist_df.iterrows(), 2):
                bbf_final = row_data.get('BBF_Final_Value', '')
                ai_conf = row_data.get('AI_Confidence', '')
                ai_reason = row_data.get('AI_Reasoning', '')

                if bbf_final and not pd.isna(bbf_final):
                    ws.cell(row=row_idx, column=bbf_final_col, value=bbf_final)
                if ai_conf and not pd.isna(ai_conf):
                    ws.cell(row=row_idx, column=ai_conf_col, value=ai_conf)
                if ai_reason and not pd.isna(ai_reason):
                    ws.cell(row=row_idx, column=ai_reason_col, value=ai_reason)

            # Style new headers
            header_fill = PatternFill(start_color='FF366092', end_color='FF366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')

            for col in [bbf_final_col, ai_conf_col, ai_reason_col]:
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font

            wb.save(excel_path)
            print(f"  Updated: {excel_path}")
            return True

        else:
            print(f"  Warning: No Picklist_Mapping sheet in {excel_path}")
            return False

    except Exception as e:
        print(f"  Error updating Excel: {e}")
        return False


def process_mapping_file(excel_path: Path, dry_run: bool = False) -> dict:
    """Process a single mapping Excel file."""
    print(f"\nProcessing: {excel_path.name}")

    # Read Picklist_Mapping sheet
    try:
        picklist_df = pd.read_excel(excel_path, sheet_name='Picklist_Mapping')
        print(f"  Total picklist rows: {len(picklist_df)}")
    except Exception as e:
        print(f"  No Picklist_Mapping sheet or error: {e}")
        return {'excel_path': str(excel_path), 'error': str(e)}

    # Count "No Match" rows
    no_match_count = len(picklist_df[picklist_df['Notes'].str.contains('No Match', na=False)])
    print(f"  'No Match' rows: {no_match_count}")

    if no_match_count == 0:
        print("  No recommendations needed")
        return {
            'excel_path': str(excel_path),
            'total_picklist_rows': len(picklist_df),
            'no_match_rows': 0,
            'recommendations_made': 0
        }

    # Process and add recommendations
    updated_df = process_picklist_sheet(picklist_df)

    # Count recommendations made
    recs_made = len(updated_df[
        (updated_df['Notes'].str.contains('No Match', na=False)) &
        (updated_df['BBF_Final_Value'].notna()) &
        (updated_df['BBF_Final_Value'] != '')
    ])

    # Update Excel file
    update_excel_file(excel_path, updated_df, dry_run=dry_run)

    return {
        'excel_path': str(excel_path),
        'total_picklist_rows': len(picklist_df),
        'no_match_rows': no_match_count,
        'recommendations_made': recs_made
    }


def process_all_mappings(dry_run: bool = False) -> List[dict]:
    """Process all mapping Excel files."""
    results = []

    # Find all mapping Excel files
    mapping_files = list(MAPPINGS_DIR.glob('ES_*_to_BBF_*_mapping.xlsx'))
    print(f"Found {len(mapping_files)} mapping files")

    for excel_path in sorted(mapping_files):
        result = process_mapping_file(excel_path, dry_run=dry_run)
        results.append(result)

    return results


def main():
    parser = argparse.ArgumentParser(
        description='AI-powered picklist value recommender for ES to BBF migration'
    )
    parser.add_argument(
        '--input',
        type=Path,
        help='Path to specific mapping Excel file'
    )
    parser.add_argument(
        '--all',
        action='store_true',
        help='Process all mapping files in day-two/mappings/'
    )
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Preview recommendations without updating files'
    )

    args = parser.parse_args()

    if args.all:
        print("=" * 60)
        print("GENERATING PICKLIST RECOMMENDATIONS FOR ALL MAPPING FILES")
        print("=" * 60)
        results = process_all_mappings(dry_run=args.dry_run)

        # Summary
        print("\n" + "=" * 60)
        print("SUMMARY")
        print("=" * 60)
        total_no_match = sum(r.get('no_match_rows', 0) for r in results)
        total_recs = sum(r.get('recommendations_made', 0) for r in results)
        print(f"Files processed: {len(results)}")
        print(f"Total 'No Match' rows: {total_no_match}")
        print(f"Total recommendations made: {total_recs}")

        for r in results:
            if 'error' not in r:
                path = Path(r['excel_path']).name
                print(f"  {path}: {r.get('recommendations_made', 0)}/{r.get('no_match_rows', 0)} recommended")

    elif args.input:
        if not args.input.exists():
            print(f"Error: File not found: {args.input}")
            sys.exit(1)

        result = process_mapping_file(args.input, dry_run=args.dry_run)

        if 'error' in result:
            print(f"Error: {result['error']}")
            sys.exit(1)

        print(f"\nRecommended {result['recommendations_made']} of {result['no_match_rows']} 'No Match' values")

    else:
        parser.print_help()
        print("\nExamples:")
        print("  python recommend_picklist_values.py --all")
        print("  python recommend_picklist_values.py --input day-two/mappings/ES_Account_to_BBF_Account_mapping.xlsx")
        print("  python recommend_picklist_values.py --all --dry-run")


if __name__ == '__main__':
    main()
