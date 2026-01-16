"""
Create ES Account to BBF Account field mapping document with semantic AI matching.
Uses AI-powered semantic understanding to match fields by meaning, not just string similarity.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

# System fields to exclude (standard Salesforce)
SYSTEM_FIELDS = {
    'Id', 'IsDeleted', 'MasterRecordId', 'CreatedDate', 'CreatedById',
    'LastModifiedDate', 'LastModifiedById', 'SystemModstamp', 'LastActivityDate',
    'LastViewedDate', 'LastReferencedDate', 'JigsawContactId', 'Jigsaw',
    'PhotoUrl', 'CleanStatus'
}

# Day 1 migration fields already populated (Account-specific)
DAY1_FIELDS = {
    'Name', 'OwnerId', 'ES_Legacy_ID__c', 'BBF_New_Id__c'
}

# Color definitions (exact hex codes)
COLORS = {
    'header': 'FF366092',  # dark blue
    'high': 'FFC6EFCE',    # light green
    'medium': 'FFFFEB9C',  # light yellow
    'low': 'FFFFC7CE',     # light red/pink
    'none': 'FFFFC7CE',    # light red/pink
}

def semantic_match_fields(bbf_api, bbf_label, bbf_type, es_fields_df):
    """
    Use semantic understanding to match BBF field to ES field.

    Considers:
    - Business/CRM domain knowledge
    - Field purposes based on names AND labels
    - Salesforce conventions
    - Data type compatibility

    Returns: (es_api_name, es_label, es_type, confidence, transformer_needed, notes)
    """

    # Exact API name match
    exact_match = es_fields_df[es_fields_df['Field API Name'] == bbf_api]
    if not exact_match.empty:
        es_row = exact_match.iloc[0]
        transformer = 'Y' if es_row['Field Type'] != bbf_type else 'N'
        return (
            es_row['Field API Name'],
            es_row['Field Label'],
            es_row['Field Type'],
            'High',
            transformer,
            'Exact API name match'
        )

    # Custom field semantic mappings based on business meaning
    semantic_mappings = {
        # BBF API Name -> (ES API Name, Reasoning)
        'Type': ('Type', 'Account classification/type field'),
        'RecordTypeId': (None, 'ES does not use record types for Account'),
        'ParentId': ('ParentId', 'Parent account hierarchy'),
        'Federal_Tax_ID__c': ('FED_TAX_ID_EIN__c', 'Federal tax ID/EIN number'),
        'State_Tax_ID__c': (None, 'No equivalent ES field for state tax ID'),
        'Classification_account__c': ('Type', 'Customer classification - ES uses Type field for similar purpose'),
        'Agent__c': ('Agent__c', 'Agent indicator flag'),
        'proposalmanager__Billing_Contact__c': ('Bill_To_Contact__c', 'Billing contact reference'),
        'proposalmanager__CLLI__c': (None, 'CLLI code - not present in ES'),
        'proposalmanager__Customer_Maintenance_Events_Email__c': (None, 'Customer maintenance email - ES specific field not present'),
        'proposalmanager__Customer_NOC_Email__c': (None, 'Customer NOC email - not in ES'),
        'proposalmanager__Customer_NOC_Phone__c': (None, 'Customer NOC phone - not in ES'),
        'proposalmanager__MSA_On_File__c': ('Master_Service_Agreement_MSA__c', 'MSA on file status'),
        'proposalmanager__Primary_State__c': ('Legally_Organized_Under__c', 'Primary state - ES uses Legally Organized Under'),
        'proposalmanager__Quote_Provider_Type__c': (None, 'Quote provider type - not in ES'),
        'proposalmanager__Quote_Provider__c': (None, 'Quote provider flag - not in ES'),
        'proposalmanager__Secondary_States__c': (None, 'Secondary states - not in ES'),
        'proposalmanager__Tax_Exempt__c': ('Sales_Tax_Exemption__c', 'Tax exemption status'),
        'proposalmanager__Org_Account_For_PMT_Quotes__c': (None, 'PMT quotes org account - not in ES'),
        'Contract_Type__c': ('Contract_Type__c', 'Contract type field'),
        'Company_Type__c': ('Legal_Entity_Required_MSA__c', 'Company legal entity type'),
        'Billing_Contact_First_Name__c': (None, 'Billing contact first name - should use Contact object'),
        'Billing_Contact_Last_Name__c': (None, 'Billing contact last name - should use Contact object'),
        'Billing_Contact_Email__c': ('Email_Address__c', 'Primary email address'),
        'Billing_Contact_Other_Phone__c': (None, 'Billing contact phone - should use Contact object'),
        'Billing_Contact_Work_Phone__c': ('Bill_To_Contact_Phone__c', 'Billing contact work phone'),
        'Engineering_Contact_Email__c': (None, 'Engineering contact - should use Contact object'),
        'Engineering_Contact_First_Name__c': (None, 'Engineering contact first name - should use Contact object'),
        'Engineering_Contact_Last_Name__c': (None, 'Engineering contact last name - should use Contact object'),
        'Engineering_Contact_Other_Phone__c': (None, 'Engineering contact other phone - should use Contact object'),
        'Engineering_Contact_Work_Phone__c': (None, 'Engineering contact work phone - should use Contact object'),
        'Maintenance_Contact_Email__c': (None, 'Maintenance contact email - should use Contact object'),
        'Maintenance_Contact_First_Name__c': (None, 'Maintenance contact first name - should use Contact object'),
        'Maintenance_Contact_Last_Name__c': (None, 'Maintenance contact last name - should use Contact object'),
        'Maintenance_Contact_Other_Phone__c': (None, 'Maintenance contact other phone - should use Contact object'),
        'Maintenance_Contact_Work_Phone__c': (None, 'Maintenance contact work phone - should use Contact object'),
        'New_Customer_Form_Notes__c': (None, 'New customer form notes - not in ES'),
        'Main_Contact_First_Name__c': (None, 'Main contact first name - should use Contact object'),
        'Main_Contact_Email__c': (None, 'Main contact email - should use Contact object'),
        'Main_Contact_Last_Name__c': (None, 'Main contact last name - should use Contact object'),
        'Bill_To__c': (None, 'Bill to/attention - not in ES'),
        'Business_Unit__c': ('OneCommunity_Entity__c', 'Business unit/entity mapping'),
        'BAN__c': (None, 'BAN field - will be populated from BAN object'),
        'Billing_Address_Extraction_Response__c': (None, 'Address API response - BBF specific'),
        'Billing_Address_Extraction_Status__c': (None, 'Address API status - BBF specific'),
        'Billing_Location__c': (None, 'Billing location lookup - BBF specific'),
        'Company__c': (None, 'Company formula field - BBF specific'),
        'Customers_To_Unified_DB_Response__c': (None, 'Unified DB response - BBF specific'),
        'Customers_To_Unified_DB_Status__c': (None, 'Unified DB status - BBF specific'),
        'Invoice_Type__c': (None, 'Invoice delivery type - not in ES'),
        'MNAMember__c': (None, 'MNA member status - BBF specific'),
        'OrgType__c': (None, 'Organization type - BBF specific'),
        'Quote_Provider__c': (None, 'Off-net quote provider flag - BBF specific'),
        'UnifiedCompanyId__c': (None, 'Unified company ID - BBF specific'),
        'businessUnit__c': ('OneCommunity_Entity__c', 'Business unit deprecated field'),
        'custCode__c': ('AccountNumber', 'Customer code/account number'),
        'invoiceEmail__c': (None, 'Invoice email flag - BBF formula field'),
        'invoicePrint__c': (None, 'Invoice print flag - BBF formula field'),
        'salesTeam__c': (None, 'Sales team - BBF specific'),
        'Cust_Code_UniqueKey__c': (None, 'Customer code unique key - BBF specific'),
        'Company_Category__c': (None, 'Company category - BBF specific'),
        'wonOpps__c': ('Number_of_Sold_Opportunities__c', 'Count of won opportunities'),
        'Match_Key__c': (None, 'Match key - BBF specific'),
        'UserIsOwner__c': ('Current_User_is_the_Account_Owner__c', 'Current user is owner formula'),
        'Campaign__c': (None, 'Campaign reference - not in ES'),
        'Owner__c': (None, 'Owner name formula field'),
        'Business_Partner__c': (None, 'Business partner flag - BBF specific'),
        'Owner_Role_Name__c': (None, 'Owner role formula field'),
        'Last_Activity_Date__c': ('LastActivityDate', 'Last activity date'),
        'Credit_Approval_Interval__c': (None, 'Credit approval interval - BBF specific'),
        'Credit_Approval_Received__c': (None, 'Credit approval received date - BBF specific'),
        'Credit_Approval_Sent__c': (None, 'Credit approval sent date - BBF specific'),
        'Master_Agent__c': (None, 'Master agent flag - BBF specific'),
        'Billing_Account_Numbers__c': ('Active_Billing_Account_Count__c', 'Count of billing accounts'),
        'Location__c': (None, 'Location lookup - BBF specific'),
        'oppCount__c': (None, 'Opportunity count formula'),
        'New_Transition__c': (None, 'Verified account flag - BBF specific'),
        'rh2__testCurrency__c': (None, 'Test currency field - ignore'),
        'Owner_Email__c': (None, 'Owner email formula field'),
        'Owner_Phone__c': (None, 'Owner phone formula field'),
        'Account_Owner_Manager__c': (None, 'Account owner manager formula'),
        'Num_Billing_Contacts__c': ('Contact_Count__c', 'Number of contacts'),
        'Num_BANs_with_Billing_Contacts__c': (None, 'BANs with billing contacts - BBF specific'),
        'Num_Active_Services_With_Charges__c': ('SOF_Count__c', 'Number of active orders/services'),
        'Num_Inactive_Services__c': (None, 'Number of inactive services - BBF specific'),
        'Num_Active_Services_Without_Charges__c': (None, 'Active services without charges - BBF specific'),
        'IT_Team__c': (None, 'IT team names formula'),
        'SOLs_for_this_Account__c': (None, 'SOLs formula field - BBF specific'),
        'Qualified_SOLs_for_this_Account__c': (None, 'Qualified SOLs formula - BBF specific'),
        'Not_Yet_Qualified_SOLs_for_this_Account__c': (None, 'Not qualified SOLs formula - BBF specific'),
        'Services_for_this_Account__c': (None, 'Services formula - BBF specific'),
        'Active_Services_for_this_Account__c': (None, 'Active services formula - BBF specific'),
        'Inactive_Services_for_this_Account__c': (None, 'Inactive services formula - BBF specific'),
        'Blankspot__c': (None, 'Blank spot field - ignore'),
        'Canceled_SOLs_for_this_Account__c': (None, 'Canceled SOLs formula - BBF specific'),
        'Full_Company_Name__c': ('Company_Legal_Name__c', 'Full legal company name'),
        'Full_Company_Address__c': ('Corporate_Address_Full__c', 'Full corporate address formula'),
        'Finalize_Master_Service_Agreement_Name__c': (None, 'MSA finalize formula - BBF specific'),
        'Finalize_Right_of_Entry_Agreement_Name__c': (None, 'ROE agreement formula - BBF specific'),
        'Finalize_Service_Agreement_Name__c': (None, 'Service agreement formula - BBF specific'),
        'Finalize_NDA_Name__c': (None, 'NDA finalize formula - BBF specific'),
        'TASKRAY__trCustomerOnboardingScore__c': (None, 'TaskRay onboarding score - BBF specific'),
        'TASKRAY__trDesiredOutcomes__c': (None, 'TaskRay desired outcomes - BBF specific'),
        'Strategic_Account__c': (None, 'Strategic account flag - BBF specific'),
        'ActOn__AccountId18Char__c': (None, 'ActOn 18 char ID formula'),
        'Agent_Type__c': (None, 'Agent type formula - BBF specific'),
        'BAN_Business_Unit__c': (None, 'BAN business unit - BBF specific'),
        'Total_Contacts__c': ('Contact_Count__c', 'Total contacts rollup'),
        'Voice_Service_Type__c': (None, 'Voice service type - BBF specific'),
        'Provider_Manufacturer__c': (None, 'Provider/manufacturer - BBF specific'),
        'Key_System_Unit_KSU__c': (None, 'Key system unit - BBF specific'),
        'circuitvision__CV_Customer_Id__c': (None, 'CircuitVision customer ID - BBF specific'),
        'Protected_Account__c': (None, 'Protected account flag - BBF specific'),
        'Lead_Source__c': ('AccountSource', 'Lead/account source field'),
    }

    # Check semantic mapping
    if bbf_api in semantic_mappings:
        es_api, reason = semantic_mappings[bbf_api]
        if es_api is None:
            return (None, None, None, 'None', 'N', reason)

        es_match = es_fields_df[es_fields_df['Field API Name'] == es_api]
        if not es_match.empty:
            es_row = es_match.iloc[0]
            # Determine confidence based on data type compatibility
            if es_row['Field Type'] == bbf_type:
                confidence = 'High'
                transformer = 'N'
            else:
                confidence = 'Medium'
                transformer = 'Y'
            return (
                es_row['Field API Name'],
                es_row['Field Label'],
                es_row['Field Type'],
                confidence,
                transformer,
                reason
            )

    # Label similarity search
    bbf_label_lower = bbf_label.lower()
    for _, es_row in es_fields_df.iterrows():
        es_label_lower = es_row['Field Label'].lower()
        if bbf_label_lower == es_label_lower:
            transformer = 'Y' if es_row['Field Type'] != bbf_type else 'N'
            return (
                es_row['Field API Name'],
                es_row['Field Label'],
                es_row['Field Type'],
                'Medium',
                transformer,
                'Exact label match but different API names'
            )

    # No match found
    return (None, None, None, 'None', 'N', 'No equivalent ES field identified')


def match_picklist_values(es_field, es_value, bbf_field, es_picklist_df, bbf_picklist_df):
    """
    Match ES picklist value to BBF picklist value.

    Returns: (suggested_mapping, notes, match_type)
    match_type: 'exact', 'close', 'none'
    """
    # Get all BBF values for this field
    bbf_values = bbf_picklist_df[
        bbf_picklist_df['Field API Name'] == bbf_field
    ]['Picklist Value'].tolist()

    if not bbf_values:
        return ('', 'BBF field has no picklist values', 'none')

    # Exact match
    if es_value in bbf_values:
        return (es_value, 'Exact Match', 'exact')

    # Case-insensitive match
    es_value_lower = es_value.lower()
    for bbf_val in bbf_values:
        if bbf_val.lower() == es_value_lower:
            return (bbf_val, 'Exact Match (case difference)', 'exact')

    # Semantic/close matches (domain knowledge)
    value_mappings = {
        # ES Type -> BBF Type
        'Customer': 'Enterprise',
        'Sales Prospect': 'Enterprise',
        'Customer (MDU Residential)': 'Enterprise',
        'Customer (Wholesaler Carrier)': 'Wholesale',
        # ES Industry -> BBF Industry
        'Arts': 'Business Services',
        'Educational': 'Education',
        'Health care': 'Healthcare',
        'Not For Profit': 'Organizations',
        'Professional Services': 'Business Services',
        'Research & Technology': 'Technology',
        # ES AccountSource -> BBF AccountSource
        'CoStar': 'ZoomInfo',
        'Directly Engaged': 'Direct Contact',
        'Inbound Phone Call': 'Inbound Web',
        'Network Map Leads': 'Marketing Campaign',
        'Sales Intern Lead Source': 'Direct Contact',
        'Tidio - Website Chatbot': 'Website',
        'Unsolicited Prospecting': 'Direct Contact',
        'Vendor Referral Program': 'Vendor',
        'Zoominfo': 'ZoomInfo',
    }

    if es_value in value_mappings and value_mappings[es_value] in bbf_values:
        return (value_mappings[es_value], 'Close Match', 'close')

    # No match - return all BBF values
    all_values = ' | '.join(bbf_values)
    return (all_values, 'No Match - Select from list', 'none')


def create_mapping_excel(es_fields_csv, bbf_fields_csv, es_picklist_csv, bbf_picklist_csv, output_file):
    """Create the field mapping Excel workbook."""

    # Load data
    es_fields = pd.read_csv(es_fields_csv)
    bbf_fields = pd.read_csv(bbf_fields_csv)
    es_picklist = pd.read_csv(es_picklist_csv)
    bbf_picklist = pd.read_csv(bbf_picklist_csv)

    # Filter out system fields and Day 1 fields
    excluded_fields = SYSTEM_FIELDS | DAY1_FIELDS
    bbf_fields = bbf_fields[~bbf_fields['Field API Name'].isin(excluded_fields)]

    # Create field mapping data
    field_mapping_data = []
    picklist_mapping_data = []

    for _, bbf_row in bbf_fields.iterrows():
        bbf_api = bbf_row['Field API Name']
        bbf_label = bbf_row['Field Label']
        bbf_type = bbf_row['Field Type']
        bbf_required = 'No' if bbf_row['Is Nillable'] else 'Yes'

        # Match to ES field
        es_api, es_label, es_type, confidence, transformer, notes = semantic_match_fields(
            bbf_api, bbf_label, bbf_type, es_fields
        )

        field_mapping_data.append({
            'BBF_Field_API_Name': bbf_api,
            'BBF_Field_Label': bbf_label,
            'BBF_Data_Type': bbf_type,
            'BBF_Is_Required': bbf_required,
            'ES_Field_API_Name': es_api if es_api else '',
            'ES_Field_Label': es_label if es_label else '',
            'ES_Data_Type': es_type if es_type else '',
            'Match_Confidence': confidence,
            'Transformer_Needed': transformer,
            'Notes': notes
        })

        # Create picklist mappings if both are picklist fields
        if bbf_type == 'picklist' and es_api and es_type == 'picklist':
            es_values = es_picklist[
                es_picklist['Field API Name'] == es_api
            ]['Picklist Value'].unique()

            for es_value in es_values:
                suggested, match_notes, match_type = match_picklist_values(
                    es_api, es_value, bbf_api, es_picklist, bbf_picklist
                )

                picklist_mapping_data.append({
                    'ES_Field': es_api,
                    'ES_Picklist_Value': es_value,
                    'BBF_Field': bbf_api,
                    'Suggested_Mapping': suggested,
                    'Notes': match_notes,
                    'BBF_Final_Value': '',
                    '_match_type': match_type
                })

    # Create DataFrames
    field_df = pd.DataFrame(field_mapping_data)
    picklist_df = pd.DataFrame(picklist_mapping_data)

    # Create Excel file
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: Field_Mapping
    ws1 = wb.create_sheet('Field_Mapping')
    for r_idx, row in enumerate(dataframe_to_rows(field_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)

            # Header row formatting
            if r_idx == 1:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                # Data row formatting based on confidence
                confidence = field_df.iloc[r_idx - 2]['Match_Confidence']
                if confidence == 'High':
                    cell.fill = PatternFill(start_color=COLORS['high'], end_color=COLORS['high'], fill_type='solid')
                elif confidence == 'Medium':
                    cell.fill = PatternFill(start_color=COLORS['medium'], end_color=COLORS['medium'], fill_type='solid')
                elif confidence in ['Low', 'None']:
                    cell.fill = PatternFill(start_color=COLORS['low'], end_color=COLORS['low'], fill_type='solid')

                cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Auto-size columns
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws1.column_dimensions[column_letter].width = adjusted_width

    # Sheet 2: Picklist_Mapping
    ws2 = wb.create_sheet('Picklist_Mapping')
    picklist_display_df = picklist_df.drop(columns=['_match_type'])

    for r_idx, row in enumerate(dataframe_to_rows(picklist_display_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)

            # Header row formatting
            if r_idx == 1:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                # Data row formatting based on match type
                match_type = picklist_df.iloc[r_idx - 2]['_match_type']
                if match_type == 'exact':
                    cell.fill = PatternFill(start_color=COLORS['high'], end_color=COLORS['high'], fill_type='solid')
                elif match_type == 'close':
                    cell.fill = PatternFill(start_color=COLORS['medium'], end_color=COLORS['medium'], fill_type='solid')
                elif match_type == 'none':
                    cell.fill = PatternFill(start_color=COLORS['low'], end_color=COLORS['low'], fill_type='solid')

                cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Auto-size columns
    for column in ws2.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws2.column_dimensions[column_letter].width = adjusted_width

    # Save workbook
    wb.save(output_file)
    print(f"✅ Mapping Excel created: {output_file}")

    # Print statistics
    high_count = len(field_df[field_df['Match_Confidence'] == 'High'])
    medium_count = len(field_df[field_df['Match_Confidence'] == 'Medium'])
    low_none_count = len(field_df[field_df['Match_Confidence'].isin(['Low', 'None'])])
    transformer_count = len(field_df[field_df['Transformer_Needed'] == 'Y'])
    picklist_fields = len(picklist_df['ES_Field'].unique())

    print(f"\n📊 Mapping Statistics:")
    print(f"   BBF Fields: {len(bbf_fields)}")
    print(f"   ES Fields: {len(es_fields)}")
    print(f"   High Confidence Matches: {high_count}")
    print(f"   Medium Confidence Matches: {medium_count}")
    print(f"   Low/No Match (Need Review): {low_none_count}")
    print(f"   Picklist Fields Mapped: {picklist_fields}")
    print(f"   Transformers Needed: {transformer_count}")


if __name__ == '__main__':
    # File paths
    base_dir = Path(__file__).parent.parent
    exports_dir = base_dir / 'exports'
    mappings_dir = base_dir / 'mappings'

    es_fields_csv = exports_dir / 'es_Account_fields_with_picklists.csv'
    bbf_fields_csv = exports_dir / 'bbf_Account_fields_with_picklists.csv'
    es_picklist_csv = exports_dir / 'es_Account_picklist_values.csv'
    bbf_picklist_csv = exports_dir / 'bbf_Account_picklist_values.csv'

    output_file = mappings_dir / 'ES_Account_to_BBF_Account_mapping.xlsx'

    print("🚀 Creating ES Account to BBF Account field mapping...")
    print(f"   ES Fields: {es_fields_csv}")
    print(f"   BBF Fields: {bbf_fields_csv}")
    print(f"   Output: {output_file}")
    print()

    create_mapping_excel(
        es_fields_csv,
        bbf_fields_csv,
        es_picklist_csv,
        bbf_picklist_csv,
        output_file
    )

    print("\n✨ Mapping complete!")
