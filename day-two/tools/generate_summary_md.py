#!/usr/bin/env python3
"""
Generate summary markdown from field mapping Excel files.

Shared utility that reads mapping Excel files and produces a structured
markdown summary with statistics, field breakdowns, and action items.

Usage:
    python generate_summary_md.py --excel mapping.xlsx --output summary.md
    python generate_summary_md.py --excel mapping.xlsx  # prints to stdout
    python generate_summary_md.py --all  # generates summaries for all mapping files
"""

import argparse
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


def read_excel_mapping(excel_path):
    """
    Read field mapping and picklist mapping from Excel file.

    Returns:
        dict with 'field_mappings' and 'picklist_mappings' lists
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    result = {
        'field_mappings': [],
        'picklist_mappings': [],
        'source_file': str(excel_path)
    }

    # Read Field_Mapping sheet
    if 'Field_Mapping' in wb.sheetnames:
        ws = wb['Field_Mapping']
        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # Skip empty rows
                continue
            mapping = dict(zip(headers, row))
            result['field_mappings'].append(mapping)

    # Read Picklist_Mapping sheet
    if 'Picklist_Mapping' in wb.sheetnames:
        ws = wb['Picklist_Mapping']
        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # Skip empty rows
                continue
            mapping = dict(zip(headers, row))
            result['picklist_mappings'].append(mapping)

    wb.close()
    return result


def calculate_statistics(data):
    """Calculate summary statistics from mapping data."""
    field_mappings = data['field_mappings']
    picklist_mappings = data['picklist_mappings']

    stats = {
        'total_fields': len(field_mappings),
        'high_confidence': 0,
        'medium_confidence': 0,
        'low_confidence': 0,
        'no_match': 0,
        'transformers_needed': 0,
        'include_yes': 0,
        'include_no': 0,
        'include_tbd': 0,
        'total_picklist_values': len(picklist_mappings),
        'picklist_exact': 0,
        'picklist_close': 0,
        'picklist_no_match': 0,
        'picklist_fields': set(),
    }

    for mapping in field_mappings:
        confidence = mapping.get('Match_Confidence', 'None')
        if confidence == 'High':
            stats['high_confidence'] += 1
        elif confidence == 'Medium':
            stats['medium_confidence'] += 1
        elif confidence == 'Low':
            stats['low_confidence'] += 1
        else:
            stats['no_match'] += 1

        if mapping.get('Transformer_Needed') == 'Y':
            stats['transformers_needed'] += 1

        include = mapping.get('Include_in_Migration', 'TBD')
        if include == 'Yes':
            stats['include_yes'] += 1
        elif include == 'No':
            stats['include_no'] += 1
        else:
            stats['include_tbd'] += 1

    for mapping in picklist_mappings:
        notes = mapping.get('Notes', '')
        if 'Exact Match' in notes:
            stats['picklist_exact'] += 1
        elif 'Close Match' in notes:
            stats['picklist_close'] += 1
        else:
            stats['picklist_no_match'] += 1

        es_field = mapping.get('ES_Field', '')
        if es_field:
            stats['picklist_fields'].add(es_field)

    stats['picklist_field_count'] = len(stats['picklist_fields'])

    return stats


def generate_markdown(data, stats, object_name=None):
    """Generate markdown summary from mapping data and statistics."""

    # Extract object name from filename if not provided
    if object_name is None:
        source = Path(data['source_file']).stem
        # Parse ES_Account_to_BBF_Account_mapping -> Account
        parts = source.replace('_mapping', '').split('_to_')
        if len(parts) == 2:
            es_obj = parts[0].replace('ES_', '')
            bbf_obj = parts[1].replace('BBF_', '')
            object_name = f"{es_obj} -> {bbf_obj}"
        else:
            object_name = source

    lines = []
    lines.append(f"# Field Mapping Summary: {object_name}")
    lines.append("")
    lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"Source: `{Path(data['source_file']).name}`")
    lines.append("")

    # Overall Statistics
    lines.append("## Overview")
    lines.append("")
    lines.append(f"| Metric | Count |")
    lines.append(f"|--------|-------|")
    lines.append(f"| Total BBF Fields (for enrichment) | {stats['total_fields']} |")
    lines.append(f"| Fields with ES Match | {stats['high_confidence'] + stats['medium_confidence']} |")
    lines.append(f"| Fields without ES Match | {stats['no_match'] + stats['low_confidence']} |")
    lines.append(f"| Transformers Needed | {stats['transformers_needed']} |")
    lines.append("")

    # Match Confidence Breakdown
    lines.append("## Match Confidence")
    lines.append("")
    lines.append(f"| Confidence | Count | % |")
    lines.append(f"|------------|-------|---|")
    total = stats['total_fields'] or 1
    lines.append(f"| High | {stats['high_confidence']} | {stats['high_confidence']*100//total}% |")
    lines.append(f"| Medium | {stats['medium_confidence']} | {stats['medium_confidence']*100//total}% |")
    lines.append(f"| Low | {stats['low_confidence']} | {stats['low_confidence']*100//total}% |")
    lines.append(f"| None | {stats['no_match']} | {stats['no_match']*100//total}% |")
    lines.append("")

    # Migration Decision Status
    lines.append("## Migration Decision Status")
    lines.append("")
    lines.append(f"| Decision | Count |")
    lines.append(f"|----------|-------|")
    lines.append(f"| Include: Yes | {stats['include_yes']} |")
    lines.append(f"| Include: No | {stats['include_no']} |")
    lines.append(f"| Include: TBD | {stats['include_tbd']} |")
    lines.append("")

    # Picklist Statistics (if any)
    if stats['total_picklist_values'] > 0:
        lines.append("## Picklist Value Mappings")
        lines.append("")
        lines.append(f"| Metric | Count |")
        lines.append(f"|--------|-------|")
        lines.append(f"| Picklist Fields | {stats['picklist_field_count']} |")
        lines.append(f"| Total Values | {stats['total_picklist_values']} |")
        lines.append(f"| Exact Match | {stats['picklist_exact']} |")
        lines.append(f"| Close Match | {stats['picklist_close']} |")
        lines.append(f"| No Match | {stats['picklist_no_match']} |")
        lines.append("")

    # Action Items
    lines.append("## Action Items")
    lines.append("")

    # Fields needing transformers
    transformer_fields = [
        m for m in data['field_mappings']
        if m.get('Transformer_Needed') == 'Y' and m.get('Match_Confidence') in ['High', 'Medium']
    ]
    if transformer_fields:
        lines.append(f"### Fields Needing Transformers ({len(transformer_fields)})")
        lines.append("")
        for m in transformer_fields:
            lines.append(f"- **{m.get('BBF_Field_API_Name')}** <- {m.get('ES_Field_API_Name')}: {m.get('BBF_Data_Type')} <- {m.get('ES_Data_Type')}")
        lines.append("")

    # TBD decisions
    tbd_fields = [
        m for m in data['field_mappings']
        if m.get('Include_in_Migration') == 'TBD'
    ]
    if tbd_fields:
        lines.append(f"### Fields Pending Decision ({len(tbd_fields)})")
        lines.append("")
        lines.append("Fields with `Include_in_Migration = TBD` that need business review:")
        lines.append("")
        # Group by confidence
        tbd_medium = [m for m in tbd_fields if m.get('Match_Confidence') == 'Medium']
        tbd_other = [m for m in tbd_fields if m.get('Match_Confidence') not in ['Medium', 'High']]

        if tbd_medium:
            lines.append(f"**Medium Confidence ({len(tbd_medium)}):**")
            for m in tbd_medium[:10]:  # Limit to 10
                lines.append(f"- {m.get('BBF_Field_API_Name')} <- {m.get('ES_Field_API_Name')}")
            if len(tbd_medium) > 10:
                lines.append(f"- ... and {len(tbd_medium) - 10} more")
            lines.append("")

        if tbd_other:
            lines.append(f"**Low/No Confidence ({len(tbd_other)}):**")
            for m in tbd_other[:10]:  # Limit to 10
                lines.append(f"- {m.get('BBF_Field_API_Name')}: {m.get('Notes', 'No match')}")
            if len(tbd_other) > 10:
                lines.append(f"- ... and {len(tbd_other) - 10} more")
            lines.append("")

    # Picklist values needing review
    picklist_no_match = [
        m for m in data['picklist_mappings']
        if 'No Match' in m.get('Notes', '')
    ]
    if picklist_no_match:
        lines.append(f"### Picklist Values Needing Mapping ({len(picklist_no_match)})")
        lines.append("")
        # Group by field
        by_field = {}
        for m in picklist_no_match:
            field = m.get('ES_Field', 'Unknown')
            if field not in by_field:
                by_field[field] = []
            by_field[field].append(m.get('ES_Picklist_Value', ''))

        for field, values in by_field.items():
            lines.append(f"**{field}:**")
            for v in values[:5]:
                lines.append(f"- `{v}`")
            if len(values) > 5:
                lines.append(f"- ... and {len(values) - 5} more")
            lines.append("")

    # High Confidence Mappings Summary
    high_conf = [m for m in data['field_mappings'] if m.get('Match_Confidence') == 'High']
    if high_conf:
        lines.append(f"## High Confidence Mappings ({len(high_conf)})")
        lines.append("")
        lines.append("| BBF Field | ES Field | Type |")
        lines.append("|-----------|----------|------|")
        for m in high_conf[:20]:  # Limit to 20
            lines.append(f"| {m.get('BBF_Field_API_Name')} | {m.get('ES_Field_API_Name')} | {m.get('BBF_Data_Type')} |")
        if len(high_conf) > 20:
            lines.append(f"| ... | {len(high_conf) - 20} more | ... |")
        lines.append("")

    return '\n'.join(lines)


def process_single_file(excel_path, output_path=None):
    """Process a single Excel mapping file."""
    excel_path = Path(excel_path)

    if not excel_path.exists():
        print(f"Error: File not found: {excel_path}")
        return False

    print(f"Reading: {excel_path}")
    data = read_excel_mapping(excel_path)
    stats = calculate_statistics(data)
    markdown = generate_markdown(data, stats)

    if output_path:
        output_path = Path(output_path)
        output_path.write_text(markdown, encoding='utf-8')
        print(f"Written: {output_path}")
    else:
        print(markdown)

    return True


def process_all_mappings(mappings_dir=None):
    """Process all mapping Excel files in the mappings directory."""
    if mappings_dir is None:
        mappings_dir = Path(__file__).parent.parent / 'mappings'
    else:
        mappings_dir = Path(mappings_dir)

    if not mappings_dir.exists():
        print(f"Error: Mappings directory not found: {mappings_dir}")
        return False

    excel_files = list(mappings_dir.glob('*_mapping.xlsx'))

    if not excel_files:
        print(f"No mapping Excel files found in: {mappings_dir}")
        return False

    print(f"Found {len(excel_files)} mapping files")
    print()

    summaries_dir = mappings_dir / 'summaries'
    summaries_dir.mkdir(exist_ok=True)

    for excel_path in sorted(excel_files):
        output_name = excel_path.stem + '_summary.md'
        output_path = summaries_dir / output_name

        try:
            data = read_excel_mapping(excel_path)
            stats = calculate_statistics(data)
            markdown = generate_markdown(data, stats)
            output_path.write_text(markdown, encoding='utf-8')
            print(f"  {excel_path.name} -> {output_name}")
        except Exception as e:
            print(f"  Error processing {excel_path.name}: {e}")

    print()
    print(f"Summaries written to: {summaries_dir}")
    return True


def main():
    parser = argparse.ArgumentParser(
        description='Generate summary markdown from field mapping Excel files'
    )
    parser.add_argument(
        '--excel',
        help='Path to mapping Excel file'
    )
    parser.add_argument(
        '--output', '-o',
        help='Output markdown file path (prints to stdout if not specified)'
    )
    parser.add_argument(
        '--all',
        action='store_true',
        help='Process all mapping files in the mappings directory'
    )
    parser.add_argument(
        '--mappings-dir',
        help='Directory containing mapping Excel files (for --all)'
    )

    args = parser.parse_args()

    if args.all:
        success = process_all_mappings(args.mappings_dir)
    elif args.excel:
        success = process_single_file(args.excel, args.output)
    else:
        parser.print_help()
        print("\nError: Either --excel or --all is required")
        sys.exit(1)

    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
