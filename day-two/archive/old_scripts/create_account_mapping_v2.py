#!/usr/bin/env python3
"""
Create Account → Account field mapping workbook with updated picklist format.
Excludes system fields and Day 1 fields before mapping.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import re

# System fields to exclude
SYSTEM_FIELDS = {
    'Id', 'IsDeleted', 'MasterRecordId', 'CreatedDate', 'CreatedById',
    'LastModifiedDate', 'LastModifiedById', 'SystemModstamp',
    'LastActivityDate', 'LastViewedDate', 'LastReferencedDate',
    'JigsawContactId', 'Jigsaw', 'PhotoUrl', 'CleanStatus'
}

# Day 1 fields to exclude
DAY1_FIELDS = {
    'Name', 'OwnerId', 'ES_Legacy_ID__c', 'BBF_New_Id__c'
}

# Combine exclusions
EXCLUDED_FIELDS = SYSTEM_FIELDS | DAY1_FIELDS

def normalize_api_name(name):
    """Normalize API name for matching (remove __c, lowercase)."""
    return name.replace('__c', '').lower()

def normalize_label(label):
    """Normalize label for matching (lowercase, remove special chars)."""
    return re.sub(r'[^a-z0-9]', '', label.lower())

def calculate_match_confidence(bbf_field, bbf_label, es_field, es_label, bbf_type, es_type):
    """Calculate match confidence level."""
    # Exact API name match
    if bbf_field == es_field:
        return 'High' if bbf_type == es_type else 'Medium'

    # Normalized API name match
    if normalize_api_name(bbf_field) == normalize_api_name(es_field):
        return 'High' if bbf_type == es_type else 'Medium'

    # Exact label match
    if bbf_label == es_label:
        return 'Medium' if bbf_type == es_type else 'Low'

    # Normalized label match
    if normalize_label(bbf_label) == normalize_label(es_label):
        return 'Medium' if bbf_type == es_type else 'Low'

    # Similar API names (substring match)
    norm_bbf = normalize_api_name(bbf_field)
    norm_es = normalize_api_name(es_field)
    if norm_bbf in norm_es or norm_es in norm_bbf:
        return 'Medium'

    # Similar labels (substring match)
    norm_bbf_label = normalize_label(bbf_label)
    norm_es_label = normalize_label(es_label)
    if len(norm_bbf_label) > 5 and len(norm_es_label) > 5:
        if norm_bbf_label in norm_es_label or norm_es_label in norm_bbf_label:
            return 'Low'

    return 'None'

def needs_transformer(bbf_type, es_type, bbf_length, es_length, match_confidence):
    """Determine if a transformer is needed."""
    # Type mismatch
    if bbf_type != es_type:
        return 'Y'

    # Picklist fields always need review
    if bbf_type == 'picklist' or bbf_type == 'multipicklist':
        return 'Y'

    # String length mismatch (ES longer)
    if bbf_type == 'string' and es_length > bbf_length:
        return 'Y'

    # Low or no match confidence
    if match_confidence in ['Low', 'None']:
        return 'Y'

    return 'N'

def find_best_match(bbf_field, bbf_label, bbf_type, es_fields_df):
    """Find best matching ES field for a BBF field."""
    best_match = None
    best_confidence = 'None'

    confidence_order = {'High': 4, 'Medium': 3, 'Low': 2, 'None': 1}

    for _, es_row in es_fields_df.iterrows():
        es_field = es_row['Field API Name']
        es_label = es_row['Field Label']
        es_type = es_row['Field Type']

        confidence = calculate_match_confidence(
            bbf_field, bbf_label, es_field, es_label, bbf_type, es_type
        )

        if confidence_order.get(confidence, 0) > confidence_order.get(best_confidence, 0):
            best_match = es_row
            best_confidence = confidence

    return best_match, best_confidence

def create_field_mapping():
    """Create field mapping sheet."""
    print("Loading BBF and ES field metadata...")
    bbf_df = pd.read_csv('day-two/exports/bbf_Account_fields_with_picklists.csv')
    es_df = pd.read_csv('day-two/exports/es_Account_fields_with_picklists.csv')

    # Filter out excluded fields
    print(f"Total BBF fields before filtering: {len(bbf_df)}")
    bbf_df = bbf_df[~bbf_df['Field API Name'].isin(EXCLUDED_FIELDS)]
    print(f"Total BBF fields after filtering: {len(bbf_df)}")

    print(f"Total ES fields: {len(es_df)}")

    # Build mapping
    mappings = []
    match_counts = {'High': 0, 'Medium': 0, 'Low': 0, 'None': 0}
    transformer_count = 0

    print("\nMapping BBF fields to ES fields...")
    for _, bbf_row in bbf_df.iterrows():
        bbf_field = bbf_row['Field API Name']
        bbf_label = bbf_row['Field Label']
        bbf_type = bbf_row['Field Type']
        bbf_required = not bbf_row['Is Nillable']
        bbf_length = bbf_row['Length']

        # Find best ES match
        es_match, confidence = find_best_match(bbf_field, bbf_label, bbf_type, es_df)

        if es_match is not None:
            es_field = es_match['Field API Name']
            es_label = es_match['Field Label']
            es_type = es_match['Field Type']
            es_length = es_match['Length']

            transformer = needs_transformer(bbf_type, es_type, bbf_length, es_length, confidence)

            # Generate notes
            notes = []
            if confidence == 'High':
                notes.append("Strong match - API name/label similar with compatible types")
            elif confidence == 'Medium':
                notes.append("Moderate match - Similar API name or label")
            elif confidence == 'Low':
                notes.append("Weak match - Semantic similarity only")
            else:
                notes.append("No obvious match found")

            if bbf_type != es_type:
                notes.append(f"Type mismatch: ES {es_type} → BBF {bbf_type}")
            if bbf_type == 'string' and es_length > bbf_length:
                notes.append(f"ES field longer ({es_length}) than BBF ({bbf_length}) - may truncate")
            if bbf_type in ['picklist', 'multipicklist']:
                notes.append("Picklist values need mapping review")
        else:
            es_field = ''
            es_label = ''
            es_type = ''
            transformer = 'N'
            notes = ["No ES field found - may need business rule or default value"]

        mappings.append({
            'BBF_Field_API_Name': bbf_field,
            'BBF_Field_Label': bbf_label,
            'BBF_Data_Type': bbf_type,
            'BBF_Is_Required': 'Yes' if bbf_required else 'No',
            'ES_Field_API_Name': es_field,
            'ES_Field_Label': es_label,
            'ES_Data_Type': es_type,
            'Match_Confidence': confidence,
            'Transformer_Needed': transformer,
            'Notes': '; '.join(notes)
        })

        match_counts[confidence] += 1
        if transformer == 'Y':
            transformer_count += 1

    print(f"\nMatch Statistics:")
    print(f"  High Confidence: {match_counts['High']}")
    print(f"  Medium Confidence: {match_counts['Medium']}")
    print(f"  Low Confidence: {match_counts['Low']}")
    print(f"  No Match: {match_counts['None']}")
    print(f"  Transformers Needed: {transformer_count}")

    return pd.DataFrame(mappings), match_counts, transformer_count

def create_picklist_mapping():
    """Create picklist mapping sheet with new column order."""
    print("\nLoading picklist values...")
    bbf_picklist_df = pd.read_csv('day-two/exports/bbf_Account_picklist_values.csv')
    es_picklist_df = pd.read_csv('day-two/exports/es_Account_picklist_values.csv')

    # Filter out excluded fields
    bbf_picklist_df = bbf_picklist_df[~bbf_picklist_df['Field API Name'].isin(EXCLUDED_FIELDS)]

    picklist_mappings = []
    exact_match_count = 0
    suggested_match_count = 0
    no_match_count = 0

    print("Mapping picklist values...")

    # Group by BBF field
    for bbf_field in bbf_picklist_df['Field API Name'].unique():
        bbf_field_values = bbf_picklist_df[bbf_picklist_df['Field API Name'] == bbf_field]

        # Try to find corresponding ES field
        # First try exact match
        es_field_values = es_picklist_df[es_picklist_df['Field API Name'] == bbf_field]

        # If no exact match, try normalized match
        if len(es_field_values) == 0:
            norm_bbf = normalize_api_name(bbf_field)
            for es_field in es_picklist_df['Field API Name'].unique():
                if normalize_api_name(es_field) == norm_bbf:
                    es_field_values = es_picklist_df[es_picklist_df['Field API Name'] == es_field]
                    break

        if len(es_field_values) > 0:
            # We have an ES field match
            es_field_name = es_field_values.iloc[0]['Field API Name']
            es_values = set(es_field_values['Picklist Value'].unique())
            bbf_values_list = bbf_field_values['Picklist Value'].tolist()

            # Map each ES value to BBF
            for es_value in sorted(es_values):
                # Try exact match first
                if es_value in bbf_values_list:
                    suggested_mapping = es_value
                    notes = "Exact Match"
                    exact_match_count += 1
                else:
                    # Try case-insensitive match
                    es_lower = es_value.lower()
                    bbf_match = None
                    for bbf_val in bbf_values_list:
                        if bbf_val.lower() == es_lower:
                            bbf_match = bbf_val
                            break

                    if bbf_match:
                        suggested_mapping = bbf_match
                        notes = "Close Match"
                        suggested_match_count += 1
                    else:
                        # No match - show all valid BBF values
                        suggested_mapping = ' | '.join(sorted(bbf_values_list))
                        notes = "No Match - Select from list"
                        no_match_count += 1

                picklist_mappings.append({
                    'ES_Field': es_field_name,
                    'ES_Picklist_Value': es_value,
                    'BBF_Field': bbf_field,
                    'Suggested_Mapping': suggested_mapping,
                    'Notes': notes,
                    'BBF_Final_Value': ''  # User decision column
                })

    print(f"\nPicklist Mapping Statistics:")
    print(f"  Exact Matches: {exact_match_count}")
    print(f"  Suggested Matches: {suggested_match_count}")
    print(f"  No Match (User Review): {no_match_count}")
    print(f"  Total Picklist Values Mapped: {len(picklist_mappings)}")

    return pd.DataFrame(picklist_mappings), exact_match_count, suggested_match_count, no_match_count

def create_excel_workbook(field_mapping_df, picklist_mapping_df):
    """Create Excel workbook with formatting."""
    print("\nCreating Excel workbook...")
    wb = Workbook()

    # Sheet 1: Field Mapping
    ws1 = wb.active
    ws1.title = "Field_Mapping"

    # Write headers
    for r_idx, row in enumerate(dataframe_to_rows(field_mapping_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)

            # Format header row
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                # Color code by match confidence
                confidence = field_mapping_df.iloc[r_idx-2]['Match_Confidence']
                if confidence == 'High':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif confidence == 'Medium':
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif confidence == 'Low':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                else:  # None
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

                cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Set column widths
    ws1.column_dimensions['A'].width = 30  # BBF_Field_API_Name
    ws1.column_dimensions['B'].width = 30  # BBF_Field_Label
    ws1.column_dimensions['C'].width = 15  # BBF_Data_Type
    ws1.column_dimensions['D'].width = 12  # BBF_Is_Required
    ws1.column_dimensions['E'].width = 30  # ES_Field_API_Name
    ws1.column_dimensions['F'].width = 30  # ES_Field_Label
    ws1.column_dimensions['G'].width = 15  # ES_Data_Type
    ws1.column_dimensions['H'].width = 15  # Match_Confidence
    ws1.column_dimensions['I'].width = 15  # Transformer_Needed
    ws1.column_dimensions['J'].width = 60  # Notes

    # Sheet 2: Picklist Mapping
    ws2 = wb.create_sheet("Picklist_Mapping")

    for r_idx, row in enumerate(dataframe_to_rows(picklist_mapping_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)

            # Format header row
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                # Color code by notes
                notes = picklist_mapping_df.iloc[r_idx-2]['Notes']
                if notes == 'Exact Match':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif notes == 'Close Match':
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:  # No Match
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Set column widths for picklist sheet
    ws2.column_dimensions['A'].width = 30  # ES_Field
    ws2.column_dimensions['B'].width = 40  # ES_Picklist_Value
    ws2.column_dimensions['C'].width = 30  # BBF_Field
    ws2.column_dimensions['D'].width = 80  # Suggested_Mapping (can be long list)
    ws2.column_dimensions['E'].width = 25  # Notes
    ws2.column_dimensions['F'].width = 30  # BBF_Final_Value

    # Save workbook
    output_file = 'day-two/mappings/ES_Account_to_BBF_Account_mapping.xlsx'
    wb.save(output_file)
    print(f"✅ Excel workbook saved: {output_file}")

    return output_file

def main():
    """Main execution."""
    print("=" * 80)
    print("Account → Account Field Mapping Generator (Updated Format)")
    print("=" * 80)

    # Create field mapping
    field_mapping_df, match_counts, transformer_count = create_field_mapping()

    # Create picklist mapping
    picklist_mapping_df, exact_count, suggested_count, no_match_count = create_picklist_mapping()

    # Create Excel workbook
    output_file = create_excel_workbook(field_mapping_df, picklist_mapping_df)

    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Output File: {output_file}")
    print(f"\nField Mapping:")
    print(f"  Total BBF Fields: {len(field_mapping_df)}")
    print(f"  High Confidence: {match_counts['High']}")
    print(f"  Medium Confidence: {match_counts['Medium']}")
    print(f"  Low Confidence: {match_counts['Low']}")
    print(f"  No Match: {match_counts['None']}")
    print(f"  Transformers Needed: {transformer_count}")
    print(f"\nPicklist Mapping:")
    print(f"  Total Values: {len(picklist_mapping_df)}")
    print(f"  Exact Matches: {exact_count}")
    print(f"  Suggested Matches: {suggested_count}")
    print(f"  No Match (Review Needed): {no_match_count}")
    print("=" * 80)

if __name__ == '__main__':
    main()
