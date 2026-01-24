#!/usr/bin/env python3
"""
Account → Account Field Mapping Generator v3
Creates Excel workbook with field mappings and picklist value mappings
Following EXACT user specifications for format and exclusions
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from difflib import SequenceMatcher
import re

# Color definitions (EXACT per specifications)
HEADER_FILL = PatternFill(start_color="FF366092", end_color="FF366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HIGH_FILL = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")
LOW_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

# Fields to EXCLUDE per user specification
EXCLUDE_FIELDS = {
    # System fields
    'Id', 'IsDeleted', 'MasterRecordId', 'CreatedDate', 'CreatedById',
    'LastModifiedDate', 'LastModifiedById', 'SystemModstamp', 'LastActivityDate',
    'LastViewedDate', 'LastReferencedDate', 'JigsawCompanyId', 'Jigsaw',
    'PhotoUrl', 'CleanStatus',
    # Day 1 fields
    'Name', 'OwnerId', 'ES_Legacy_ID__c', 'BBF_New_Id__c'
}


def similarity_score(a, b):
    """Calculate similarity between two strings"""
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def normalize_api_name(name):
    """Remove common prefixes/suffixes for comparison"""
    name = re.sub(r'__c$', '', name)
    name = re.sub(r'^[a-z0-9]+__', '', name, flags=re.IGNORECASE)
    return name.lower()


def find_best_match(bbf_field, es_fields_df):
    """Find best matching ES field for a BBF field"""
    bbf_api = bbf_field['Field API Name']
    bbf_label = bbf_field['Field Label']
    bbf_type = bbf_field['Field Type']

    best_match = None
    best_score = 0.0
    confidence = "None"

    for _, es_field in es_fields_df.iterrows():
        es_api = es_field['Field API Name']
        es_label = es_field['Field Label']
        es_type = es_field['Field Type']

        # Exact API name match
        if bbf_api == es_api:
            return {
                'es_field': es_field,
                'confidence': 'High',
                'score': 1.0
            }

        # Normalized API name match
        norm_bbf = normalize_api_name(bbf_api)
        norm_es = normalize_api_name(es_api)
        api_sim = similarity_score(norm_bbf, norm_es)

        # Label similarity
        label_sim = similarity_score(bbf_label, es_label)

        # Calculate composite score
        score = max(api_sim * 0.7, label_sim * 0.3)

        if score > best_score:
            best_score = score
            best_match = es_field

            # Determine confidence
            if api_sim >= 0.9 and bbf_type == es_type:
                confidence = "High"
            elif api_sim >= 0.7 or label_sim >= 0.8:
                confidence = "Medium"
            elif api_sim >= 0.5 or label_sim >= 0.6:
                confidence = "Low"
            else:
                confidence = "None"

    if best_score >= 0.5:
        return {
            'es_field': best_match,
            'confidence': confidence,
            'score': best_score
        }

    return None


def needs_transformer(bbf_field, es_field, bbf_picklist_values, es_picklist_values):
    """Determine if a transformer is needed"""
    if es_field is None or (isinstance(es_field, pd.Series) and es_field.empty):
        return False

    bbf_type = bbf_field['Field Type']
    es_type = es_field['Field Type']
    bbf_api = bbf_field['Field API Name']
    es_api = es_field['Field API Name']

    # Type mismatch
    if bbf_type != es_type:
        return True

    # Length mismatch (BBF shorter than ES)
    bbf_len = bbf_field.get('Length', 0)
    es_len = es_field.get('Length', 0)
    if bbf_len and es_len and es_len > bbf_len:
        return True

    # Picklist value mismatch
    if bbf_type in ['picklist', 'multipicklist']:
        bbf_values = set(bbf_picklist_values.get(bbf_api, []))
        es_values = set(es_picklist_values.get(es_api, []))

        if es_values and bbf_values:
            # Check if all ES values have exact matches in BBF
            unmatched = es_values - bbf_values
            if unmatched:
                return True

    return False


def load_picklist_values(filepath):
    """Load picklist values from CSV into dict"""
    df = pd.read_csv(filepath)
    picklist_dict = {}

    for _, row in df.iterrows():
        field = row['Field API Name']
        value = row['Picklist Value']

        if field not in picklist_dict:
            picklist_dict[field] = []
        picklist_dict[field].append(value)

    return picklist_dict


def create_field_mapping_sheet(bbf_df, es_df, bbf_picklist, es_picklist):
    """Create Field_Mapping sheet data"""
    mapping_data = []

    # Filter BBF fields
    bbf_filtered = bbf_df[~bbf_df['Field API Name'].isin(EXCLUDE_FIELDS)].copy()

    stats = {
        'high': 0,
        'medium': 0,
        'low': 0,
        'none': 0,
        'transformer': 0
    }

    for _, bbf_field in bbf_filtered.iterrows():
        bbf_api = bbf_field['Field API Name']
        bbf_label = bbf_field['Field Label']
        bbf_type = bbf_field['Field Type']
        bbf_required = "No" if bbf_field['Is Nillable'] else "Yes"

        # Find best match
        match_result = find_best_match(bbf_field, es_df)

        if match_result:
            es_field = match_result['es_field']
            confidence = match_result['confidence']

            es_api = es_field['Field API Name']
            es_label = es_field['Field Label']
            es_type = es_field['Field Type']

            # Check if transformer needed
            transformer = "Y" if needs_transformer(
                bbf_field, es_field, bbf_picklist, es_picklist
            ) else "N"

            notes = f"Mapped based on {confidence.lower()} confidence match"
            if transformer == "Y":
                notes += "; Transformation required"

            stats[confidence.lower()] += 1
            if transformer == "Y":
                stats['transformer'] += 1
        else:
            es_api = ""
            es_label = ""
            es_type = ""
            confidence = "None"
            transformer = "N"
            notes = "No obvious ES source field identified"
            stats['none'] += 1

        mapping_data.append({
            'BBF_Field_API_Name': bbf_api,
            'BBF_Field_Label': bbf_label,
            'BBF_Data_Type': bbf_type,
            'BBF_Is_Required': bbf_required,
            'ES_Field_API_Name': es_api,
            'ES_Field_Label': es_label,
            'ES_Data_Type': es_type,
            'Match_Confidence': confidence,
            'Transformer_Needed': transformer,
            'Notes': notes
        })

    return pd.DataFrame(mapping_data), stats


def create_picklist_mapping_sheet(bbf_df, es_df, bbf_picklist, es_picklist):
    """Create Picklist_Mapping sheet data following EXACT left-to-right workflow"""
    picklist_data = []

    # Filter BBF fields
    bbf_filtered = bbf_df[~bbf_df['Field API Name'].isin(EXCLUDE_FIELDS)].copy()

    # Get picklist fields from BBF
    bbf_picklist_fields = bbf_filtered[
        bbf_filtered['Field Type'].isin(['picklist', 'multipicklist'])
    ]

    for _, bbf_field in bbf_picklist_fields.iterrows():
        bbf_api = bbf_field['Field API Name']

        # Find matching ES field
        match_result = find_best_match(bbf_field, es_df)

        if not match_result:
            continue

        es_field = match_result['es_field']
        es_api = es_field['Field API Name']

        # Get picklist values
        bbf_values = bbf_picklist.get(bbf_api, [])
        es_values = es_picklist.get(es_api, [])

        if not es_values:
            continue

        # Map each ES value
        for es_value in es_values:
            # Check for exact match
            if es_value in bbf_values:
                suggested = es_value
                notes = "Exact Match"
            else:
                # Find close match
                best_match = None
                best_score = 0.0

                for bbf_value in bbf_values:
                    score = similarity_score(es_value, bbf_value)
                    if score > best_score:
                        best_score = score
                        best_match = bbf_value

                if best_score >= 0.8:
                    suggested = best_match
                    notes = "Close Match"
                else:
                    # Show ALL BBF values separated by " | "
                    suggested = " | ".join(bbf_values)
                    notes = "No Match - Select from list"

            picklist_data.append({
                'ES_Field': es_api,
                'ES_Picklist_Value': es_value,
                'BBF_Field': bbf_api,
                'Suggested_Mapping': suggested,
                'Notes': notes,
                'BBF_Final_Value': ''  # Empty user decision column
            })

    return pd.DataFrame(picklist_data)


def apply_field_mapping_formatting(ws):
    """Apply formatting to Field_Mapping sheet"""
    # Header row
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Data rows - color by confidence
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        confidence = row[7].value  # Match_Confidence column

        if confidence == "High":
            fill = HIGH_FILL
        elif confidence == "Medium":
            fill = MEDIUM_FILL
        elif confidence in ["Low", "None"]:
            fill = LOW_FILL
        else:
            fill = None

        if fill:
            for cell in row:
                cell.fill = fill

        # Alignment
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Column widths
    ws.column_dimensions['A'].width = 25  # BBF_Field_API_Name
    ws.column_dimensions['B'].width = 30  # BBF_Field_Label
    ws.column_dimensions['C'].width = 15  # BBF_Data_Type
    ws.column_dimensions['D'].width = 15  # BBF_Is_Required
    ws.column_dimensions['E'].width = 25  # ES_Field_API_Name
    ws.column_dimensions['F'].width = 30  # ES_Field_Label
    ws.column_dimensions['G'].width = 15  # ES_Data_Type
    ws.column_dimensions['H'].width = 15  # Match_Confidence
    ws.column_dimensions['I'].width = 15  # Transformer_Needed
    ws.column_dimensions['J'].width = 50  # Notes

    # Freeze top row
    ws.freeze_panes = 'A2'


def apply_picklist_mapping_formatting(ws):
    """Apply formatting to Picklist_Mapping sheet"""
    # Header row
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Data rows - color by match type
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        notes = row[4].value  # Notes column

        if notes == "Exact Match":
            fill = HIGH_FILL
        elif notes == "Close Match":
            fill = MEDIUM_FILL
        elif notes and "No Match" in notes:
            fill = LOW_FILL
        else:
            fill = None

        if fill:
            for cell in row:
                cell.fill = fill

        # Alignment
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Column widths
    ws.column_dimensions['A'].width = 30  # ES_Field
    ws.column_dimensions['B'].width = 35  # ES_Picklist_Value
    ws.column_dimensions['C'].width = 30  # BBF_Field
    ws.column_dimensions['D'].width = 60  # Suggested_Mapping (wide for " | " lists)
    ws.column_dimensions['E'].width = 35  # Notes
    ws.column_dimensions['F'].width = 25  # BBF_Final_Value

    # Freeze top row
    ws.freeze_panes = 'A2'


def main():
    print("🔄 Loading field metadata...")

    # Load ES and BBF field metadata
    es_df = pd.read_csv('day-two/exports/es_Account_fields_with_picklists.csv')
    bbf_df = pd.read_csv('day-two/exports/bbf_Account_fields_with_picklists.csv')

    print(f"   ES fields: {len(es_df)}")
    print(f"   BBF fields: {len(bbf_df)}")
    print(f"   Excluded fields: {len(EXCLUDE_FIELDS)}")

    # Load picklist values
    print("\n🔄 Loading picklist values...")
    es_picklist = load_picklist_values('day-two/exports/es_Account_picklist_values.csv')
    bbf_picklist = load_picklist_values('day-two/exports/bbf_Account_picklist_values.csv')

    print(f"   ES picklist fields: {len(es_picklist)}")
    print(f"   BBF picklist fields: {len(bbf_picklist)}")

    # Create field mapping
    print("\n🔄 Creating field mapping...")
    field_mapping_df, stats = create_field_mapping_sheet(bbf_df, es_df, bbf_picklist, es_picklist)

    print(f"   Total BBF fields mapped: {len(field_mapping_df)}")
    print(f"   High confidence: {stats['high']}")
    print(f"   Medium confidence: {stats['medium']}")
    print(f"   Low confidence: {stats['low']}")
    print(f"   No match: {stats['none']}")
    print(f"   Transformers needed: {stats['transformer']}")

    # Create picklist mapping
    print("\n🔄 Creating picklist mapping...")
    picklist_mapping_df = create_picklist_mapping_sheet(bbf_df, es_df, bbf_picklist, es_picklist)

    print(f"   Total picklist value mappings: {len(picklist_mapping_df)}")
    print(f"   Exact matches: {len(picklist_mapping_df[picklist_mapping_df['Notes'] == 'Exact Match'])}")
    print(f"   Close matches: {len(picklist_mapping_df[picklist_mapping_df['Notes'] == 'Close Match'])}")
    print(f"   No matches: {len(picklist_mapping_df[picklist_mapping_df['Notes'].str.contains('No Match', na=False)])}")

    # Create Excel workbook
    print("\n🔄 Creating Excel workbook...")
    wb = Workbook()

    # Sheet 1: Field_Mapping
    ws1 = wb.active
    ws1.title = "Field_Mapping"

    for r in dataframe_to_rows(field_mapping_df, index=False, header=True):
        ws1.append(r)

    apply_field_mapping_formatting(ws1)

    # Sheet 2: Picklist_Mapping
    ws2 = wb.create_sheet("Picklist_Mapping")

    for r in dataframe_to_rows(picklist_mapping_df, index=False, header=True):
        ws2.append(r)

    apply_picklist_mapping_formatting(ws2)

    # Save workbook
    output_path = 'day-two/mappings/ES_Account_to_BBF_Account_mapping.xlsx'
    wb.save(output_path)

    print(f"\n✅ Mapping created successfully!")
    print(f"   Output: {output_path}")
    print(f"\n📊 Summary:")
    print(f"   - BBF Fields: {len(field_mapping_df)}")
    print(f"   - ES Fields: {len(es_df)}")
    print(f"   - High Confidence: {stats['high']}")
    print(f"   - Medium Confidence: {stats['medium']}")
    print(f"   - Low/No Match: {stats['low'] + stats['none']}")
    print(f"   - Transformers Needed: {stats['transformer']}")
    print(f"   - Picklist Values Mapped: {len(picklist_mapping_df)}")


if __name__ == "__main__":
    main()
