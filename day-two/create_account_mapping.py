"""
ES Account to BBF Account Field Mapping Generator
Creates Excel workbook with field mappings and picklist value mappings
"""

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict
import re

# System fields to exclude
SYSTEM_FIELDS = {
    'Id', 'IsDeleted', 'MasterRecordId', 'CreatedDate', 'CreatedById',
    'LastModifiedDate', 'LastModifiedById', 'SystemModstamp',
    'LastActivityDate', 'LastViewedDate', 'LastReferencedDate',
    'JigsawContactId', 'Jigsaw', 'PhotoUrl', 'CleanStatus'
}

# Day 1 fields (already handled)
DAY1_FIELDS = {
    'Name', 'OwnerId', 'ES_Legacy_ID__c', 'BBF_New_Id__c'
}

def load_field_metadata(csv_path):
    """Load field metadata from CSV export"""
    fields = {}
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            api_name = row['Field API Name']
            fields[api_name] = {
                'label': row['Field Label'],
                'type': row['Field Type'],
                'is_nillable': row['Is Nillable'] == 'True',
                'custom': row['Custom'] == 'True',
                'picklist_values': row.get('Picklist Values', ''),
                'length': row.get('Length', '0')
            }
    return fields

def load_picklist_values(csv_path):
    """Load picklist values from CSV export"""
    picklists = defaultdict(lambda: {'bbf_values': [], 'es_values': []})

    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            field_api_name = row['Field API Name']
            picklist_value = row['Picklist Value']
            picklists[field_api_name]['values'] = picklists[field_api_name].get('values', [])
            picklists[field_api_name]['values'].append(picklist_value)

    return picklists

def normalize_for_matching(text):
    """Normalize field names for fuzzy matching"""
    if not text:
        return ""
    # Remove common prefixes
    text = re.sub(r'^(proposalmanager__|c2g__CODA|c2g__|fferpcore__|smartystreets__|SBQQ__|fw1__|dupcheck__|ffbf__|TASKRAY__|ActOn__|circuitvision__|DOZISF__)', '', text)
    # Remove common suffixes
    text = re.sub(r'(__c|__r)$', '', text)
    # Convert to lowercase and remove underscores
    text = text.lower().replace('_', '').replace(' ', '')
    return text

def calculate_similarity(str1, str2):
    """Calculate simple similarity score between two strings"""
    norm1 = normalize_for_matching(str1)
    norm2 = normalize_for_matching(str2)

    if norm1 == norm2:
        return 1.0

    # Check if one contains the other
    if norm1 in norm2 or norm2 in norm1:
        return 0.8

    # Simple character overlap
    set1 = set(norm1)
    set2 = set(norm2)
    if not set1 or not set2:
        return 0.0

    intersection = len(set1 & set2)
    union = len(set1 | set2)
    return intersection / union if union > 0 else 0.0

def find_best_match(bbf_field_api, bbf_field_label, es_fields):
    """Find best matching ES field for a BBF field"""
    best_match = None
    best_score = 0.0
    match_type = "None"

    for es_api, es_data in es_fields.items():
        # Exact API name match (highest priority)
        if bbf_field_api == es_api:
            return es_api, es_data, 1.0, "High"

        # API name similarity
        api_sim = calculate_similarity(bbf_field_api, es_api)
        if api_sim > best_score:
            best_score = api_sim
            best_match = (es_api, es_data)
            match_type = "High" if api_sim >= 0.8 else "Medium"

        # Label similarity
        label_sim = calculate_similarity(bbf_field_label, es_data['label'])
        if label_sim > best_score:
            best_score = label_sim
            best_match = (es_api, es_data)
            match_type = "Medium" if label_sim >= 0.7 else "Low"

    if best_score >= 0.4:  # Threshold for considering a match
        return best_match[0], best_match[1], best_score, match_type

    return None, None, 0.0, "None"

def needs_transformer(bbf_field, es_field, bbf_data, es_data):
    """Determine if a transformer is needed"""
    if not es_field:
        return False

    # Type mismatch
    if bbf_data['type'] != es_data['type']:
        return True

    # Length mismatch (ES field is longer)
    try:
        bbf_len = int(bbf_data.get('length', 0))
        es_len = int(es_data.get('length', 0))
        if es_len > bbf_len and bbf_len > 0:
            return True
    except (ValueError, TypeError):
        pass

    # Picklist with different values
    if bbf_data['type'] == 'picklist' and bbf_data.get('picklist_values') and es_data.get('picklist_values'):
        bbf_values = set(bbf_data['picklist_values'].split('; '))
        es_values = set(es_data['picklist_values'].split('; '))
        if bbf_values != es_values:
            return True

    return False

def match_picklist_value(es_value, bbf_values):
    """Find best matching BBF picklist value for ES value"""
    # Exact match
    if es_value in bbf_values:
        return es_value, "Exact"

    # Normalize and check
    norm_es = normalize_for_matching(es_value)
    for bbf_val in bbf_values:
        norm_bbf = normalize_for_matching(bbf_val)
        if norm_es == norm_bbf:
            return bbf_val, "Exact"

    # Fuzzy match
    best_match = None
    best_score = 0.0
    for bbf_val in bbf_values:
        score = calculate_similarity(es_value, bbf_val)
        if score > best_score:
            best_score = score
            best_match = bbf_val

    if best_score >= 0.6:
        return best_match, "Suggested"

    return None, "No Match"

def create_mapping_workbook(bbf_fields_csv, es_fields_csv, bbf_picklists_csv, es_picklists_csv, output_path):
    """Create Excel workbook with field and picklist mappings"""

    # Load metadata
    print("Loading BBF fields...")
    bbf_fields = load_field_metadata(bbf_fields_csv)
    print(f"Loaded {len(bbf_fields)} BBF fields")

    print("Loading ES fields...")
    es_fields = load_field_metadata(es_fields_csv)
    print(f"Loaded {len(es_fields)} ES fields")

    print("Loading BBF picklists...")
    bbf_picklists = load_picklist_values(bbf_picklists_csv)
    print(f"Loaded {len(bbf_picklists)} BBF picklist fields")

    print("Loading ES picklists...")
    es_picklists = load_picklist_values(es_picklists_csv)
    print(f"Loaded {len(es_picklists)} ES picklist fields")

    # Filter out excluded fields
    excluded_fields = SYSTEM_FIELDS | DAY1_FIELDS
    bbf_fields = {k: v for k, v in bbf_fields.items() if k not in excluded_fields}

    print(f"\nAfter exclusions: {len(bbf_fields)} BBF fields to map")

    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create Field_Mapping sheet
    ws_fields = wb.create_sheet("Field_Mapping")

    # Headers
    headers = [
        'BBF_Field_API_Name', 'BBF_Field_Label', 'BBF_Data_Type', 'BBF_Is_Required',
        'ES_Field_API_Name', 'ES_Field_Label', 'ES_Data_Type',
        'Match_Confidence', 'Transformer_Needed', 'Notes'
    ]

    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws_fields.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Color fills for confidence levels
    high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
    low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red

    # Match statistics
    stats = {'High': 0, 'Medium': 0, 'Low': 0, 'None': 0, 'Transformers': 0}

    # Process each BBF field
    row_num = 2
    for bbf_api, bbf_data in sorted(bbf_fields.items()):
        bbf_label = bbf_data['label']
        bbf_type = bbf_data['type']
        bbf_required = "Yes" if not bbf_data['is_nillable'] else "No"

        # Find matching ES field
        es_api, es_data, score, confidence = find_best_match(bbf_api, bbf_label, es_fields)

        es_label = es_data['label'] if es_data else ""
        es_type = es_data['type'] if es_data else ""

        transformer = "Y" if needs_transformer(bbf_api, es_api, bbf_data, es_data) else "N"

        # Generate notes
        notes = []
        if confidence == "High":
            notes.append("Strong match based on API name similarity")
        elif confidence == "Medium":
            notes.append("Moderate match based on label or API name")
        elif confidence == "Low":
            notes.append("Weak match - requires review")
        else:
            notes.append("No obvious ES field match found")

        if transformer == "Y":
            if es_type and bbf_type != es_type:
                notes.append(f"Type mismatch: ES={es_type}, BBF={bbf_type}")
            if bbf_type == 'picklist':
                notes.append("Picklist values may not match exactly")

        notes_text = "; ".join(notes)

        # Write row
        ws_fields.cell(row=row_num, column=1, value=bbf_api)
        ws_fields.cell(row=row_num, column=2, value=bbf_label)
        ws_fields.cell(row=row_num, column=3, value=bbf_type)
        ws_fields.cell(row=row_num, column=4, value=bbf_required)
        ws_fields.cell(row=row_num, column=5, value=es_api or "")
        ws_fields.cell(row=row_num, column=6, value=es_label)
        ws_fields.cell(row=row_num, column=7, value=es_type)
        ws_fields.cell(row=row_num, column=8, value=confidence)
        ws_fields.cell(row=row_num, column=9, value=transformer)
        ws_fields.cell(row=row_num, column=10, value=notes_text)

        # Apply color based on confidence
        if confidence == "High":
            for col in range(1, 11):
                ws_fields.cell(row=row_num, column=col).fill = high_fill
        elif confidence == "Medium":
            for col in range(1, 11):
                ws_fields.cell(row=row_num, column=col).fill = medium_fill
        elif confidence in ["Low", "None"]:
            for col in range(1, 11):
                ws_fields.cell(row=row_num, column=col).fill = low_fill

        # Update stats
        stats[confidence] += 1
        if transformer == "Y":
            stats['Transformers'] += 1

        row_num += 1

    # Adjust column widths
    ws_fields.column_dimensions['A'].width = 30
    ws_fields.column_dimensions['B'].width = 35
    ws_fields.column_dimensions['C'].width = 18
    ws_fields.column_dimensions['D'].width = 18
    ws_fields.column_dimensions['E'].width = 30
    ws_fields.column_dimensions['F'].width = 35
    ws_fields.column_dimensions['G'].width = 18
    ws_fields.column_dimensions['H'].width = 20
    ws_fields.column_dimensions['I'].width = 20
    ws_fields.column_dimensions['J'].width = 60

    # Create Picklist_Mapping sheet
    ws_picklists = wb.create_sheet("Picklist_Mapping")

    # Headers for picklist sheet
    pk_headers = [
        'BBF_Field', 'BBF_Picklist_Value', 'ES_Field', 'ES_Picklist_Value',
        'Suggested_Mapping', 'Notes'
    ]

    for col_num, header in enumerate(pk_headers, 1):
        cell = ws_picklists.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Process picklist mappings
    pk_row = 2
    pk_stats = {'Exact': 0, 'Suggested': 0, 'No Match': 0}

    for bbf_field, bbf_pk_data in sorted(bbf_picklists.items()):
        if bbf_field in excluded_fields:
            continue

        # Find corresponding ES field
        es_field = None
        for es_api, es_data in es_fields.items():
            if normalize_for_matching(bbf_field) == normalize_for_matching(es_api):
                es_field = es_api
                break

        if not es_field or es_field not in es_picklists:
            continue

        bbf_values = bbf_pk_data.get('values', [])
        es_values = es_picklists[es_field].get('values', [])

        # Map each ES value to BBF value
        for es_val in es_values:
            bbf_match, match_type = match_picklist_value(es_val, bbf_values)

            notes = []
            if match_type == "Exact":
                notes.append("Exact match found")
            elif match_type == "Suggested":
                notes.append("Fuzzy match - review recommended")
            else:
                notes.append("No suitable BBF value found - business decision needed")
                bbf_match = "[SELECT FROM: " + ", ".join(bbf_values[:5]) + ("..." if len(bbf_values) > 5 else "") + "]"

            ws_picklists.cell(row=pk_row, column=1, value=bbf_field)
            ws_picklists.cell(row=pk_row, column=2, value="")
            ws_picklists.cell(row=pk_row, column=3, value=es_field)
            ws_picklists.cell(row=pk_row, column=4, value=es_val)
            ws_picklists.cell(row=pk_row, column=5, value=bbf_match)
            ws_picklists.cell(row=pk_row, column=6, value="; ".join(notes))

            # Color code
            if match_type == "Exact":
                for col in range(1, 7):
                    ws_picklists.cell(row=pk_row, column=col).fill = high_fill
            elif match_type == "Suggested":
                for col in range(1, 7):
                    ws_picklists.cell(row=pk_row, column=col).fill = medium_fill
            else:
                for col in range(1, 7):
                    ws_picklists.cell(row=pk_row, column=col).fill = low_fill

            pk_stats[match_type] += 1
            pk_row += 1

    # Adjust picklist column widths
    ws_picklists.column_dimensions['A'].width = 30
    ws_picklists.column_dimensions['B'].width = 30
    ws_picklists.column_dimensions['C'].width = 30
    ws_picklists.column_dimensions['D'].width = 30
    ws_picklists.column_dimensions['E'].width = 30
    ws_picklists.column_dimensions['F'].width = 60

    # Save workbook
    wb.save(output_path)

    print(f"\n✅ Mapping workbook created: {output_path}")
    print(f"\nField Mapping Statistics:")
    print(f"  - High Confidence: {stats['High']}")
    print(f"  - Medium Confidence: {stats['Medium']}")
    print(f"  - Low Confidence: {stats['Low']}")
    print(f"  - No Match: {stats['None']}")
    print(f"  - Transformers Needed: {stats['Transformers']}")
    print(f"\nPicklist Mapping Statistics:")
    print(f"  - Exact Matches: {pk_stats['Exact']}")
    print(f"  - Suggested Matches: {pk_stats['Suggested']}")
    print(f"  - No Match (Need Review): {pk_stats['No Match']}")

    return stats, pk_stats

if __name__ == "__main__":
    BASE_DIR = "/mnt/c/Users/vjero/Documents/repos/python-repo/python_scripts/ES-to-BBF-migration"

    bbf_fields_csv = f"{BASE_DIR}/day-two/exports/bbf_Account_fields_with_picklists.csv"
    es_fields_csv = f"{BASE_DIR}/day-two/exports/es_Account_fields_with_picklists.csv"
    bbf_picklists_csv = f"{BASE_DIR}/day-two/exports/bbf_Account_picklist_values.csv"
    es_picklists_csv = f"{BASE_DIR}/day-two/exports/es_Account_picklist_values.csv"
    import datetime
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"{BASE_DIR}/day-two/mappings/ES_Account_to_BBF_Account_mapping_{timestamp}.xlsx"

    stats, pk_stats = create_mapping_workbook(
        bbf_fields_csv, es_fields_csv,
        bbf_picklists_csv, es_picklists_csv,
        output_path
    )

    # Try to rename to final name
    final_path = f"{BASE_DIR}/day-two/mappings/ES_Account_to_BBF_Account_mapping.xlsx"
    try:
        import os
        if os.path.exists(final_path):
            os.remove(final_path)
        os.rename(output_path, final_path)
        print(f"\n✅ Final file: {final_path}")
    except Exception as e:
        print(f"\n⚠️  Could not rename to final path (file may be open): {e}")
        print(f"✅ File saved as: {output_path}")
