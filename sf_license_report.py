"""
Salesforce License Report Generator
Connects to a Salesforce org and generates a formatted Excel report
of all User Licenses, Permission Set Licenses, and Feature Licenses.

Usage:
    python sf_license_report.py

You will be prompted for credentials (nothing is stored).
"""

import getpass
import sys
from datetime import datetime
from simple_salesforce import Salesforce
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


def connect_to_salesforce():
    print("\n=== Salesforce License Report Generator ===\n")
    username = input("Salesforce Username: ").strip()
    password = getpass.getpass("Password: ")
    security_token = getpass.getpass("Security Token: ")

    env = input("Environment (production/sandbox) [production]: ").strip().lower()
    domain = "test" if env == "sandbox" else None

    print("\nConnecting to Salesforce...")
    try:
        sf = Salesforce(
            username=username,
            password=password,
            security_token=security_token,
            domain=domain,
        )
        print(f"Connected successfully to: {sf.sf_instance}")
        return sf
    except Exception as e:
        print(f"Connection failed: {e}")
        sys.exit(1)


def query_licenses(sf):
    print("Querying license data...")

    user_licenses = sf.query_all(
        "SELECT Name, TotalLicenses, UsedLicenses, Status "
        "FROM UserLicense ORDER BY Name"
    )["records"]

    perm_set_licenses = sf.query_all(
        "SELECT MasterLabel, TotalLicenses, UsedLicenses, Status "
        "FROM PermissionSetLicense ORDER BY MasterLabel"
    )["records"]

    # Feature licenses don't always exist in every org
    try:
        feature_licenses = sf.query_all(
            "SELECT MasterLabel, TotalLicenses, UsedLicenses "
            "FROM ActiveFeatureLicenseMetric ORDER BY MasterLabel"
        )["records"]
    except Exception:
        feature_licenses = []

    print(f"  User Licenses:           {len(user_licenses)}")
    print(f"  Permission Set Licenses: {len(perm_set_licenses)}")
    print(f"  Feature Licenses:        {len(feature_licenses)}")

    return user_licenses, perm_set_licenses, feature_licenses


def build_report(user_licenses, perm_set_licenses, feature_licenses, org_instance):
    wb = Workbook()

    # --- Style definitions ---
    DARK_BLUE = "1F3864"
    MED_BLUE = "2E75B6"
    LIGHT_BLUE = "D6E4F0"
    LIGHT_GRAY = "F2F2F2"
    WHITE = "FFFFFF"
    GREEN = "548235"
    RED = "C00000"
    AMBER = "BF8F00"

    title_font = Font(name="Arial", size=18, bold=True, color=WHITE)
    subtitle_font = Font(name="Arial", size=11, color=WHITE)
    header_font = Font(name="Arial", size=11, bold=True, color=WHITE)
    data_font = Font(name="Arial", size=11, color="333333")
    data_font_bold = Font(name="Arial", size=11, bold=True, color="333333")
    section_font = Font(name="Arial", size=13, bold=True, color=DARK_BLUE)
    pct_good_font = Font(name="Arial", size=11, color=GREEN)
    pct_warn_font = Font(name="Arial", size=11, bold=True, color=AMBER)
    pct_crit_font = Font(name="Arial", size=11, bold=True, color=RED)

    title_fill = PatternFill("solid", fgColor=DARK_BLUE)
    header_fill = PatternFill("solid", fgColor=MED_BLUE)
    alt_row_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    white_fill = PatternFill("solid", fgColor=WHITE)
    light_blue_fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    thin_border = Border(bottom=Side(style="thin", color="B4C6E7"))
    header_border = Border(bottom=Side(style="medium", color=DARK_BLUE))

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ================================================================
    #  SUMMARY SHEET
    # ================================================================
    ws = wb.active
    ws.title = "License Summary"
    ws.sheet_properties.tabColor = DARK_BLUE

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 14

    # Title banner (merged across columns)
    ws.merge_cells("A1:G2")
    cell = ws["A1"]
    cell.value = "Salesforce License Report"
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 8):
        for row in range(1, 3):
            ws.cell(row=row, column=col).fill = title_fill

    # Subtitle row
    ws.merge_cells("A3:G3")
    cell = ws["A3"]
    report_date = datetime.now().strftime("%B %d, %Y at %I:%M %p")
    cell.value = f"Generated: {report_date}  |  Org: {org_instance}"
    cell.font = subtitle_font
    cell.fill = PatternFill("solid", fgColor="2B4F82")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 8):
        ws.cell(row=3, column=col).fill = PatternFill("solid", fgColor="2B4F82")

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 10
    ws.row_dimensions[3].height = 22

    row = 5

    def write_section(
        ws, start_row, section_title, headers, records, name_key, has_status=True
    ):
        r = start_row

        # Section title
        ws.merge_cells(f"B{r}:G{r}")
        cell = ws.cell(row=r, column=2, value=section_title)
        cell.font = section_font
        cell.alignment = left
        ws.row_dimensions[r].height = 28
        r += 1

        # Column headers
        for col_idx, header in enumerate(headers, start=2):
            cell = ws.cell(row=r, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = header_border
        ws.row_dimensions[r].height = 24
        r += 1

        if not records:
            ws.merge_cells(f"B{r}:G{r}")
            cell = ws.cell(row=r, column=2, value="No records found")
            cell.font = Font(name="Arial", size=11, italic=True, color="888888")
            cell.alignment = center
            return r + 2

        data_start = r
        for i, rec in enumerate(records):
            name = rec.get(name_key, "Unknown")
            total = rec.get("TotalLicenses", 0) or 0
            used = rec.get("UsedLicenses", 0) or 0
            available = total - used

            bg = alt_row_fill if i % 2 == 0 else white_fill

            # Name
            cell = ws.cell(row=r, column=2, value=name)
            cell.font = data_font
            cell.fill = bg
            cell.alignment = wrap
            cell.border = thin_border

            # Total
            cell = ws.cell(row=r, column=3, value=total)
            cell.font = data_font
            cell.fill = bg
            cell.alignment = center
            cell.number_format = "#,##0"
            cell.border = thin_border

            # Used
            cell = ws.cell(row=r, column=4, value=used)
            cell.font = data_font_bold
            cell.fill = bg
            cell.alignment = center
            cell.number_format = "#,##0"
            cell.border = thin_border

            # Available
            cell = ws.cell(row=r, column=5, value=available)
            cell.font = data_font
            cell.fill = bg
            cell.alignment = center
            cell.number_format = "#,##0"
            cell.border = thin_border

            # Utilization % (formula)
            pct_cell = ws.cell(row=r, column=6)
            pct_cell.fill = bg
            pct_cell.alignment = center
            pct_cell.border = thin_border
            if total > 0:
                pct_cell.value = f"=D{r}/C{r}"
                pct_cell.number_format = "0.0%"
                pct = used / total
                if pct >= 0.90:
                    pct_cell.font = pct_crit_font
                elif pct >= 0.75:
                    pct_cell.font = pct_warn_font
                else:
                    pct_cell.font = pct_good_font
            else:
                pct_cell.value = "N/A"
                pct_cell.font = Font(name="Arial", size=11, italic=True, color="888888")

            # Status
            if has_status:
                status = rec.get("Status", "N/A") or "N/A"
                cell = ws.cell(row=r, column=7, value=status)
                cell.font = data_font
                cell.fill = bg
                cell.alignment = center
                cell.border = thin_border
            else:
                cell = ws.cell(row=r, column=7, value="Active")
                cell.font = data_font
                cell.fill = bg
                cell.alignment = center
                cell.border = thin_border

            ws.row_dimensions[r].height = 20
            r += 1

        # Totals row
        cell = ws.cell(row=r, column=2, value="TOTAL")
        cell.font = Font(name="Arial", size=11, bold=True, color=DARK_BLUE)
        cell.fill = light_blue_fill
        cell.alignment = left
        cell.border = Border(
            top=Side(style="medium", color=DARK_BLUE),
            bottom=Side(style="medium", color=DARK_BLUE),
        )

        for col in [3, 4, 5]:
            col_letter = get_column_letter(col)
            cell = ws.cell(row=r, column=col)
            cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{r-1})"
            cell.font = Font(name="Arial", size=11, bold=True, color=DARK_BLUE)
            cell.fill = light_blue_fill
            cell.alignment = center
            cell.number_format = "#,##0"
            cell.border = Border(
                top=Side(style="medium", color=DARK_BLUE),
                bottom=Side(style="medium", color=DARK_BLUE),
            )

        # Overall utilization
        cell = ws.cell(row=r, column=6)
        cell.value = f"=IF(C{r}=0,0,D{r}/C{r})"
        cell.font = Font(name="Arial", size=11, bold=True, color=DARK_BLUE)
        cell.fill = light_blue_fill
        cell.alignment = center
        cell.number_format = "0.0%"
        cell.border = Border(
            top=Side(style="medium", color=DARK_BLUE),
            bottom=Side(style="medium", color=DARK_BLUE),
        )

        cell = ws.cell(row=r, column=7)
        cell.fill = light_blue_fill
        cell.border = Border(
            top=Side(style="medium", color=DARK_BLUE),
            bottom=Side(style="medium", color=DARK_BLUE),
        )

        return r + 3

    # Write each section
    headers_with_status = [
        "License Name",
        "Total",
        "Used",
        "Available",
        "Utilization %",
        "Status",
    ]
    headers_no_status = [
        "License Name",
        "Total",
        "Used",
        "Available",
        "Utilization %",
        "Status",
    ]

    row = write_section(
        ws,
        row,
        "User Licenses",
        headers_with_status,
        user_licenses,
        "Name",
        has_status=True,
    )
    row = write_section(
        ws,
        row,
        "Permission Set Licenses",
        headers_with_status,
        perm_set_licenses,
        "MasterLabel",
        has_status=True,
    )
    row = write_section(
        ws,
        row,
        "Feature Licenses",
        headers_no_status,
        feature_licenses,
        "MasterLabel",
        has_status=False,
    )

    # Footer note
    ws.merge_cells(f"B{row}:G{row}")
    cell = ws.cell(row=row, column=2)
    cell.value = (
        "Utilization color coding:  Green = <75%  |  Amber = 75-89%  |  Red = 90%+"
    )
    cell.font = Font(name="Arial", size=9, italic=True, color="666666")
    cell.alignment = Alignment(horizontal="center")

    # Print settings
    ws.print_title_rows = "1:3"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # Freeze panes below title
    ws.freeze_panes = "A4"

    return wb


def main():
    sf = connect_to_salesforce()
    user_lic, perm_lic, feat_lic = query_licenses(sf)

    print("\nBuilding report...")
    wb = build_report(user_lic, perm_lic, feat_lic, sf.sf_instance)

    timestamp = datetime.now().strftime("%Y-%m-%d")
    filename = f"Salesforce_License_Report_{timestamp}.xlsx"
    output_path = f"{filename}"

    wb.save(output_path)
    print(f"\nReport saved: {output_path}")
    return output_path, filename


if __name__ == "__main__":
    output_path, filename = main()
